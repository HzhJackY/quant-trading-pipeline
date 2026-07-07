import gc
import json
import math
import re
import traceback
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")


TASK_NAME = "DGTW_Benchmark_Source_Audit"
OUT_DIR = Path("output/dgtw_benchmark_source_audit_stock_matching_feasibility_v1")
RUN_DIR = Path("output/_agent_runs") / TASK_NAME
CSMAR_DIR = Path("data/csmar_exports")

BENCH_FILE = CSMAR_DIR / "STK_MKT_DGTWBENCH.xlsx"
BENCH_DES = CSMAR_DIR / "STK_MKT_DGTWBENCH[DES][xlsx].txt"
FLAG_WEIGHTS = Path("output/flag_based_top50_buffer_portfolio_construction_run_v0/flag_based_top50_buffer_research_weights_v0.parquet")
ROBUST_WEIGHTS = Path("output/robust_formation_portfolio_construction_run_v0/robust_formation_research_weights_v0.parquet")
MONTH_REF = Path("output/unified_robust_portfolio_evaluation_run_v0/unified_portfolio_monthly_net_return_by_cost.csv")
AKSHARE_SUMMARY = Path("output/akshare_csi_index_supplement_monthly_alignment_v0/akshare_csi_index_supplement_monthly_alignment_summary.json")

PORTFOLIOS = [
    "ROBUST_VQ_TOP20_EXCLUDE_SOFT_ANOMALY_EQUAL_WEIGHT",
    "ROBUST_VQ_FLAG_CLEAN_TOP50_EQUAL_WEIGHT",
    "ROBUST_VQ_FLAG_CLEAN_TOP50_BUFFER_EQUAL_WEIGHT",
    "ROBUST_VQ_D7_D9_BAND_EQUAL_WEIGHT",
    "ROBUST_VQ_TOP30_PERCENT_EQUAL_WEIGHT",
]

FIELD_ALIASES = {
    "TradingMonth": ["TradingMonth", "交易月份", "交易月", "统计月份"],
    "TradingYear": ["TradingYear", "交易年份", "统计年份", "投资组合更新年份", "组合更新年份"],
    "Symbol": ["Symbol", "证券代码", "股票代码", "stkcd", "Stkcd"],
    "ShortName": ["ShortName", "证券简称", "股票简称"],
    "MarketValue": ["MarketValue", "市值", "市值分类", "市值组合"],
    "BooktoMarket": ["BooktoMarket", "账面市值比", "账面市值比分类", "BM", "B/M"],
    "Momentum": ["Momentum", "动量", "动量分类"],
    "IsNotBSE": ["IsNotBSE", "是否剔除北交所股票", "是否剔除北交所", "剔除北交所"],
    "BenchmarkReturns": ["BenchmarkReturns", "基准收益", "基准收益率", "BenchmarkReturn"],
    "Exchange": ["Exchange", "交易所"],
    "IndustryCode": ["IndustryCode", "行业代码"],
    "INDUSTRYNAME": ["INDUSTRYNAME", "行业名称"],
}


def ensure_dirs():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)


def write_state(status, checkpoint, notes=None):
    lines = [
        "# DGTW Benchmark Source Audit & Stock Matching Feasibility v1",
        "",
        f"- task_name: {TASK_NAME}",
        f"- status: {status}",
        f"- last_checkpoint: {checkpoint}",
        f"- updated_at: {datetime.now().isoformat(timespec='seconds')}",
        "- resume_command: `python scripts\\audit_dgtw_stock_matching_feasibility_v1.py > output\\_agent_runs\\DGTW_Benchmark_Source_Audit\\run_stdout.txt 2> output\\_agent_runs\\DGTW_Benchmark_Source_Audit\\run_stderr.txt`",
        "- guardrails: no portfolio DGTW-adjusted return; no benchmark-relative return; no alpha/beta; no IR; no TE; no weights edits; no production",
    ]
    if notes:
        lines += ["", "## Notes"] + [f"- {x}" for x in notes]
    (RUN_DIR / "RUN_STATE.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def norm_text(x):
    if x is None:
        return ""
    return re.sub(r"\s+", "", str(x)).strip().lower()


def detect_field(value):
    v = norm_text(value)
    if not v:
        return None
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            a = norm_text(alias)
            if v == a or a in v:
                return field
    return None


def read_des(path):
    if not path.exists():
        return ""
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return path.read_text(encoding="gbk", errors="ignore")


def find_assignment_file():
    keywords = ["DGTWASSIGN", "DGTWASSING", "DGTW股票分配", "DGTW_ASSIGN", "STK_MKT_DGTW"]
    candidates = []
    for p in CSMAR_DIR.iterdir():
        if not p.is_file() or p.suffix.lower() != ".xlsx":
            continue
        name = p.name.upper()
        if p.name == BENCH_FILE.name:
            continue
        if any(k.upper() in name for k in keywords):
            candidates.append(p)
    best = None
    best_score = -1
    for p in candidates:
        try:
            info = inspect_excel_header(p, ["Symbol", "TradingYear", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE"])
            score = len(info["field_to_col"])
        except Exception:
            score = 0
        if score > best_score:
            best = p
            best_score = score
    return best, candidates


def inspect_excel_header(path, required_fields):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        best = None
        for ws in wb.worksheets:
            max_scan = min(ws.max_row or 0, 40)
            # Some CSMAR workbooks report ws.max_column as 1 even when rows contain
            # multiple cells. Scan a bounded width instead of trusting metadata.
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, max_col=80, values_only=True), start=1):
                detected = {}
                for col_idx, val in enumerate(row, start=1):
                    field = detect_field(val)
                    if field and field not in detected:
                        detected[field] = col_idx
                score = sum(1 for f in required_fields if f in detected)
                if best is None or score > best["score"]:
                    best = {
                        "sheet_name": ws.title,
                        "detected_header_row": row_idx,
                        "detected_data_start_row": row_idx + 1,
                        "field_to_col": detected,
                        "score": score,
                    }
        return best or {"sheet_name": None, "detected_header_row": None, "detected_data_start_row": None, "field_to_col": {}, "score": 0}
    finally:
        wb.close()


def stream_excel_required(path, required_fields, optional_fields=None):
    optional_fields = optional_fields or []
    info = inspect_excel_header(path, required_fields)
    fields = [f for f in required_fields + optional_fields if f in info["field_to_col"]]
    rows = []
    if not info["sheet_name"] or not fields:
        return pd.DataFrame(), info
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[info["sheet_name"]]
        col_map = {f: info["field_to_col"][f] for f in fields}
        max_col = max(col_map.values())
        for row in ws.iter_rows(min_row=info["detected_data_start_row"], max_col=max_col, values_only=True):
            rec = {}
            nonempty = False
            for f, c in col_map.items():
                val = row[c - 1] if c - 1 < len(row) else None
                rec[f] = val
                if val not in (None, ""):
                    nonempty = True
            if nonempty:
                rows.append(rec)
    finally:
        wb.close()
    df = pd.DataFrame(rows)
    return df, info


def parse_month(s):
    if pd.isna(s):
        return pd.NaT
    if isinstance(s, (pd.Timestamp, datetime)):
        return pd.Timestamp(s).to_period("M").to_timestamp()
    text = str(s).strip()
    if re.fullmatch(r"\d{6}", text):
        return pd.to_datetime(text + "01", format="%Y%m%d", errors="coerce")
    dt = pd.to_datetime(text, errors="coerce")
    if pd.isna(dt):
        return pd.NaT
    return dt.to_period("M").to_timestamp()


def parse_symbol(s):
    if pd.isna(s):
        return None
    text = str(s).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits.zfill(6) if digits else None


def num(s):
    return pd.to_numeric(pd.Series(s).astype(str).str.replace(",", "", regex=False).str.replace("%", "", regex=False), errors="coerce")


def load_benchmark():
    required = ["TradingMonth", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE", "BenchmarkReturns"]
    df, info = stream_excel_required(BENCH_FILE, required)
    if not df.empty:
        df["TradingMonth_dt"] = df["TradingMonth"].map(parse_month)
        df["TradingMonth"] = df["TradingMonth_dt"].dt.strftime("%Y-%m")
        for c in ["MarketValue", "BooktoMarket", "Momentum", "IsNotBSE", "BenchmarkReturns"]:
            df[c] = num(df[c]).values
        df = df.dropna(subset=["TradingMonth", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE"])
        for c in ["MarketValue", "BooktoMarket", "Momentum", "IsNotBSE"]:
            df[c] = df[c].astype(int)
    return df, info


def load_assignment(path):
    required = ["Symbol", "TradingYear", "IsNotBSE", "MarketValue", "BooktoMarket", "Momentum"]
    optional = ["ShortName", "Exchange", "IndustryCode", "INDUSTRYNAME"]
    df, info = stream_excel_required(path, required, optional)
    if not df.empty:
        df["Symbol"] = df["Symbol"].map(parse_symbol)
        for c in ["TradingYear", "IsNotBSE", "MarketValue", "BooktoMarket", "Momentum"]:
            df[c] = num(df[c]).values
        df = df.dropna(subset=["Symbol", "TradingYear", "IsNotBSE", "MarketValue", "BooktoMarket", "Momentum"])
        for c in ["TradingYear", "IsNotBSE", "MarketValue", "BooktoMarket", "Momentum"]:
            df[c] = df[c].astype(int)
    return df, info


def detect_return_unit(series):
    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return "UNKNOWN", False
    abs_s = s.abs()
    if abs_s.quantile(0.99) < 1:
        return "DECIMAL_RETURN", False
    if abs_s.median() < 30 and abs_s.quantile(0.99) > 1:
        return "PERCENT_RETURN", True
    return "UNKNOWN_REVIEW_REQUIRED", False


def audit_benchmark(df, info, unit, conv):
    required = ["TradingMonth", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE", "BenchmarkReturns"]
    missing = [f for f in required if f not in info["field_to_col"]]
    keys = ["TradingMonth", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE"]
    dup = int(df.duplicated(keys).sum()) if not df.empty and all(k in df.columns for k in keys) else None
    row = {
        "file_name": str(BENCH_FILE),
        "file_exists": BENCH_FILE.exists(),
        "sheet_name": info.get("sheet_name"),
        "row_count": int(len(df)),
        "column_count": int(len(df.columns)),
        "detected_header_row": info.get("detected_header_row"),
        "detected_data_start_row": info.get("detected_data_start_row"),
        "columns_detected": "|".join(info.get("field_to_col", {}).keys()),
        "required_fields_detected": len(missing) == 0,
        "missing_required_fields": "|".join(missing),
        "min_trading_month": df["TradingMonth"].min() if not df.empty else None,
        "max_trading_month": df["TradingMonth"].max() if not df.empty else None,
        "duplicate_key_count": dup,
        "benchmark_return_missing_count": int(df["BenchmarkReturns"].isna().sum()) if "BenchmarkReturns" in df.columns else None,
        "benchmark_return_unit_detected": unit,
        "unit_conversion_needed": conv,
        "dgtw_cell_key_unique": dup == 0 if dup is not None else False,
        "schema_status": "OK" if len(missing) == 0 and not df.empty else "SOURCE_UNUSABLE",
    }
    return pd.DataFrame([row])


def audit_assignment(df, info, path):
    required = ["Symbol", "TradingYear", "IsNotBSE", "MarketValue", "BooktoMarket", "Momentum"]
    missing = [f for f in required if f not in info["field_to_col"]]
    dup_sy = int(df.duplicated(["Symbol", "TradingYear"]).sum()) if not df.empty else None
    dup_syi = int(df.duplicated(["Symbol", "TradingYear", "IsNotBSE"]).sum()) if not df.empty else None
    row = {
        "file_name": str(path) if path else None,
        "file_exists": bool(path and path.exists()),
        "sheet_name": info.get("sheet_name"),
        "row_count": int(len(df)),
        "column_count": int(len(df.columns)),
        "detected_header_row": info.get("detected_header_row"),
        "detected_data_start_row": info.get("detected_data_start_row"),
        "columns_detected": "|".join(info.get("field_to_col", {}).keys()),
        "required_fields_detected": len(missing) == 0,
        "missing_required_fields": "|".join(missing),
        "min_trading_year": int(df["TradingYear"].min()) if not df.empty else None,
        "max_trading_year": int(df["TradingYear"].max()) if not df.empty else None,
        "duplicate_symbol_year_count": dup_sy,
        "duplicate_symbol_year_isnotbse_count": dup_syi,
        "symbol_format_example": df["Symbol"].dropna().iloc[0] if not df.empty else None,
        "assignment_schema_status": "OK" if len(missing) == 0 and not df.empty else "SOURCE_UNUSABLE",
    }
    return pd.DataFrame([row])


def duplicate_detail(df):
    rows = []
    if df.empty:
        return pd.DataFrame(columns=["key_variant", "duplicate_group_count", "duplicate_row_count", "sample_keys"])
    for variant, keys in [
        ("Symbol+TradingYear", ["Symbol", "TradingYear"]),
        ("Symbol+TradingYear+IsNotBSE", ["Symbol", "TradingYear", "IsNotBSE"]),
    ]:
        g = df.groupby(keys, dropna=False).size().reset_index(name="row_count")
        d = g[g["row_count"] > 1]
        sample = ["|".join(map(str, row)) for row in d.head(10).itertuples(index=False, name=None)]
        rows.append({
            "key_variant": variant,
            "duplicate_group_count": int(len(d)),
            "duplicate_row_count": int(d["row_count"].sum()) if len(d) else 0,
            "sample_keys": "; ".join(sample),
        })
    return pd.DataFrame(rows)


def coverage(df):
    if df.empty:
        empty = pd.DataFrame()
        summary = pd.DataFrame([{
            "month_count": 0, "min_month": None, "max_month": None, "avg_cell_count": 0,
            "min_cell_count": 0, "max_cell_count": 0, "expected_cell_count_inferred": 0,
            "complete_cell_month_ratio": 0.0, "dgtw_cell_coverage_pass": False,
        }])
        return empty, summary
    by = df.groupby("TradingMonth").agg(
        cell_count=("BenchmarkReturns", "size"),
        marketvalue_group_count=("MarketValue", "nunique"),
        booktomarket_group_count=("BooktoMarket", "nunique"),
        momentum_group_count=("Momentum", "nunique"),
        isnotbse_value_count=("IsNotBSE", "nunique"),
        missing_benchmarkreturns_count=("BenchmarkReturns", lambda x: int(pd.isna(x).sum())),
    ).reset_index()
    by["isnotbse_values"] = df.groupby("TradingMonth")["IsNotBSE"].apply(lambda x: "|".join(map(str, sorted(x.dropna().unique())))).values
    expected = int(by["cell_count"].mode().iloc[0]) if len(by) else 0
    complete_ratio = float((by["cell_count"] == expected).mean()) if expected else 0.0
    summ = pd.DataFrame([{
        "month_count": int(len(by)),
        "min_month": by["TradingMonth"].min(),
        "max_month": by["TradingMonth"].max(),
        "avg_cell_count": float(by["cell_count"].mean()),
        "min_cell_count": int(by["cell_count"].min()),
        "max_cell_count": int(by["cell_count"].max()),
        "expected_cell_count_inferred": expected,
        "complete_cell_month_ratio": complete_ratio,
        "dgtw_cell_coverage_pass": bool(complete_ratio >= 0.98 and by["missing_benchmarkreturns_count"].sum() == 0),
    }])
    return by, summ


def unit_audit(df, unit, conv):
    s = pd.to_numeric(df["BenchmarkReturns"], errors="coerce").dropna() if "BenchmarkReturns" in df else pd.Series(dtype=float)
    return pd.DataFrame([{
        "field_name": "BenchmarkReturns",
        "min_value": float(s.min()) if len(s) else np.nan,
        "p1": float(s.quantile(0.01)) if len(s) else np.nan,
        "median_value": float(s.median()) if len(s) else np.nan,
        "p99": float(s.quantile(0.99)) if len(s) else np.nan,
        "max_value": float(s.max()) if len(s) else np.nan,
        "unit_detected": unit,
        "unit_conversion_needed": conv,
        "warning": unit.startswith("UNKNOWN"),
    }])


def read_months():
    m = pd.read_csv(MONTH_REF, usecols=["month_end"])
    months = pd.to_datetime(m["month_end"], errors="coerce").dropna().drop_duplicates().sort_values()
    del m
    gc.collect()
    return months


def expected_dgtw_month(month_end):
    return (pd.Timestamp(month_end).to_period("M") + 1).strftime("%Y-%m")


def map_year(month_str, rule):
    p = pd.Period(month_str, freq="M")
    if rule == "RULE_A_CALENDAR_YEAR":
        return p.year
    if rule == "RULE_B_JULY_TO_JUNE":
        return p.year if p.month >= 7 else p.year - 1
    if rule == "RULE_C_PREVIOUS_YEAR":
        return p.year - 1
    raise ValueError(rule)


def load_weights():
    cols = ["portfolio_name", "symbol", "month_end", "weight"]
    frames = []
    for path in [FLAG_WEIGHTS, ROBUST_WEIGHTS]:
        if not path.exists():
            continue
        df = pd.read_parquet(path, columns=cols)
        df = df[df["portfolio_name"].isin(PORTFOLIOS)].copy()
        df["symbol"] = df["symbol"].map(parse_symbol)
        df["month_end"] = pd.to_datetime(df["month_end"], errors="coerce")
        df["source_weights_file"] = str(path)
        frames.append(df)
    if not frames:
        return pd.DataFrame(columns=cols + ["source_weights_file"])
    out = pd.concat(frames, ignore_index=True)
    out = out.dropna(subset=["portfolio_name", "symbol", "month_end", "weight"])
    return out


def mapping_rule_audit(months, assignment, bench, weights, des_text):
    expected_months = pd.Series([expected_dgtw_month(x) for x in months]).drop_duplicates()
    bench_months = set(bench["TradingMonth"].dropna().unique()) if not bench.empty else set()
    years_available = set(assignment["TradingYear"].dropna().astype(int).unique()) if not assignment.empty else set()
    doc_support_b = any(x in des_text for x in ["投资组合更新年份", "组合更新年份", "DGTW"])
    rows = []
    for rule in ["RULE_A_CALENDAR_YEAR", "RULE_B_JULY_TO_JUNE", "RULE_C_PREVIOUS_YEAR"]:
        mapped_years = expected_months.map(lambda x: map_year(x, rule))
        assign_cov = float(mapped_years.isin(years_available).mean()) if len(mapped_years) else 0.0
        bench_cov = float(expected_months.isin(bench_months).mean()) if len(expected_months) else 0.0
        prelim = 0.0
        if not weights.empty and not assignment.empty and not bench.empty:
            tmp = weights[["symbol", "month_end"]].drop_duplicates().copy()
            tmp["expected_dgtw_trading_month"] = tmp["month_end"].map(expected_dgtw_month)
            tmp["mapped_trading_year"] = tmp["expected_dgtw_trading_month"].map(lambda x: map_year(x, rule))
            akeys = assignment[["Symbol", "TradingYear"]].drop_duplicates().rename(columns={"Symbol": "symbol", "TradingYear": "mapped_trading_year"})
            bmonths = pd.DataFrame({"expected_dgtw_trading_month": list(bench_months)})
            tmp = tmp.merge(akeys, on=["symbol", "mapped_trading_year"], how="left", indicator="a_ind")
            tmp = tmp.merge(bmonths, on="expected_dgtw_trading_month", how="left", indicator="b_ind")
            prelim = float(((tmp["a_ind"] == "both") & (tmp["b_ind"] == "both")).mean()) if len(tmp) else 0.0
            del tmp, akeys, bmonths
        support = "DES_SUPPORTS_UPDATE_YEAR_OR_DGTW_CONVENTION" if rule == "RULE_B_JULY_TO_JUNE" and doc_support_b else ""
        rows.append({
            "candidate_rule": rule,
            "portfolio_month_count": int(len(months)),
            "expected_dgtw_trading_month_count": int(len(expected_months)),
            "assignment_coverage_ratio": assign_cov,
            "benchmark_cell_coverage_ratio": bench_cov,
            "stock_month_match_coverage_ratio_preliminary": prelim,
            "missing_assignment_month_count": int((~mapped_years.isin(years_available)).sum()),
            "missing_benchmark_month_count": int((~expected_months.isin(bench_months)).sum()),
            "documentation_support": support,
            "rule_recommendation_label": "NOT_RECOMMENDED",
        })
    df = pd.DataFrame(rows)
    viable = df[(df["assignment_coverage_ratio"] >= 0.98) & (df["benchmark_cell_coverage_ratio"] >= 0.98)]
    if viable.empty:
        df.loc[df["assignment_coverage_ratio"].idxmax(), "rule_recommendation_label"] = "MANUAL_REVIEW_REQUIRED"
    else:
        if doc_support_b and "RULE_B_JULY_TO_JUNE" in viable["candidate_rule"].values:
            rec = "RULE_B_JULY_TO_JUNE"
        else:
            rec = viable.sort_values(["stock_month_match_coverage_ratio_preliminary", "assignment_coverage_ratio"], ascending=False).iloc[0]["candidate_rule"]
        df.loc[df["candidate_rule"] == rec, "rule_recommendation_label"] = "RECOMMENDED"
        df.loc[(df["candidate_rule"] != rec) & (df.index.isin(viable.index)), "rule_recommendation_label"] = "VIABLE_BUT_SECONDARY"
    return df


def dedup_assignment(assignment):
    return assignment.sort_values(["Symbol", "TradingYear", "IsNotBSE"]).drop_duplicates(["Symbol", "TradingYear", "IsNotBSE"], keep="last")


def dedup_bench(bench):
    keys = ["TradingMonth", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE"]
    return bench.sort_values(keys).drop_duplicates(keys, keep="last")


def build_matches(weights, assignment, bench, rule, policy, unit, conv, assignment_file):
    if weights.empty:
        return pd.DataFrame()
    w = weights.copy()
    w["expected_dgtw_trading_month"] = w["month_end"].map(expected_dgtw_month)
    w["mapped_trading_year"] = w["expected_dgtw_trading_month"].map(lambda x: map_year(x, rule))
    w["is_not_bse_policy"] = policy
    a = dedup_assignment(assignment)
    a = a[a["IsNotBSE"] == policy].rename(columns={"Symbol": "symbol", "TradingYear": "mapped_trading_year"})
    keep_a = ["symbol", "mapped_trading_year", "IsNotBSE", "MarketValue", "BooktoMarket", "Momentum"]
    w = w.merge(a[keep_a], on=["symbol", "mapped_trading_year"], how="left")
    w["dgtw_assignment_match_flag"] = w["MarketValue"].notna()
    b = dedup_bench(bench)
    b = b[b["IsNotBSE"] == policy].rename(columns={"TradingMonth": "expected_dgtw_trading_month", "BenchmarkReturns": "dgtw_benchmark_return"})
    keep_b = ["expected_dgtw_trading_month", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE", "dgtw_benchmark_return"]
    out = w.merge(b[keep_b], on=["expected_dgtw_trading_month", "MarketValue", "BooktoMarket", "Momentum", "IsNotBSE"], how="left")
    out["dgtw_cell_match_flag"] = out["dgtw_benchmark_return"].notna()
    out["dgtw_benchmark_return_unit"] = unit
    out["dgtw_benchmark_return_decimal"] = out["dgtw_benchmark_return"] / 100.0 if conv else out["dgtw_benchmark_return"]
    out["source_assignment_file"] = str(assignment_file)
    out["source_benchmark_file"] = str(BENCH_FILE)
    return out


def feasibility(weights, assignment, bench, rule, unit, conv, assignment_file):
    policy_results = {}
    for policy in [1, 0]:
        m = build_matches(weights, assignment, bench, rule, policy, unit, conv, assignment_file)
        policy_results[policy] = m
    ratios = {}
    for policy, m in policy_results.items():
        ratios[policy] = float((m["dgtw_assignment_match_flag"] & m["dgtw_cell_match_flag"]).mean()) if len(m) else 0.0
    recommended_policy = 1 if ratios.get(1, 0.0) >= 0.95 or ratios.get(1, 0.0) >= ratios.get(0, 0.0) else 0
    m = policy_results[recommended_policy]
    rows = []
    for pname, g in m.groupby("portfolio_name"):
        assign_ratio = float(g["dgtw_assignment_match_flag"].mean()) if len(g) else 0.0
        cell_ratio = float(g["dgtw_cell_match_flag"].mean()) if len(g) else 0.0
        final_ratio = float((g["dgtw_assignment_match_flag"] & g["dgtw_cell_match_flag"]).mean()) if len(g) else 0.0
        if final_ratio >= 0.98:
            status = "DIRECT_MATCH_READY"
        elif final_ratio >= 0.95:
            status = "MATCH_READY_WITH_MINOR_GAPS"
        elif final_ratio >= 0.90:
            status = "WATCH_COVERAGE_GAPS"
        else:
            status = "FAIL_INSUFFICIENT_MATCHING"
        rows.append({
            "portfolio_name": pname,
            "weight_row_count": int(len(g)),
            "unique_symbol_count": int(g["symbol"].nunique()),
            "portfolio_month_count": int(g["month_end"].nunique()),
            "assignment_match_count": int(g["dgtw_assignment_match_flag"].sum()),
            "assignment_match_ratio": assign_ratio,
            "benchmark_cell_match_count": int(g["dgtw_cell_match_flag"].sum()),
            "benchmark_cell_match_ratio": cell_ratio,
            "final_dgtw_benchmark_match_count": int((g["dgtw_assignment_match_flag"] & g["dgtw_cell_match_flag"]).sum()),
            "final_dgtw_benchmark_match_ratio": final_ratio,
            "missing_assignment_count": int((~g["dgtw_assignment_match_flag"]).sum()),
            "missing_benchmark_cell_count": int((g["dgtw_assignment_match_flag"] & ~g["dgtw_cell_match_flag"]).sum()),
            "recommended_is_not_bse_policy": recommended_policy,
            "isnotbse_1_overall_final_match_ratio": ratios.get(1, 0.0),
            "isnotbse_0_overall_final_match_ratio": ratios.get(0, 0.0),
            "matching_feasibility_status": status,
        })
    return pd.DataFrame(rows), m, recommended_policy


def guardrail():
    row = {
        "portfolio_weights_modified": False,
        "portfolio_weights_reconstructed": False,
        "portfolio_dgtw_adjusted_return_calculated": False,
        "portfolio_benchmark_relative_return_calculated": False,
        "alpha_beta_regression_calculated": False,
        "information_ratio_calculated": False,
        "tracking_error_calculated": False,
        "training_run": False,
        "shap_calculated": False,
        "production_modified": False,
    }
    row["guardrail_pass"] = not any(row.values())
    return pd.DataFrame([row])


def main():
    ensure_dirs()
    write_state("running", "checking prerequisites")
    assignment_file, assignment_candidates = find_assignment_file()
    prereq = {
        "benchmark_file": str(BENCH_FILE),
        "benchmark_file_found": BENCH_FILE.exists(),
        "benchmark_des_found": BENCH_DES.exists(),
        "assignment_candidates": [str(x) for x in assignment_candidates],
        "assignment_file_selected": str(assignment_file) if assignment_file else None,
        "assignment_file_found": bool(assignment_file and assignment_file.exists()),
        "flag_weights_found": FLAG_WEIGHTS.exists(),
        "robust_weights_found": ROBUST_WEIGHTS.exists(),
        "month_reference_found": MONTH_REF.exists(),
        "akshare_summary_found": AKSHARE_SUMMARY.exists(),
    }
    prereq["prerequisites_passed"] = all([prereq["benchmark_file_found"], prereq["assignment_file_found"], prereq["month_reference_found"], prereq["flag_weights_found"], prereq["robust_weights_found"]])
    (OUT_DIR / "dgtw_source_prerequisite_check.json").write_text(json.dumps(prereq, ensure_ascii=False, indent=2), encoding="utf-8")

    write_state("running", "streaming DGTW Excel sources")
    bench, bench_info = load_benchmark() if BENCH_FILE.exists() else (pd.DataFrame(), {"field_to_col": {}})
    assignment, assign_info = load_assignment(assignment_file) if assignment_file else (pd.DataFrame(), {"field_to_col": {}})
    des_text = read_des(BENCH_DES)
    if assignment_file:
        des_text += "\n" + read_des(assignment_file.with_name(assignment_file.stem + "[DES][xlsx].txt"))
    unit, conv = detect_return_unit(bench["BenchmarkReturns"] if "BenchmarkReturns" in bench else pd.Series(dtype=float))

    bench_audit = audit_benchmark(bench, bench_info, unit, conv)
    assign_audit = audit_assignment(assignment, assign_info, assignment_file)
    bench_audit.to_csv(OUT_DIR / "dgtw_benchmark_cell_schema_audit.csv", index=False, encoding="utf-8-sig")
    assign_audit.to_csv(OUT_DIR / "dgtw_stock_assignment_schema_audit.csv", index=False, encoding="utf-8-sig")
    duplicate_detail(assignment).to_csv(OUT_DIR / "dgtw_assignment_duplicate_key_detail.csv", index=False, encoding="utf-8-sig")

    by_month, cov_summary = coverage(bench)
    by_month.to_csv(OUT_DIR / "dgtw_cell_coverage_by_month.csv", index=False, encoding="utf-8-sig")
    cov_summary.to_csv(OUT_DIR / "dgtw_cell_coverage_summary.csv", index=False, encoding="utf-8-sig")
    unit_audit(bench, unit, conv).to_csv(OUT_DIR / "dgtw_benchmark_return_unit_audit.csv", index=False, encoding="utf-8-sig")

    write_state("running", "building mapping rule audit and stock-month matching feasibility")
    months = read_months()
    weights = load_weights()
    rule_df = mapping_rule_audit(months, assignment, bench, weights, des_text)
    rule_df.to_csv(OUT_DIR / "dgtw_tradingyear_tradingmonth_mapping_rule_audit.csv", index=False, encoding="utf-8-sig")
    rec_rows = rule_df[rule_df["rule_recommendation_label"] == "RECOMMENDED"]
    recommended_rule = rec_rows["candidate_rule"].iloc[0] if len(rec_rows) else None
    manual_review = recommended_rule is None
    if recommended_rule is None:
        recommended_rule = rule_df.sort_values("stock_month_match_coverage_ratio_preliminary", ascending=False)["candidate_rule"].iloc[0]

    feas, matched, policy = feasibility(weights, assignment, bench, recommended_rule, unit, conv, assignment_file)
    feas.to_csv(OUT_DIR / "dgtw_portfolio_matching_feasibility_by_portfolio.csv", index=False, encoding="utf-8-sig")

    avg_final = float(feas["final_dgtw_benchmark_match_ratio"].mean()) if len(feas) else 0.0
    candidate_generated = avg_final >= 0.95 and not matched.empty
    candidate_cols = [
        "portfolio_name", "symbol", "month_end", "expected_dgtw_trading_month", "mapped_trading_year",
        "is_not_bse_policy", "MarketValue", "BooktoMarket", "Momentum", "dgtw_benchmark_return",
        "dgtw_benchmark_return_unit", "dgtw_benchmark_return_decimal", "dgtw_assignment_match_flag",
        "dgtw_cell_match_flag", "weight", "source_assignment_file", "source_benchmark_file",
    ]
    if candidate_generated:
        cand = matched[candidate_cols].copy()
        cand.to_parquet(OUT_DIR / "dgtw_stock_month_matched_benchmark_candidate.parquet", index=False)
        cand.head(200).to_csv(OUT_DIR / "dgtw_stock_month_matched_benchmark_candidate_sample.csv", index=False, encoding="utf-8-sig")
        candidate_rows = int(len(cand))
    else:
        cand = pd.DataFrame(columns=candidate_cols)
        cand.to_parquet(OUT_DIR / "dgtw_stock_month_matched_benchmark_candidate.parquet", index=False)
        cand.to_csv(OUT_DIR / "dgtw_stock_month_matched_benchmark_candidate_sample.csv", index=False, encoding="utf-8-sig")
        candidate_rows = 0

    policy_doc = {
        "portfolio_month_end_definition": "portfolio month_end is the holding formation/evaluation month-end from portfolio weights",
        "expected_dgtw_trading_month_rule": "expected_dgtw_trading_month = month after portfolio month_end",
        "recommended_tradingyear_mapping_rule": recommended_rule,
        "is_not_bse_policy": int(policy),
        "no_double_shift_required": True,
        "example_1": {
            "portfolio_month_end": "2017-01-31",
            "portfolio_fwd_return_month": "2017-02",
            "expected_dgtw_trading_month": "2017-02",
            "mapped_trading_year": map_year("2017-02", recommended_rule),
        },
        "example_2": {
            "portfolio_month_end": "2017-07-31",
            "portfolio_fwd_return_month": "2017-08",
            "expected_dgtw_trading_month": "2017-08",
            "mapped_trading_year": map_year("2017-08", recommended_rule),
        },
        "caveats": [
            "本任务只生成 stock-month DGTW matched benchmark candidate，不计算组合 DGTW-adjusted return。",
            "TradingYear mapping 以 coverage 和 DES/source convention evidence 审计；不使用 future return 表现选择规则。",
        ],
    }
    (OUT_DIR / "dgtw_alignment_policy.json").write_text(json.dumps(policy_doc, ensure_ascii=False, indent=2), encoding="utf-8")

    guard = guardrail()
    guard.to_csv(OUT_DIR / "dgtw_guardrail_qa.csv", index=False, encoding="utf-8-sig")

    source_ready = bool(bench_audit.loc[0, "schema_status"] == "OK" and assign_audit.loc[0, "assignment_schema_status"] == "OK")
    if not source_ready:
        recommendation = "DGTW_MATCH_FAIL_SOURCE_UNUSABLE"
    elif manual_review and avg_final >= 0.95:
        recommendation = "DGTW_MATCH_WATCH_MAPPING_RULE_MANUAL_REVIEW_REQUIRED"
    elif avg_final >= 0.95 and candidate_generated:
        recommendation = "DGTW_STOCK_MONTH_MATCH_READY_FOR_ADJUSTED_EVAL_PREP"
    elif avg_final >= 0.90:
        recommendation = "DGTW_MATCH_WATCH_COVERAGE_GAPS"
    else:
        recommendation = "DGTW_MATCH_FAIL_INSUFFICIENT_MATCHING"
    integration = pd.DataFrame([{
        "recommendation": recommendation,
        "dgtw_table_type": "month-cell return + stock-year assignment",
        "direct_stock_match_ready": bool(avg_final >= 0.95),
        "assignment_table_ready": bool(assign_audit.loc[0, "assignment_schema_status"] == "OK"),
        "benchmark_cell_table_ready": bool(bench_audit.loc[0, "schema_status"] == "OK"),
        "stock_month_candidate_generated": candidate_generated,
        "recommended_mapping_rule": recommended_rule,
        "recommended_is_not_bse_policy": int(policy),
        "missing_requirements": "" if source_ready else "required source fields or parse failed",
        "next_step_task": "DGTW-adjusted portfolio evaluation prep",
        "caveat": "不含组合层 DGTW-adjusted return；下一阶段再按权重聚合。",
    }])
    integration.to_csv(OUT_DIR / "dgtw_integration_recommendation.csv", index=False, encoding="utf-8-sig")

    final_decision = {
        "DGTW_STOCK_MONTH_MATCH_READY_FOR_ADJUSTED_EVAL_PREP": "DGTW_STOCK_MATCH_READY_FOR_ADJUSTED_EVAL_PREP",
        "DGTW_MATCH_WATCH_MAPPING_RULE_MANUAL_REVIEW_REQUIRED": "DGTW_STOCK_MATCH_WATCH_MAPPING_RULE_MANUAL_REVIEW_REQUIRED",
        "DGTW_MATCH_WATCH_COVERAGE_GAPS": "DGTW_STOCK_MATCH_WATCH_COVERAGE_GAPS",
        "DGTW_MATCH_FAIL_INSUFFICIENT_MATCHING": "DGTW_STOCK_MATCH_FAIL_INSUFFICIENT_MATCHING",
        "DGTW_MATCH_FAIL_SOURCE_UNUSABLE": "DGTW_STOCK_MATCH_FAIL_SOURCE_UNUSABLE",
    }[recommendation]
    if not bool(guard.loc[0, "guardrail_pass"]):
        final_decision = "DGTW_STOCK_MATCH_FAIL_GUARDRAIL"

    summary = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "prerequisites_passed": bool(prereq["prerequisites_passed"]),
        "dgtw_benchmark_file_found": BENCH_FILE.exists(),
        "dgtw_assignment_file_found": bool(assignment_file and assignment_file.exists()),
        "dgtw_benchmark_required_fields_detected": bool(bench_audit.loc[0, "required_fields_detected"]),
        "dgtw_assignment_required_fields_detected": bool(assign_audit.loc[0, "required_fields_detected"]),
        "dgtw_table_type": "month-cell return + stock-year assignment",
        "benchmark_row_count": int(len(bench)),
        "assignment_row_count": int(len(assignment)),
        "min_trading_month": bench_audit.loc[0, "min_trading_month"],
        "max_trading_month": bench_audit.loc[0, "max_trading_month"],
        "min_trading_year": int(assign_audit.loc[0, "min_trading_year"]) if pd.notna(assign_audit.loc[0, "min_trading_year"]) else None,
        "max_trading_year": int(assign_audit.loc[0, "max_trading_year"]) if pd.notna(assign_audit.loc[0, "max_trading_year"]) else None,
        "dgtw_cell_key_unique": bool(bench_audit.loc[0, "dgtw_cell_key_unique"]),
        "assignment_symbol_year_isnotbse_key_unique": bool(assign_audit.loc[0, "duplicate_symbol_year_isnotbse_count"] == 0),
        "benchmark_return_unit_detected": unit,
        "unit_conversion_needed": bool(conv),
        "dgtw_cell_coverage_pass": bool(cov_summary.loc[0, "dgtw_cell_coverage_pass"]),
        "recommended_tradingyear_mapping_rule": recommended_rule,
        "recommended_is_not_bse_policy": int(policy),
        "avg_assignment_match_ratio": float(feas["assignment_match_ratio"].mean()) if len(feas) else 0.0,
        "avg_benchmark_cell_match_ratio": float(feas["benchmark_cell_match_ratio"].mean()) if len(feas) else 0.0,
        "avg_final_dgtw_match_ratio": avg_final,
        "lowest_portfolio_final_match_ratio": float(feas["final_dgtw_benchmark_match_ratio"].min()) if len(feas) else 0.0,
        "stock_month_candidate_generated": candidate_generated,
        "stock_month_candidate_row_count": candidate_rows,
        "dgtw_adjusted_eval_prep_allowed": bool(candidate_generated and avg_final >= 0.95),
        "mapping_rule_manual_review_required": bool(manual_review),
        "coverage_gap_detected": bool(avg_final < 0.98),
        "portfolio_weights_modified": False,
        "portfolio_weights_reconstructed": False,
        "portfolio_dgtw_adjusted_return_calculated": False,
        "portfolio_benchmark_relative_return_calculated": False,
        "alpha_beta_regression_calculated": False,
        "information_ratio_calculated": False,
        "tracking_error_calculated": False,
        "training_run": False,
        "shap_calculated": False,
        "production_modified": False,
        "final_decision": final_decision,
        "recommended_next_step": "进入 DGTW-adjusted portfolio evaluation prep；下一阶段才允许按 weight 聚合 matched DGTW benchmark return。" if candidate_generated else "先处理 source/mapping/coverage gap，再进入 DGTW-adjusted evaluation prep。",
    }
    (OUT_DIR / "dgtw_benchmark_source_audit_stock_matching_feasibility_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    report = [
        "# DGTW Benchmark Source Audit & Stock Matching Feasibility v1",
        "",
        f"- final_decision: {final_decision}",
        f"- benchmark rows: {len(bench)}",
        f"- assignment rows: {len(assignment)}",
        f"- recommended TradingYear mapping: {recommended_rule}",
        f"- recommended IsNotBSE policy: {policy}",
        f"- avg final DGTW match ratio: {avg_final:.6f}",
        f"- stock-month candidate generated: {candidate_generated}",
        "",
        "## Guardrail",
        "",
        "- 未修改 portfolio weights。",
        "- 未计算 portfolio DGTW-adjusted return、benchmark-relative return、alpha/beta、information ratio、tracking error。",
        "- 未训练、未 SHAP、未写 production。",
    ]
    if "投资组合更新年份" in des_text or "DGTW" in des_text:
        report += ["", "## Mapping Note", "", "- DES/source text contains update-year or DGTW convention evidence, so July-to-June mapping is preferred when coverage is viable."]
    (OUT_DIR / "dgtw_benchmark_source_audit_stock_matching_feasibility_report.md").write_text("\n".join(report) + "\n", encoding="utf-8")

    final_qa = pd.DataFrame([{
        "final_decision": final_decision,
        "source_ready": source_ready,
        "candidate_generated": candidate_generated,
        "guardrail_pass": bool(guard.loc[0, "guardrail_pass"]),
        "required_outputs_created": True,
    }])
    final_qa.to_csv(OUT_DIR / "final_qa.csv", index=False, encoding="utf-8-sig")
    terminal_summary = {
        "task_name": TASK_NAME,
        "status": "completed",
        "stdout_log": str(RUN_DIR / "run_stdout.txt"),
        "stderr_log": str(RUN_DIR / "run_stderr.txt"),
        "summary_json": str(OUT_DIR / "dgtw_benchmark_source_audit_stock_matching_feasibility_summary.json"),
        "final_decision": final_decision,
    }
    (OUT_DIR / "terminal_summary.json").write_text(json.dumps(terminal_summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "task_completion_card.md").write_text(
        "\n".join([
            "# Task Completion Card",
            "",
            f"- task_name: {TASK_NAME}",
            "- status: completed",
            f"- final_decision: {final_decision}",
            f"- summary_json: {OUT_DIR / 'dgtw_benchmark_source_audit_stock_matching_feasibility_summary.json'}",
            f"- final_qa: {OUT_DIR / 'final_qa.csv'}",
        ]) + "\n",
        encoding="utf-8",
    )
    write_state("completed", "all required outputs written", [f"final_decision: {final_decision}"])
    print(json.dumps(terminal_summary, ensure_ascii=False))

    del bench, assignment, weights, matched
    gc.collect()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        ensure_dirs()
        err = traceback.format_exc()
        (RUN_DIR / "last_error.txt").write_text(err, encoding="utf-8")
        write_state("failed", "exception captured in last_error.txt")
        raise
