from __future__ import annotations

import csv
import gc
import json
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

ROOT = Path(r"C:\dev\quant")
TASK_NAME = "debug_trd_co_excel_ingestion_v0"
XLSX_PATH = ROOT / "data" / "csmar_exports" / "TRD_Co.xlsx"
OUT_DIR = ROOT / "output" / TASK_NAME
RUN_DIR = ROOT / "output" / "_agent_runs" / TASK_NAME

REQUIRED_FIELDS = [
    "Stkcd",
    "Stknme",
    "Listdt",
    "Indcd",
    "Indnme",
    "Nindcd",
    "Nindnme",
    "IndcdZX",
    "IndnmeZX",
    "Statdt",
    "Sctcd",
    "Statco",
    "FormerCode",
]

INDUSTRY_FIELDS = ["Indcd", "Indnme", "Nindcd", "Nindnme", "IndcdZX", "IndnmeZX"]


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_col(col: Any) -> str:
    text = as_text(col)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def is_security_code(value: Any) -> bool:
    text = as_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return bool(re.fullmatch(r"\d{6}", text))


def discover_effective_max_column(ws: Any, min_scan_cols: int = 80) -> int:
    max_col = max(ws.max_column or 1, min_scan_cols)
    last = 1
    for row_idx in range(1, min(ws.max_row or 1, 10) + 1):
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                last = max(last, col_idx)
    return last


def sheet_preview(ws: Any, max_cols: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(1, min(ws.max_row or 1, 5) + 1):
        row_values = [as_text(ws.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, min(max_cols, 25) + 1)]
        rows.append(
            {
                "sheet_name": ws.title,
                "row_number": row_idx,
                "values_json": json.dumps(row_values, ensure_ascii=False),
            }
        )
    return rows


def read_sheet_to_dataframe(ws: Any, effective_max_col: int) -> pd.DataFrame:
    headers = [normalize_col(ws.cell(row=1, column=col_idx).value) for col_idx in range(1, effective_max_col + 1)]
    keep_indices = [idx for idx, header in enumerate(headers, start=1) if header]
    columns = [headers[idx - 1] for idx in keep_indices]
    rows: list[list[str | None]] = []
    for row_idx in range(2, (ws.max_row or 1) + 1):
        rows.append([as_text(ws.cell(row=row_idx, column=col_idx).value) or None for col_idx in keep_indices])
    return pd.DataFrame(rows, columns=columns, dtype="string")


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    run_timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    print(f"[{run_timestamp}] start {TASK_NAME}")

    workbook_opened = False
    sheet_names: list[str] = []
    selected_sheet_name: str | None = None
    selected_sheet_max_row: int | None = None
    selected_sheet_max_column: int | None = None

    try:
        wb = load_workbook(XLSX_PATH, read_only=False, data_only=True)
        workbook_opened = True
        sheet_names = list(wb.sheetnames)
    except Exception as exc:  # noqa: BLE001
        summary = {
            "xlsx_path": str(XLSX_PATH),
            "workbook_opened": False,
            "sheet_names": [],
            "selected_sheet_name": None,
            "selected_sheet_max_row": None,
            "selected_sheet_max_column": None,
            "raw_column_count": 0,
            "cleaned_column_count": 0,
            "row_count_before_metadata_drop": 0,
            "row_count_after_metadata_drop": 0,
            "required_columns_found_count": 0,
            "required_columns_missing": REQUIRED_FIELDS,
            "industry_columns_found": [],
            "previous_stkcd_only_bug_reproduced": False,
            "ingestion_fixed": False,
            "final_decision": "TRD_CO_EXCEL_INGESTION_FAIL",
            "error": f"{type(exc).__name__}: {exc}",
        }
        write_json(OUT_DIR / "debug_trd_co_excel_ingestion_summary.json", summary)
        return 1

    sheet_profile_rows: list[dict[str, Any]] = []
    preview_rows: list[dict[str, Any]] = []
    best_sheet = None
    best_score = -1
    best_effective_max_col = 1
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        effective_max_col = discover_effective_max_column(ws)
        first_row = [normalize_col(ws.cell(row=1, column=col_idx).value) for col_idx in range(1, effective_max_col + 1)]
        found_required = [field for field in REQUIRED_FIELDS if field in first_row]
        found_industry = [field for field in INDUSTRY_FIELDS if field in first_row]
        score = len(found_required) * 10 + len(found_industry)
        sheet_profile_rows.append(
            {
                "sheet_name": sheet_name,
                "max_row": ws.max_row,
                "max_column_reported": ws.max_column,
                "effective_max_column": effective_max_col,
                "required_columns_found_count": len(found_required),
                "industry_columns_found": ";".join(found_industry),
                "first_row_values_json": json.dumps(first_row[:25], ensure_ascii=False),
            }
        )
        preview_rows.extend(sheet_preview(ws, effective_max_col))
        if score > best_score:
            best_score = score
            best_sheet = ws
            best_effective_max_col = effective_max_col

    write_csv(
        OUT_DIR / "trd_co_workbook_sheet_profile.csv",
        sheet_profile_rows,
        [
            "sheet_name",
            "max_row",
            "max_column_reported",
            "effective_max_column",
            "required_columns_found_count",
            "industry_columns_found",
            "first_row_values_json",
        ],
    )
    write_csv(OUT_DIR / "trd_co_workbook_first5x25_preview.csv", preview_rows, ["sheet_name", "row_number", "values_json"])

    if best_sheet is None:
        wb.close()
        raise RuntimeError("No sheet found.")

    selected_sheet_name = best_sheet.title
    selected_sheet_max_row = best_sheet.max_row
    selected_sheet_max_column = best_effective_max_col

    raw_df = read_sheet_to_dataframe(best_sheet, best_effective_max_col)
    wb.close()

    raw_columns = [str(col) for col in raw_df.columns]
    previous_stkcd_only_bug_reproduced = raw_columns == ["Stkcd"]
    row_count_before = int(len(raw_df))

    cleaned_df = raw_df.copy()
    if "Stkcd" in cleaned_df.columns:
        metadata_mask = ~cleaned_df["Stkcd"].map(is_security_code)
        cleaned_df = cleaned_df.loc[~metadata_mask].copy()
        cleaned_df["Stkcd"] = cleaned_df["Stkcd"].map(lambda x: as_text(x).zfill(6) if is_security_code(x) else as_text(x))
    cleaned_df = cleaned_df.dropna(how="all")
    cleaned_columns = [str(col) for col in cleaned_df.columns]
    row_count_after = int(len(cleaned_df))

    check_rows = []
    found_required: list[str] = []
    for field in REQUIRED_FIELDS:
        matched = field if field in cleaned_df.columns else ""
        found = bool(matched)
        if found:
            found_required.append(field)
            series = cleaned_df[matched]
            sample_values = ";".join([as_text(v) for v in series.dropna().astype(str).head(5).tolist()])
            non_null_count = int(series.notna().sum())
        else:
            sample_values = ""
            non_null_count = 0
        check_rows.append(
            {
                "field_name": field,
                "found": found,
                "matched_column_name": matched,
                "non_null_count": non_null_count,
                "sample_values": sample_values,
            }
        )
    write_csv(
        OUT_DIR / "trd_co_required_column_check.csv",
        check_rows,
        ["field_name", "found", "matched_column_name", "non_null_count", "sample_values"],
    )
    cleaned_df.head(20).to_csv(OUT_DIR / "trd_co_cleaned_preview.csv", index=False, encoding="utf-8-sig")

    industry_columns_found = [field for field in INDUSTRY_FIELDS if field in cleaned_df.columns]
    required_missing = [field for field in REQUIRED_FIELDS if field not in cleaned_df.columns]
    raw_column_count = len(raw_columns)
    cleaned_column_count = len(cleaned_columns)
    ingestion_fixed = raw_column_count > 1 and bool(industry_columns_found) and row_count_after > 0

    if ingestion_fixed:
        final_decision = "TRD_CO_EXCEL_INGESTION_FIXED"
    elif len(sheet_names) > 1:
        final_decision = "TRD_CO_EXCEL_INGESTION_WATCH_SHEET_AMBIGUOUS"
    else:
        final_decision = "TRD_CO_EXCEL_INGESTION_FAIL"

    summary = {
        "xlsx_path": str(XLSX_PATH),
        "workbook_opened": workbook_opened,
        "sheet_names": sheet_names,
        "selected_sheet_name": selected_sheet_name,
        "selected_sheet_max_row": selected_sheet_max_row,
        "selected_sheet_max_column": selected_sheet_max_column,
        "raw_columns": raw_columns,
        "cleaned_columns": cleaned_columns,
        "raw_column_count": raw_column_count,
        "cleaned_column_count": cleaned_column_count,
        "row_count_before_metadata_drop": row_count_before,
        "row_count_after_metadata_drop": row_count_after,
        "required_columns_found_count": len(found_required),
        "required_columns_missing": required_missing,
        "industry_columns_found": industry_columns_found,
        "previous_stkcd_only_bug_reproduced": previous_stkcd_only_bug_reproduced,
        "ingestion_fixed": ingestion_fixed,
        "final_decision": final_decision,
    }
    write_json(OUT_DIR / "debug_trd_co_excel_ingestion_summary.json", summary)

    report = f"""# Debug TRD_Co Excel Ingestion v0

## 结论

`{final_decision}`

## Workbook

- xlsx_path: `{XLSX_PATH}`
- sheet_names: `{", ".join(sheet_names)}`
- selected_sheet_name: `{selected_sheet_name}`
- selected_sheet_max_row: `{selected_sheet_max_row}`
- selected_sheet_max_column: `{selected_sheet_max_column}`

## 修复点

本次读取不再使用 DES 白名单或 `usecols` 限制；按第 1 行英文字段名读取，并删除 `Stkcd` 不是 6 位证券代码的元数据行，例如中文字段名行和单位/说明行。

## 结果

- raw_column_count: `{raw_column_count}`
- cleaned_column_count: `{cleaned_column_count}`
- row_count_before_metadata_drop: `{row_count_before}`
- row_count_after_metadata_drop: `{row_count_after}`
- required_columns_found_count: `{len(found_required)}`
- required_columns_missing: `{", ".join(required_missing)}`
- industry_columns_found: `{", ".join(industry_columns_found)}`

本任务没有做行业中性化、score、join、IC、收益、回测、训练、SHAP 或 production。
"""
    (OUT_DIR / "debug_trd_co_excel_ingestion_report.md").write_text(report, encoding="utf-8")

    terminal_summary = {
        "task_name": TASK_NAME,
        "run_timestamp": run_timestamp,
        "stdout_log": rel(RUN_DIR / "run_stdout.txt"),
        "stderr_log": rel(RUN_DIR / "run_stderr.txt"),
        "output_directory": rel(OUT_DIR),
        "final_decision": final_decision,
        "exit_code": 0,
    }
    write_json(OUT_DIR / "terminal_summary.json", terminal_summary)
    write_csv(
        OUT_DIR / "final_qa.csv",
        [
            {"check": "workbook_opened", "passed": workbook_opened, "notes": ""},
            {"check": "multi_column_read", "passed": raw_column_count > 1, "notes": f"raw_column_count={raw_column_count}"},
            {"check": "industry_columns_found", "passed": bool(industry_columns_found), "notes": ";".join(industry_columns_found)},
            {"check": "metadata_rows_dropped", "passed": row_count_after < row_count_before, "notes": f"{row_count_before}->{row_count_after}"},
            {"check": "no_forbidden_work", "passed": True, "notes": "No score/join/IC/return/backtest/training/production."},
        ],
        ["check", "passed", "notes"],
    )
    (OUT_DIR / "task_completion_card.md").write_text(
        f"""# Task Completion Card

- task_name: `{TASK_NAME}`
- final_decision: `{final_decision}`
- output_directory: `{rel(OUT_DIR)}`
- ingestion_fixed: `{ingestion_fixed}`
- no_forbidden_work: `True`
""",
        encoding="utf-8",
    )

    (RUN_DIR / "RUN_STATE.md").write_text(
        f"""# RUN_STATE

任务：{TASK_NAME}
状态：完成

已读取：
- {XLSX_PATH}

输出目录：
- {OUT_DIR}

final_decision: {final_decision}
ingestion_fixed: {ingestion_fixed}

禁止项确认：
- 未做行业中性化
- 未生成 neutral score
- 未 join score panel
- 未计算 IC / 收益 / 回测
- 未训练 / SHAP
- 未写 production
""",
        encoding="utf-8",
    )

    del raw_df, cleaned_df
    gc.collect()
    print(f"final_decision={final_decision}")
    print(f"raw_column_count={raw_column_count}")
    print(f"row_count_after_metadata_drop={row_count_after}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
