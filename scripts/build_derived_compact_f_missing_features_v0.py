from __future__ import annotations

import gc
import json
import math
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / "derived_compact_f_missing_features_v0"

V3_PANEL = ROOT / "output" / "csmar_pit_clean_core_financial_factors_v3" / "pit_clean_core_financial_factors_monthly_v3.parquet"
PATCH_SUMMARY = ROOT / "output" / "existing_fs_missing_feature_constructibility_patch_v0" / "existing_fs_missing_feature_constructibility_patch_summary.json"
FIELD_AVAILABILITY = ROOT / "output" / "existing_fs_missing_feature_constructibility_patch_v0" / "fs_field_availability_check.csv"
RECLASSIFIED = ROOT / "output" / "existing_fs_missing_feature_constructibility_patch_v0" / "compact_f_missing_feature_reclassified.csv"
FORMULAS = ROOT / "output" / "existing_fs_missing_feature_constructibility_patch_v0" / "derived_feature_formula_candidates.csv"
COMINS = ROOT / "data" / "csmar_exports" / "FS_Comins.xlsx"
COMBAS = ROOT / "data" / "csmar_exports" / "FS_Combas.xlsx"
COMSCFD = ROOT / "data" / "csmar_exports" / "FS_Comscfd.xlsx"

REQUIRED_INPUTS = [V3_PANEL, PATCH_SUMMARY, FIELD_AVAILABILITY, RECLASSIFIED, FORMULAS]
V3_REQUIRED = [
    "symbol",
    "month_end",
    "selected_report_period",
    "selected_pit_date",
    "ttm_complete_flag",
    "ttm_quarters_available",
    "uses_pre_2017_buffer_flag",
    "factor_validity_flags",
    "net_profit_parent_ttm",
    "net_profit_ttm",
    "revenue_ttm",
    "total_assets",
    "total_liabilities",
    "equity_parent",
    "total_equity",
]

INCOME_COLS = [
    "Stkcd",
    "Accper",
    "Typrep",
    "DeclareDate",
    "B001101000",
    "B001100000",
    "B001300000",
    "B002000000",
    "B002000101",
    "B003000000",
    "B004000000",
]
BALANCE_COLS = [
    "Stkcd",
    "Accper",
    "Typrep",
    "DeclareDate",
    "A001100000",
    "A001123000",
    "A002100000",
    "A001000000",
    "A003100000",
    "A003000000",
]
CASHFLOW_COLS = ["Stkcd", "Accper", "Typrep", "DeclareDate", "C001000000"]
EPS = 1e-12


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def write_json(path: Path, payload: dict[str, Any] | list[Any]) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=json_default), encoding="utf-8")


def json_default(value: Any) -> Any:
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        if math.isnan(float(value)):
            return None
        return float(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    return str(value)


def missing_input_report(missing: list[Path]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Missing Input Report",
        "",
        "Derived Compact-F missing feature panel was not built because required inputs are missing.",
        "",
        "## Missing files",
        "",
    ]
    lines.extend(f"- {p.as_posix()}" for p in missing)
    (OUT_DIR / "missing_input_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    write_json(
        OUT_DIR / "derived_compact_f_missing_features_summary.json",
        {"run_timestamp": now_iso(), "final_decision": "DERIVED_COMPACT_F_MISSING_FEATURES_FAIL_BLOCK_INTEGRATION", "missing_inputs": [str(p) for p in missing]},
    )


def symbol_norm(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    if "." in text:
        text = text.split(".", 1)[0]
    return text.zfill(6)


def date_norm(value: Any) -> pd.Timestamp | pd.NaT:
    return pd.to_datetime(value, errors="coerce")


def period_key(value: Any) -> str | None:
    dt = date_norm(value)
    if pd.isna(dt):
        return None
    return dt.strftime("%Y-%m-%d")


def lag4_period(value: Any) -> str | None:
    dt = date_norm(value)
    if pd.isna(dt):
        return None
    return (dt - pd.DateOffset(years=1)).strftime("%Y-%m-%d")


def numeric(value: Any) -> float:
    if value is None:
        return np.nan
    return pd.to_numeric(value, errors="coerce")


def read_excel_filtered(path: Path, wanted_cols: list[str], symbols: set[str], periods: set[str]) -> tuple[pd.DataFrame, list[str]]:
    if not path.exists():
        return pd.DataFrame(columns=wanted_cols), wanted_cols
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # CSMAR workbooks can declare an incorrect worksheet dimension (for example A1:A...).
    # Reset dimensions so read-only iteration sees all columns while still streaming rows.
    ws.reset_dimensions()
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(name): idx for idx, name in enumerate(header) if name is not None}
    present_cols = [col for col in wanted_cols if col in header_map]
    missing_cols = [col for col in wanted_cols if col not in header_map]
    required_for_filter = [col for col in ["Stkcd", "Accper"] if col in header_map]
    read_cols = sorted(set(present_cols + required_for_filter), key=lambda c: header_map[c])
    indices = [header_map[col] for col in read_cols]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        stk = row[header_map["Stkcd"]] if "Stkcd" in header_map else None
        acc = row[header_map["Accper"]] if "Accper" in header_map else None
        sym = symbol_norm(stk)
        per = period_key(acc)
        if sym not in symbols or per not in periods:
            continue
        rows.append({col: row[idx] for col, idx in zip(read_cols, indices)})
    wb.close()
    df = pd.DataFrame(rows)
    for col in wanted_cols:
        if col not in df.columns:
            df[col] = np.nan
    return df[wanted_cols], missing_cols


def prepare_source(df: pd.DataFrame, prefix: str, value_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        cols = ["symbol", "report_period", f"{prefix}_typrep", f"{prefix}_declare_date"] + value_cols
        return pd.DataFrame(columns=cols)
    out = df.copy()
    out["symbol"] = out["Stkcd"].apply(symbol_norm)
    out["report_period"] = out["Accper"].apply(period_key)
    out[f"{prefix}_declare_date"] = pd.to_datetime(out["DeclareDate"], errors="coerce") if "DeclareDate" in out else pd.NaT
    out[f"{prefix}_typrep"] = out["Typrep"].astype(str) if "Typrep" in out else ""
    for col in value_cols:
        out[col] = pd.to_numeric(out[col], errors="coerce") if col in out else np.nan
    out["_typrep_priority"] = np.where(out[f"{prefix}_typrep"].str.upper().eq("A"), 0, 1)
    out = out.sort_values(["symbol", "report_period", "_typrep_priority", f"{prefix}_declare_date"])
    out = out.drop_duplicates(["symbol", "report_period"], keep="first")
    keep = ["symbol", "report_period", f"{prefix}_typrep", f"{prefix}_declare_date"] + value_cols
    return out[keep]


def denominator_invalid(series: pd.Series, negative_invalid: bool = True) -> pd.Series:
    numeric_series = pd.to_numeric(series, errors="coerce")
    invalid = numeric_series.isna() | (numeric_series.abs() <= EPS)
    if negative_invalid:
        invalid = invalid | (numeric_series < 0)
    return invalid


def ratio(numer: pd.Series, denom: pd.Series) -> pd.Series:
    n = pd.to_numeric(numer, errors="coerce")
    d = pd.to_numeric(denom, errors="coerce")
    out = n / d.replace(0, np.nan)
    return out.replace([np.inf, -np.inf], np.nan)


def add_coverage(rows: list[dict[str, Any]], panel: pd.DataFrame, feature: str) -> None:
    s = panel[feature]
    rows.append(
        {
            "feature_name": feature,
            "non_null_count": int(s.notna().sum()),
            "missing_count": int(s.isna().sum()),
            "coverage": float(s.notna().mean()) if len(s) else 0.0,
        }
    )


def quantile_row(panel: pd.DataFrame, feature: str) -> dict[str, Any]:
    s = pd.to_numeric(panel[feature], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    qs = s.quantile([0.001, 0.01, 0.05, 0.5, 0.95, 0.99, 0.999]) if len(s) else pd.Series(dtype=float)
    return {
        "feature_name": feature,
        "non_null_count": int(len(s)),
        "min": float(s.min()) if len(s) else None,
        "p001": float(qs.get(0.001)) if len(s) else None,
        "p01": float(qs.get(0.01)) if len(s) else None,
        "p05": float(qs.get(0.05)) if len(s) else None,
        "p50": float(qs.get(0.5)) if len(s) else None,
        "p95": float(qs.get(0.95)) if len(s) else None,
        "p99": float(qs.get(0.99)) if len(s) else None,
        "p999": float(qs.get(0.999)) if len(s) else None,
        "max": float(s.max()) if len(s) else None,
    }


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    missing = [p for p in REQUIRED_INPUTS if not p.exists()]
    source_missing = [p for p in [COMINS, COMBAS, COMSCFD] if not p.exists()]
    if missing or source_missing:
        missing_input_report(missing + source_missing)
        print(f"Missing required inputs: {len(missing) + len(source_missing)}")
        return 0

    patch_summary = json.loads(PATCH_SUMMARY.read_text(encoding="utf-8"))
    availability = pd.read_csv(FIELD_AVAILABILITY)
    source_files_used = [str(COMINS.relative_to(ROOT)), str(COMBAS.relative_to(ROOT)), str(COMSCFD.relative_to(ROOT))]

    v3_columns = pq.ParquetFile(V3_PANEL).schema.names
    v3_read_cols = [col for col in V3_REQUIRED if col in v3_columns]
    v3_missing_cols = [col for col in V3_REQUIRED if col not in v3_columns]
    base = pd.read_parquet(V3_PANEL, columns=v3_read_cols)
    base["symbol"] = base["symbol"].apply(symbol_norm)
    base["month_end"] = pd.to_datetime(base["month_end"], errors="coerce")
    base["selected_report_period"] = pd.to_datetime(base["selected_report_period"], errors="coerce")
    base["selected_pit_date"] = pd.to_datetime(base["selected_pit_date"], errors="coerce")
    base["report_period_key"] = base["selected_report_period"].apply(period_key)
    base["lag4_report_period_key"] = base["selected_report_period"].apply(lag4_period)

    symbols = set(base["symbol"].dropna().astype(str))
    periods = set(base["report_period_key"].dropna().astype(str))
    eps_periods = periods | set(base["lag4_report_period_key"].dropna().astype(str))

    income_raw, income_missing = read_excel_filtered(COMINS, INCOME_COLS, symbols, eps_periods)
    balance_raw, balance_missing = read_excel_filtered(COMBAS, BALANCE_COLS, symbols, periods)
    cash_raw, cash_missing = read_excel_filtered(COMSCFD, CASHFLOW_COLS, symbols, periods)

    income = prepare_source(income_raw, "income", ["B001101000", "B001100000", "B001300000", "B002000000", "B002000101", "B003000000", "B004000000"])
    balance = prepare_source(balance_raw, "balance", ["A001100000", "A001123000", "A002100000", "A001000000", "A003100000", "A003000000"])
    cash = prepare_source(cash_raw, "cashflow", ["C001000000"])

    panel = base.merge(income, how="left", left_on=["symbol", "report_period_key"], right_on=["symbol", "report_period"])
    panel = panel.drop(columns=["report_period"], errors="ignore")
    panel = panel.merge(balance, how="left", left_on=["symbol", "report_period_key"], right_on=["symbol", "report_period"])
    panel = panel.drop(columns=["report_period"], errors="ignore")
    panel = panel.merge(cash, how="left", left_on=["symbol", "report_period_key"], right_on=["symbol", "report_period"])
    panel = panel.drop(columns=["report_period"], errors="ignore")

    eps_lag = income[["symbol", "report_period", "B003000000", "B004000000"]].rename(
        columns={"report_period": "lag4_report_period_key", "B003000000": "basic_eps_lag4", "B004000000": "diluted_eps_lag4"}
    )
    panel = panel.merge(eps_lag, how="left", on=["symbol", "lag4_report_period_key"])

    rename_map = {
        "A001100000": "current_assets",
        "A001123000": "inventories",
        "A002100000": "current_liabilities",
        "A001000000": "total_assets_component",
        "A003100000": "equity_parent_component",
        "A003000000": "total_equity_component",
        "B001300000": "operating_profit",
        "B001101000": "operating_revenue",
        "B001100000": "total_revenue",
        "B002000101": "net_profit_parent",
        "B002000000": "net_profit",
        "B003000000": "basic_eps",
        "B004000000": "diluted_eps",
        "C001000000": "operating_cash_flow",
    }
    panel = panel.rename(columns=rename_map)
    for col in rename_map.values():
        if col not in panel.columns:
            panel[col] = np.nan

    panel["current_ratio_denominator_invalid"] = denominator_invalid(panel["current_liabilities"])
    panel["quick_ratio_denominator_invalid"] = denominator_invalid(panel["current_liabilities"])
    panel["eps_yoy_denominator_invalid"] = denominator_invalid(panel["basic_eps_lag4"], negative_invalid=True)
    panel["equity_multiplier_parent_denominator_invalid"] = denominator_invalid(panel["equity_parent_component"])
    panel["equity_multiplier_total_denominator_invalid"] = denominator_invalid(panel["total_equity_component"])
    panel["operating_margin_denominator_invalid"] = denominator_invalid(panel["operating_revenue"])
    panel["cfo_to_earnings_parent_denominator_invalid"] = denominator_invalid(panel["net_profit_parent"], negative_invalid=True)
    panel["cfo_to_earnings_total_denominator_invalid"] = denominator_invalid(panel["net_profit"], negative_invalid=True)

    panel["current_ratio_raw"] = ratio(panel["current_assets"], panel["current_liabilities"])
    panel["quick_ratio_raw"] = ratio(panel["current_assets"] - panel["inventories"], panel["current_liabilities"])
    panel["eps_yoy_raw"] = ratio(panel["basic_eps"], panel["basic_eps_lag4"]) - 1
    panel["diluted_eps_yoy_raw"] = ratio(panel["diluted_eps"], panel["diluted_eps_lag4"]) - 1
    panel["equity_multiplier_parent_raw"] = ratio(panel["total_assets_component"], panel["equity_parent_component"])
    panel["equity_multiplier_total_raw"] = ratio(panel["total_assets_component"], panel["total_equity_component"])
    panel["operating_margin_raw"] = ratio(panel["operating_profit"], panel["operating_revenue"])
    panel["operating_margin_total_revenue_raw"] = ratio(panel["operating_profit"], panel["total_revenue"])
    panel["cfo_to_earnings_parent_raw"] = ratio(panel["operating_cash_flow"], panel["net_profit_parent"])
    panel["cfo_to_earnings_total_raw"] = ratio(panel["operating_cash_flow"], panel["net_profit"])

    publish_cols = ["income_declare_date", "balance_declare_date", "cashflow_declare_date"]
    typrep_cols = ["income_typrep", "balance_typrep", "cashflow_typrep"]
    for col in publish_cols:
        if col not in panel.columns:
            panel[col] = pd.NaT
    for col in typrep_cols:
        if col not in panel.columns:
            panel[col] = ""
    panel["source_lacks_publish_date_flag"] = panel[publish_cols].isna().any(axis=1)
    panel["typrep_not_A_flag"] = ~panel[typrep_cols].fillna("").apply(lambda row: all(str(x).upper() == "A" for x in row if str(x) != ""), axis=1)
    source_publish_violation = pd.Series(False, index=panel.index)
    for col in publish_cols:
        source_publish_violation = source_publish_violation | (pd.to_datetime(panel[col], errors="coerce") > panel["month_end"])
    panel["source_publish_date_after_month_end_flag"] = source_publish_violation

    component_flag_map = {
        "current_ratio_component_missing_flag": ["current_assets", "current_liabilities"],
        "quick_ratio_component_missing_flag": ["current_assets", "inventories", "current_liabilities"],
        "eps_yoy_component_missing_flag": ["basic_eps", "basic_eps_lag4"],
        "equity_multiplier_component_missing_flag": ["total_assets_component", "equity_parent_component"],
        "operating_margin_component_missing_flag": ["operating_profit", "operating_revenue"],
        "cfo_to_earnings_component_missing_flag": ["operating_cash_flow", "net_profit_parent"],
    }
    for flag, cols in component_flag_map.items():
        panel[flag] = panel[cols].isna().any(axis=1)

    panel["source_panel_version"] = "csmar_pit_clean_core_financial_factors_v3"
    panel["derived_feature_version"] = "derived_compact_f_missing_features_v0"

    output_cols = [
        "symbol",
        "month_end",
        "selected_report_period",
        "selected_pit_date",
        "source_panel_version",
        "derived_feature_version",
        "current_ratio_raw",
        "quick_ratio_raw",
        "eps_yoy_raw",
        "diluted_eps_yoy_raw",
        "equity_multiplier_parent_raw",
        "equity_multiplier_total_raw",
        "operating_margin_raw",
        "operating_margin_total_revenue_raw",
        "cfo_to_earnings_parent_raw",
        "cfo_to_earnings_total_raw",
        "current_ratio_denominator_invalid",
        "quick_ratio_denominator_invalid",
        "eps_yoy_denominator_invalid",
        "equity_multiplier_parent_denominator_invalid",
        "equity_multiplier_total_denominator_invalid",
        "operating_margin_denominator_invalid",
        "cfo_to_earnings_parent_denominator_invalid",
        "cfo_to_earnings_total_denominator_invalid",
        "source_lacks_publish_date_flag",
        "typrep_not_A_flag",
        "source_publish_date_after_month_end_flag",
        "current_ratio_component_missing_flag",
        "quick_ratio_component_missing_flag",
        "eps_yoy_component_missing_flag",
        "equity_multiplier_component_missing_flag",
        "operating_margin_component_missing_flag",
        "cfo_to_earnings_component_missing_flag",
        "current_assets",
        "inventories",
        "current_liabilities",
        "total_assets_component",
        "equity_parent_component",
        "total_equity_component",
        "operating_profit",
        "operating_revenue",
        "total_revenue",
        "net_profit_parent",
        "net_profit",
        "operating_cash_flow",
        "basic_eps",
        "diluted_eps",
        "basic_eps_lag4",
        "diluted_eps_lag4",
        "income_declare_date",
        "balance_declare_date",
        "cashflow_declare_date",
        "income_typrep",
        "balance_typrep",
        "cashflow_typrep",
    ]
    out_panel = panel[output_cols].copy()
    out_panel.to_parquet(OUT_DIR / "derived_compact_f_missing_features_v0.parquet", index=False)

    features = [
        "current_ratio_raw",
        "quick_ratio_raw",
        "eps_yoy_raw",
        "diluted_eps_yoy_raw",
        "equity_multiplier_parent_raw",
        "equity_multiplier_total_raw",
        "operating_margin_raw",
        "operating_margin_total_revenue_raw",
        "cfo_to_earnings_parent_raw",
        "cfo_to_earnings_total_raw",
    ]
    coverage_rows: list[dict[str, Any]] = []
    quantile_rows: list[dict[str, Any]] = []
    for feature in features:
        add_coverage(coverage_rows, out_panel, feature)
        quantile_rows.append(quantile_row(out_panel, feature))
    pd.DataFrame(coverage_rows).to_csv(OUT_DIR / "derived_feature_coverage.csv", index=False)
    pd.DataFrame(quantile_rows).to_csv(OUT_DIR / "derived_feature_extreme_quantiles.csv", index=False)

    invalid_flags = [
        "current_ratio_denominator_invalid",
        "quick_ratio_denominator_invalid",
        "eps_yoy_denominator_invalid",
        "equity_multiplier_parent_denominator_invalid",
        "equity_multiplier_total_denominator_invalid",
        "operating_margin_denominator_invalid",
        "cfo_to_earnings_parent_denominator_invalid",
        "cfo_to_earnings_total_denominator_invalid",
        "source_lacks_publish_date_flag",
        "typrep_not_A_flag",
        "source_publish_date_after_month_end_flag",
        "current_ratio_component_missing_flag",
        "quick_ratio_component_missing_flag",
        "eps_yoy_component_missing_flag",
        "equity_multiplier_component_missing_flag",
        "operating_margin_component_missing_flag",
        "cfo_to_earnings_component_missing_flag",
    ]
    invalid_rows = [{"flag_name": flag, "true_count": int(out_panel[flag].fillna(False).sum()), "true_rate": float(out_panel[flag].fillna(False).mean())} for flag in invalid_flags]
    pd.DataFrame(invalid_rows).to_csv(OUT_DIR / "derived_feature_invalid_flags.csv", index=False)

    rows = int(len(out_panel))
    symbols_count = int(out_panel["symbol"].nunique())
    months_count = int(out_panel["month_end"].nunique())
    duplicate_count = int(out_panel.duplicated(["symbol", "month_end"]).sum())
    one_row = duplicate_count == 0
    selected_pit_violation = int((out_panel["selected_pit_date"] > out_panel["month_end"]).fillna(False).sum())
    future_report_violation = int((out_panel["selected_report_period"] > out_panel["month_end"]).fillna(False).sum())
    infinite_value_count = int(np.isinf(out_panel[features].to_numpy(dtype=float, na_value=np.nan)).sum())
    source_lacks_publish_date_count = int(out_panel["source_lacks_publish_date_flag"].sum())
    source_publish_date_available = bool(out_panel[publish_cols].notna().any().any())

    source_join_qa = pd.DataFrame(
        [
            {
                "source": "income_statement",
                "path": str(COMINS.relative_to(ROOT)),
                "rows_loaded": int(len(income)),
                "missing_columns": ";".join(income_missing),
                "join_match_count": int(panel["operating_profit"].notna().sum()),
                "typrep_not_A_count": int((panel["income_typrep"].fillna("").str.upper() != "A").sum()),
                "publish_date_missing_count": int(panel["income_declare_date"].isna().sum()),
                "publish_after_month_end_count": int((panel["income_declare_date"] > panel["month_end"]).fillna(False).sum()),
            },
            {
                "source": "balance_sheet",
                "path": str(COMBAS.relative_to(ROOT)),
                "rows_loaded": int(len(balance)),
                "missing_columns": ";".join(balance_missing),
                "join_match_count": int(panel["current_assets"].notna().sum()),
                "typrep_not_A_count": int((panel["balance_typrep"].fillna("").str.upper() != "A").sum()),
                "publish_date_missing_count": int(panel["balance_declare_date"].isna().sum()),
                "publish_after_month_end_count": int((panel["balance_declare_date"] > panel["month_end"]).fillna(False).sum()),
            },
            {
                "source": "cash_flow_statement",
                "path": str(COMSCFD.relative_to(ROOT)),
                "rows_loaded": int(len(cash)),
                "missing_columns": ";".join(cash_missing),
                "join_match_count": int(panel["operating_cash_flow"].notna().sum()),
                "typrep_not_A_count": int((panel["cashflow_typrep"].fillna("").str.upper() != "A").sum()),
                "publish_date_missing_count": int(panel["cashflow_declare_date"].isna().sum()),
                "publish_after_month_end_count": int((panel["cashflow_declare_date"] > panel["month_end"]).fillna(False).sum()),
            },
        ]
    )
    source_join_qa.to_csv(OUT_DIR / "derived_feature_source_join_qa.csv", index=False)

    sample_cols = ["symbol", "month_end", "selected_report_period"] + features + [
        "current_assets",
        "inventories",
        "current_liabilities",
        "operating_cash_flow",
        "basic_eps",
        "basic_eps_lag4",
    ]
    out_panel[sample_cols].head(200).to_csv(OUT_DIR / "derived_feature_component_audit_sample.csv", index=False)

    built = {
        "current_ratio_built": out_panel["current_ratio_raw"].notna().any(),
        "quick_ratio_built": out_panel["quick_ratio_raw"].notna().any(),
        "eps_yoy_built": out_panel["eps_yoy_raw"].notna().any(),
        "equity_multiplier_built": out_panel["equity_multiplier_parent_raw"].notna().any(),
        "operating_margin_built": out_panel["operating_margin_raw"].notna().any(),
        "cfo_to_earnings_built": out_panel["cfo_to_earnings_parent_raw"].notna().any(),
    }

    fatal = any(
        [
            rows != 77538,
            symbols_count != 1352,
            months_count != 114,
            duplicate_count > 0,
            selected_pit_violation > 0,
            future_report_violation > 0,
            infinite_value_count > 0,
            not all(built.values()),
        ]
    )
    watch = bool(source_publish_date_available and source_join_qa["publish_after_month_end_count"].sum() > 0) or source_lacks_publish_date_count > 0
    denominator_watch = any(row["true_rate"] > 0.3 for row in invalid_rows if row["flag_name"].endswith("_denominator_invalid"))
    coverage_watch = any(row["coverage"] < 0.5 for row in coverage_rows if row["feature_name"] in ["current_ratio_raw", "quick_ratio_raw", "eps_yoy_raw", "equity_multiplier_parent_raw", "operating_margin_raw", "cfo_to_earnings_parent_raw"])
    if fatal:
        final_decision = "DERIVED_COMPACT_F_MISSING_FEATURES_FAIL_BLOCK_INTEGRATION"
        recommended_next_step = "Fix source joins, PIT alignment, or missing feature construction before integration review."
    elif watch or denominator_watch or coverage_watch:
        final_decision = "DERIVED_COMPACT_F_MISSING_FEATURES_BUILT_WATCH_REVIEW_REQUIRED"
        recommended_next_step = "Review coverage, denominator invalid flags, extreme values, and source publish date warnings before integration."
    else:
        final_decision = "DERIVED_COMPACT_F_MISSING_FEATURES_BUILT_READY_FOR_INTEGRATION_REVIEW"
        recommended_next_step = "Derived Feature Integration Review v0."

    summary = {
        "run_timestamp": now_iso(),
        "base_v3_panel_used": str(V3_PANEL.relative_to(ROOT)),
        "source_files_used": source_files_used,
        "rows": rows,
        "symbols": symbols_count,
        "months": months_count,
        "one_row_per_symbol_month": one_row,
        "duplicate_symbol_month_count": duplicate_count,
        "selected_pit_date_violation_count": selected_pit_violation,
        "future_report_period_violation_count": future_report_violation,
        **{k: bool(v) for k, v in built.items()},
        "source_publish_date_available": source_publish_date_available,
        "source_lacks_publish_date_count": source_lacks_publish_date_count,
        "infinite_value_count": infinite_value_count,
        "v3_missing_input_columns": v3_missing_cols,
        "production_modified": False,
        "v3_modified": False,
        "transformed_panel_modified": False,
        "training_run": False,
        "backtest_run": False,
        "ic_calculated": False,
        "final_decision": final_decision,
        "recommended_next_step": recommended_next_step,
    }
    write_json(OUT_DIR / "derived_compact_f_missing_features_summary.json", summary)

    report = [
        "# Build Derived Compact-F Missing Features Candidate Panel v0",
        "",
        "## 1. Scope",
        "",
        "This run only builds a candidate derived feature panel. It does not train, backtest, calculate IC, or modify production, v3, or the transformed panel.",
        "",
        "## 2. Inputs",
        "",
        f"- Base v3 panel: {V3_PANEL.relative_to(ROOT).as_posix()}",
        *[f"- Source file: {p}" for p in source_files_used],
        "",
        "## 3. Join / PIT Alignment",
        "",
        "The v3 panel is the master table keyed by symbol and month_end. Source statements are joined by symbol and v3 selected_report_period, with Typrep=A preferred and DeclareDate audited against month_end.",
        "",
        "## 4. Derived Feature Formulas",
        "",
        "- Current_Ratio = A001100000 / A002100000",
        "- Quick_Ratio = (A001100000 - A001123000) / A002100000",
        "- EPS_YoY = B003000000 / lag4(B003000000) - 1; diluted EPS variant also output",
        "- Equity_Multiplier = A001000000 / A003100000; total equity variant also output",
        "- Operating_Margin = B001300000 / B001101000; total revenue variant also output",
        "- CFO_to_Earnings = C001000000 / B002000101; total net profit variant also output",
        "",
        "## 5. Coverage and Invalid Flags",
        "",
        "See derived_feature_coverage.csv and derived_feature_invalid_flags.csv.",
        "",
        "## 6. Extreme Value Review",
        "",
        "See derived_feature_extreme_quantiles.csv. Values are recorded raw; no winsor, rank, zscore, or IC is computed.",
        "",
        "## 7. QA Results",
        "",
        f"- Rows / symbols / months: {rows} / {symbols_count} / {months_count}",
        f"- One row per symbol-month: {one_row}",
        f"- Duplicate symbol-month count: {duplicate_count}",
        f"- Selected PIT date violation count: {selected_pit_violation}",
        f"- Future report period violation count: {future_report_violation}",
        f"- Infinite value count: {infinite_value_count}",
        "",
        "## 8. Decision",
        "",
        final_decision,
        "",
        "## 9. Recommended Next Step",
        "",
        recommended_next_step,
        "",
    ]
    (OUT_DIR / "derived_compact_f_missing_features_report.md").write_text("\n".join(report), encoding="utf-8")

    task_card = [
        "# Task Completion Card",
        "",
        "- task_name: Build Derived Compact-F Missing Features Candidate Panel v0",
        f"- completed_at: {now_iso()}",
        f"- final_decision: {final_decision}",
        f"- output_dir: {OUT_DIR.relative_to(ROOT).as_posix()}",
    ]
    (OUT_DIR / "task_completion_card.md").write_text("\n".join(task_card) + "\n", encoding="utf-8")
    write_json(
        OUT_DIR / "terminal_summary.json",
        {
            "script": "scripts/build_derived_compact_f_missing_features_v0.py",
            "status": "completed",
            "stdout_log": "output/_agent_runs/derived_compact_f_missing_features_v0/run_stdout.txt",
            "stderr_log": "output/_agent_runs/derived_compact_f_missing_features_v0/run_stderr.txt",
            "final_decision": final_decision,
        },
    )
    pd.DataFrame([summary]).to_csv(OUT_DIR / "final_qa.csv", index=False)

    del base, panel, out_panel, income_raw, balance_raw, cash_raw, income, balance, cash
    gc.collect()
    print(f"final_decision={final_decision}")
    print(f"output_dir={OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
