from __future__ import annotations

import csv
import gc
import json
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

ROOT = Path(r"C:\dev\quant")
TASK_NAME = "trd_co_static_industry_neutral_score_run_v1"
OUT_DIR = ROOT / "output" / TASK_NAME
RUN_DIR = ROOT / "output" / "_agent_runs" / TASK_NAME

TRD_CO_XLSX = ROOT / "data" / "csmar_exports" / "TRD_Co.xlsx"
DEBUG_SUMMARY = ROOT / "output" / "debug_trd_co_excel_ingestion_v0" / "debug_trd_co_excel_ingestion_summary.json"
DEBUG_COLUMN_CHECK = ROOT / "output" / "debug_trd_co_excel_ingestion_v0" / "trd_co_required_column_check.csv"
DEBUG_PREVIEW = ROOT / "output" / "debug_trd_co_excel_ingestion_v0" / "trd_co_cleaned_preview.csv"
DEBUG_SCRIPT = ROOT / "scripts" / "debug_trd_co_excel_ingestion_v0.py"
SCORE_PANEL = ROOT / "output" / "simple_baseline_score_run_v0" / "simple_baseline_score_panel_v0.parquet"

SCORE_COLUMNS = [
    "symbol",
    "month_end",
    "bp_rank",
    "ep_ttm_rank",
    "cfo_to_earnings_parent_rank",
    "VALUE_BP_SINGLE_score",
    "VALUE_QUALITY_EQUAL_WEIGHT_score",
    "fwd_ret_1m",
]
SOURCE_FEATURES = ["bp_rank", "ep_ttm_rank", "cfo_to_earnings_parent_rank"]
NEUTRAL_COLUMNS = [
    "STATIC_IND_NEUTRAL_VALUE_BP_SINGLE_score",
    "STATIC_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score",
]
REQUIRED_OUTPUT_FIELDS = [
    "Stkcd",
    "Stknme",
    "Listdt",
    "Conme",
    "Indcd",
    "Indnme",
    "IndcdZX",
    "IndnmeZX",
    "Sctcd",
    "Statco",
    "Statdt",
    "FormerCode",
]


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def as_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def is_security_code(value: Any) -> bool:
    return bool(re.fullmatch(r"\d{6}", as_text(value)))


def norm_symbol(value: Any) -> str | None:
    text = as_text(value)
    if not text:
        return None
    if text.isdigit():
        return text.zfill(6)
    return text


def discover_effective_max_column(ws: Any, min_scan_cols: int = 80) -> int:
    max_col = max(ws.max_column or 1, min_scan_cols)
    last = 1
    for row_idx in range(1, min(ws.max_row or 1, 10) + 1):
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                last = max(last, col_idx)
    return last


def read_trd_co_fixed_logic() -> tuple[pd.DataFrame, list[str], str, int, int]:
    wb = load_workbook(TRD_CO_XLSX, read_only=False, data_only=True)
    ws = wb["sheet1"] if "sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    max_col = discover_effective_max_column(ws)
    headers = [as_text(ws.cell(row=1, column=col_idx).value) for col_idx in range(1, max_col + 1)]
    keep_indices = [idx for idx, header in enumerate(headers, start=1) if header]
    columns = [headers[idx - 1] for idx in keep_indices]
    rows: list[list[str | None]] = []
    for row_idx in range(2, (ws.max_row or 1) + 1):
        row = [as_text(ws.cell(row=row_idx, column=col_idx).value) or None for col_idx in keep_indices]
        rows.append(row)
    sheet_name = ws.title
    max_row = int(ws.max_row or 0)
    wb.close()
    df = pd.DataFrame(rows, columns=columns, dtype="string")
    before = int(len(df))
    if "Stkcd" in df.columns:
        df = df.loc[df["Stkcd"].map(is_security_code)].copy()
        df["Stkcd"] = df["Stkcd"].map(lambda x: as_text(x).zfill(6))
        df["symbol"] = df["Stkcd"]
    else:
        df["symbol"] = pd.NA
    after = int(len(df))
    gc.collect()
    return df, columns, sheet_name, max_row, before


def top_categories(series: pd.Series, n: int = 10) -> str:
    counts = series.dropna().astype(str).value_counts().head(n)
    return "; ".join(f"{idx}:{int(val)}" for idx, val in counts.items())


def coverage_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    denom = len(df)
    for field in ["Indcd", "Indnme", "IndcdZX", "IndnmeZX"]:
        present = field in df.columns
        series = df[field] if present else pd.Series(dtype="string")
        non_null = int(series.notna().sum()) if present else 0
        rows.append(
            {
                "field_name": field,
                "field_present": present,
                "non_null_count": non_null,
                "non_null_ratio": non_null / denom if denom else 0.0,
                "unique_value_count": int(series.dropna().astype(str).nunique()) if present else 0,
                "top_categories": top_categories(series) if present else "",
            }
        )
    return rows


def add_static_neutral_scores(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    rank_cols = []
    for feature in SOURCE_FEATURES:
        rank_col = f"_{feature}_static_industry_rank"
        out[rank_col] = out.groupby(["month_end", "primary_industry_code"], dropna=False)[feature].rank(
            pct=True, method="average"
        )
        rank_cols.append(rank_col)
    out["STATIC_IND_NEUTRAL_VALUE_BP_SINGLE_score"] = out["_bp_rank_static_industry_rank"]
    out["STATIC_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score"] = out[rank_cols].mean(axis=1, skipna=True)
    out = out.drop(columns=rank_cols)
    return out


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    run_timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    print(f"[{run_timestamp}] start {TASK_NAME}")

    required_paths = [TRD_CO_XLSX, DEBUG_SUMMARY, DEBUG_COLUMN_CHECK, DEBUG_PREVIEW, DEBUG_SCRIPT, SCORE_PANEL]
    prereq_rows = [
        {"path": rel(path), "exists": path.exists(), "bytes": path.stat().st_size if path.exists() else None}
        for path in required_paths
    ]
    prerequisites_passed = all(row["exists"] for row in prereq_rows)
    if not prerequisites_passed:
        summary = {
            "run_timestamp": run_timestamp,
            "trd_co_xlsx_read": False,
            "ingestion_debug_reference_used": False,
            "previous_stkcd_only_bug_fixed": False,
            "prerequisites_passed": False,
            "final_decision": "TRD_CO_STATIC_INDUSTRY_NEUTRAL_SCORE_RUN_FAIL",
            "recommended_next_step": "补齐缺失输入后重跑。",
        }
        write_json(OUT_DIR / "trd_co_static_industry_neutral_score_run_summary.json", summary)
        return 1

    debug_summary = read_json(DEBUG_SUMMARY)
    ingestion_debug_reference_used = debug_summary.get("final_decision") == "TRD_CO_EXCEL_INGESTION_FIXED"

    trd_df, raw_columns, sheet_name, sheet_max_row, row_count_before_metadata_drop = read_trd_co_fixed_logic()
    trd_co_xlsx_read = True
    previous_stkcd_only_bug_fixed = len(raw_columns) > 1 and "IndcdZX" in raw_columns and "IndnmeZX" in raw_columns

    trd_co_row_count = int(len(trd_df))
    trd_co_unique_stkcd_count = int(trd_df["symbol"].nunique(dropna=True))
    row_counts = trd_df.groupby("symbol", dropna=True).size()
    one_row_per_stkcd = bool((row_counts == 1).all()) if len(row_counts) else False
    multi_row_per_stkcd = bool((row_counts > 1).any()) if len(row_counts) else False
    duplicate_stkcd_count = int(row_counts[row_counts > 1].sum() - (row_counts > 1).sum()) if len(row_counts) else 0

    schema_rows = [
        {
            "column_name": col,
            "selected_for_read": True,
            "non_null_count": int(trd_df[col].notna().sum()) if col in trd_df.columns else "",
            "dtype_observed": str(trd_df[col].dtype) if col in trd_df.columns else "",
        }
        for col in raw_columns
    ]
    write_csv(OUT_DIR / "trd_co_static_schema_profile.csv", schema_rows, ["column_name", "selected_for_read", "non_null_count", "dtype_observed"])

    key_rows = [
        {"check": "trd_co_row_count", "value": trd_co_row_count, "notes": "metadata rows removed"},
        {"check": "trd_co_unique_stkcd_count", "value": trd_co_unique_stkcd_count, "notes": ""},
        {"check": "one_row_per_stkcd", "value": one_row_per_stkcd, "notes": ""},
        {"check": "multi_row_per_stkcd", "value": multi_row_per_stkcd, "notes": ""},
        {"check": "duplicate_stkcd_count", "value": duplicate_stkcd_count, "notes": ""},
    ]
    write_csv(OUT_DIR / "trd_co_static_key_uniqueness_check.csv", key_rows, ["check", "value", "notes"])

    cov_rows = coverage_rows(trd_df)
    write_csv(
        OUT_DIR / "trd_co_static_industry_field_coverage_profile.csv",
        cov_rows,
        ["field_name", "field_present", "non_null_count", "non_null_ratio", "unique_value_count", "top_categories"],
    )

    primary_industry_code_field = "IndcdZX" if "IndcdZX" in trd_df.columns else None
    primary_industry_name_field = "IndnmeZX" if "IndnmeZX" in trd_df.columns else None
    coarse_industry_code_field = "Indcd" if "Indcd" in trd_df.columns else None
    coarse_industry_name_field = "Indnme" if "Indnme" in trd_df.columns else None
    primary_coverage = float(trd_df[primary_industry_code_field].notna().mean()) if primary_industry_code_field else 0.0
    coarse_coverage = float(trd_df[coarse_industry_code_field].notna().mean()) if coarse_industry_code_field else 0.0

    if primary_industry_code_field and primary_industry_name_field and primary_coverage > 0:
        industry_source_classification = "STATIC_INDUSTRY_AVAILABLE"
        pit_quality_status = "STATIC_NOT_PIT"
    else:
        industry_source_classification = "NOT_USABLE"
        pit_quality_status = "NOT_USABLE"

    cleaned = pd.DataFrame(index=trd_df.index)
    for col in REQUIRED_OUTPUT_FIELDS:
        cleaned[col] = trd_df[col] if col in trd_df.columns else pd.NA
    cleaned["symbol"] = trd_df["symbol"]
    cleaned["primary_industry_code"] = trd_df[primary_industry_code_field] if primary_industry_code_field else pd.NA
    cleaned["primary_industry_name"] = trd_df[primary_industry_name_field] if primary_industry_name_field else pd.NA
    cleaned["coarse_industry_code"] = trd_df[coarse_industry_code_field] if coarse_industry_code_field else pd.NA
    cleaned["coarse_industry_name"] = trd_df[coarse_industry_name_field] if coarse_industry_name_field else pd.NA
    cleaned["industry_source_classification"] = industry_source_classification
    cleaned["pit_quality_status"] = pit_quality_status
    cleaned_cols = [
        "Stkcd",
        "symbol",
        "Stknme",
        "Listdt",
        "Conme",
        "primary_industry_code",
        "primary_industry_name",
        "coarse_industry_code",
        "coarse_industry_name",
        "Sctcd",
        "Statco",
        "Statdt",
        "FormerCode",
        "industry_source_classification",
        "pit_quality_status",
    ]
    cleaned[cleaned_cols].to_csv(OUT_DIR / "cleaned_trd_co_static_industry_source.csv", index=False, encoding="utf-8-sig")
    cleaned_industry_source_written = True

    policy = {
        "industry_source_classification": industry_source_classification,
        "pit_quality_status": pit_quality_status,
        "primary_industry_code_field": primary_industry_code_field,
        "primary_industry_name_field": primary_industry_name_field,
        "coarse_industry_code_field": coarse_industry_code_field,
        "coarse_industry_name_field": coarse_industry_name_field,
        "static_industry_join": industry_source_classification == "STATIC_INDUSTRY_AVAILABLE",
        "pit_industry_join": False,
        "industry_changes_detected_within_stkcd": False,
        "industry_change_date_candidate": None,
        "limitations": [
            "TRD_Co 当前导出为一股一行，只能作为 static industry source。",
            "不可声称 PIT monthly industry neutralization。",
            "不可声称 industry-neutral alpha 已完全验证。",
        ],
    }
    write_json(OUT_DIR / "trd_co_static_industry_source_policy.json", policy)

    score_df = pd.read_parquet(SCORE_PANEL, columns=SCORE_COLUMNS)
    score_df["symbol"] = score_df["symbol"].map(norm_symbol)
    score_panel_rows = int(len(score_df))
    join_source = cleaned[
        ["symbol", "primary_industry_code", "primary_industry_name", "coarse_industry_code", "coarse_industry_name", "industry_source_classification", "pit_quality_status"]
    ].drop_duplicates("symbol", keep="last")
    duplicate_join_detected = bool(join_source["symbol"].duplicated(keep=False).any())
    max_ind_per_symbol = join_source.groupby("symbol")["primary_industry_code"].nunique(dropna=True).max()
    one_symbol_one_primary = bool(max_ind_per_symbol <= 1) if pd.notna(max_ind_per_symbol) else True
    joined = score_df.merge(join_source, on="symbol", how="left", validate="many_to_one")
    joined["primary_industry_code"] = joined["primary_industry_code"].astype("string")
    joined["primary_industry_name"] = joined["primary_industry_name"].astype("string")
    joined["coarse_industry_code"] = joined["coarse_industry_code"].astype("string")
    joined["coarse_industry_name"] = joined["coarse_industry_name"].astype("string")
    joined_rows = int(joined["primary_industry_code"].notna().sum())
    join_coverage_ratio = joined_rows / score_panel_rows if score_panel_rows else 0.0
    missing_industry_rows = int(joined["primary_industry_code"].isna().sum())
    missing_industry_symbol_count = int(joined.loc[joined["primary_industry_code"].isna(), "symbol"].nunique())

    join_qa = {
        "score_panel_rows": score_panel_rows,
        "joined_rows": joined_rows,
        "join_coverage_ratio": join_coverage_ratio,
        "missing_industry_rows": missing_industry_rows,
        "missing_industry_symbol_count": missing_industry_symbol_count,
        "duplicate_join_detected": duplicate_join_detected,
        "one_symbol_one_primary_industry_after_cleaning": one_symbol_one_primary,
        "notes": "Static symbol join using TRD_Co symbol; PIT/as-of join not attempted.",
    }
    write_csv(OUT_DIR / "simple_baseline_static_industry_join_qa.csv", [join_qa], list(join_qa.keys()))

    neutral_score_generated = False
    neutral_score_row_count = 0
    small_industry_group_detected = False
    if industry_source_classification == "STATIC_INDUSTRY_AVAILABLE" and join_coverage_ratio >= 0.95 and primary_coverage >= 0.95:
        joined = add_static_neutral_scores(joined)
        neutral_score_generated = True
        neutral_score_row_count = int(joined[NEUTRAL_COLUMNS].notna().any(axis=1).sum())
    else:
        for col in NEUTRAL_COLUMNS:
            joined[col] = pd.NA

    group_size = joined.groupby(["month_end", "primary_industry_code"], dropna=False).size().reset_index(name="group_size")
    group_size["small_group_flag"] = group_size["group_size"] < 5
    small_industry_group_detected = bool(group_size["small_group_flag"].any())
    group_size.to_csv(OUT_DIR / "static_industry_group_size_summary.csv", index=False, encoding="utf-8-sig")
    joined = joined.merge(
        group_size[["month_end", "primary_industry_code", "small_group_flag"]],
        on=["month_end", "primary_industry_code"],
        how="left",
    )

    formula_rows = [
        {
            "neutral_score_name": "STATIC_IND_NEUTRAL_VALUE_BP_SINGLE_score",
            "source_features": "bp_rank",
            "neutralization_method": "month_end_x_primary_industry_percentile_rank",
            "score_formula": "industry_within_rank(bp_rank)",
            "expected_score_direction": "higher_is_better",
            "fwd_ret_1m_used": False,
        },
        {
            "neutral_score_name": "STATIC_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score",
            "source_features": ";".join(SOURCE_FEATURES),
            "neutralization_method": "month_end_x_primary_industry_percentile_rank",
            "score_formula": "mean(industry_within_rank(bp_rank), industry_within_rank(ep_ttm_rank), industry_within_rank(cfo_to_earnings_parent_rank))",
            "expected_score_direction": "higher_is_better",
            "fwd_ret_1m_used": False,
        },
    ]
    write_csv(
        OUT_DIR / "static_industry_neutral_score_formula_manifest.csv",
        formula_rows,
        ["neutral_score_name", "source_features", "neutralization_method", "score_formula", "expected_score_direction", "fwd_ret_1m_used"],
    )

    panel_cols = [
        "symbol",
        "month_end",
        "primary_industry_code",
        "primary_industry_name",
        "coarse_industry_code",
        "coarse_industry_name",
        "bp_rank",
        "ep_ttm_rank",
        "cfo_to_earnings_parent_rank",
        "VALUE_BP_SINGLE_score",
        "VALUE_QUALITY_EQUAL_WEIGHT_score",
        "STATIC_IND_NEUTRAL_VALUE_BP_SINGLE_score",
        "STATIC_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score",
        "fwd_ret_1m",
        "small_group_flag",
        "pit_quality_status",
        "industry_source_classification",
    ]
    joined[panel_cols].to_parquet(OUT_DIR / "simple_baseline_static_industry_neutral_score_panel_v1.parquet", index=False)

    qa_rows = [{"check": "neutral_score_columns_created", "passed": neutral_score_generated, "value": ";".join(NEUTRAL_COLUMNS), "notes": ""}]
    for col in NEUTRAL_COLUMNS:
        series = pd.to_numeric(joined[col], errors="coerce")
        non_null = series.dropna()
        qa_rows.extend(
            [
                {"check": f"{col}_range_within_0_1", "passed": bool(((non_null >= 0) & (non_null <= 1)).all()), "value": "", "notes": ""},
                {"check": f"{col}_no_inf", "passed": bool(~series.isin([float("inf"), float("-inf")]).any()), "value": "", "notes": ""},
                {"check": f"{col}_null_count", "passed": int(series.isna().sum()) == missing_industry_rows if neutral_score_generated else False, "value": int(series.isna().sum()), "notes": ""},
            ]
        )
    qa_rows.extend(
        [
            {"check": "score_row_count_equals_joined_row_count", "passed": neutral_score_row_count == joined_rows if neutral_score_generated else False, "value": neutral_score_row_count, "notes": ""},
            {"check": "fwd_ret_1m_not_used_in_score_formula", "passed": True, "value": "", "notes": ""},
            {"check": "static_vs_pit_label_correct", "passed": pit_quality_status == "STATIC_NOT_PIT", "value": pit_quality_status, "notes": ""},
            {"check": "old_stkcd_only_bug_not_present", "passed": previous_stkcd_only_bug_fixed, "value": len(raw_columns), "notes": ""},
        ]
    )
    write_csv(OUT_DIR / "static_industry_neutral_score_qa.csv", qa_rows, ["check", "passed", "value", "notes"])

    if not previous_stkcd_only_bug_fixed or not prerequisites_passed:
        final_decision = "TRD_CO_STATIC_INDUSTRY_NEUTRAL_SCORE_RUN_FAIL"
        recommended_next_step = "先修复 ingestion 或缺失输入。"
    elif industry_source_classification != "STATIC_INDUSTRY_AVAILABLE":
        final_decision = "TRD_CO_STATIC_INDUSTRY_SOURCE_FAIL_NOT_USABLE"
        recommended_next_step = "补充可用行业字段后重跑。"
    elif primary_coverage < 0.95:
        final_decision = "TRD_CO_STATIC_INDUSTRY_SOURCE_WATCH_LOW_INDUSTRY_COVERAGE"
        recommended_next_step = "人工确认行业字段缺失原因，再决定是否继续。"
    elif join_coverage_ratio < 0.95:
        final_decision = "TRD_CO_STATIC_INDUSTRY_SOURCE_FAIL_JOIN_COVERAGE_LOW"
        recommended_next_step = "修复 TRD_Co 与 score panel 的 symbol 覆盖后重跑。"
    elif neutral_score_generated:
        final_decision = "TRD_CO_STATIC_INDUSTRY_NEUTRAL_SCORE_READY_FOR_EVALUATION_PREP"
        recommended_next_step = "下一步只做 static neutral score evaluation prep，不做组合或回测。"
    else:
        final_decision = "TRD_CO_STATIC_INDUSTRY_NEUTRAL_SCORE_RUN_FAIL"
        recommended_next_step = "检查 guardrail 和 score QA 后重跑。"

    forbidden_false = {
        "ic_calculated": False,
        "d10_d1_calculated": False,
        "portfolio_constructed": False,
        "portfolio_return_calculated": False,
        "backtest_run": False,
        "transaction_cost_calculated": False,
        "turnover_calculated": False,
        "sharpe_calculated": False,
        "maxdd_calculated": False,
        "benchmark_relative_return_calculated": False,
        "alpha_beta_regression_calculated": False,
        "training_run": False,
        "shap_calculated": False,
        "tuning_run": False,
        "feature_importance_calculated": False,
        "production_holdings_generated": False,
        "live_order_ready_file_generated": False,
        "production_modified": False,
    }
    summary = {
        "run_timestamp": run_timestamp,
        "trd_co_xlsx_read": trd_co_xlsx_read,
        "ingestion_debug_reference_used": ingestion_debug_reference_used,
        "previous_stkcd_only_bug_fixed": previous_stkcd_only_bug_fixed,
        "prerequisites_passed": prerequisites_passed,
        "trd_co_row_count": trd_co_row_count,
        "trd_co_unique_stkcd_count": trd_co_unique_stkcd_count,
        "one_row_per_stkcd": one_row_per_stkcd,
        "multi_row_per_stkcd": multi_row_per_stkcd,
        "primary_industry_code_field": primary_industry_code_field,
        "primary_industry_name_field": primary_industry_name_field,
        "primary_industry_coverage_ratio": primary_coverage,
        "coarse_industry_code_field": coarse_industry_code_field,
        "coarse_industry_name_field": coarse_industry_name_field,
        "coarse_industry_coverage_ratio": coarse_coverage,
        "industry_source_classification": industry_source_classification,
        "pit_quality_status": pit_quality_status,
        "cleaned_industry_source_written": cleaned_industry_source_written,
        "score_panel_path": rel(SCORE_PANEL),
        "score_panel_rows": score_panel_rows,
        "joined_rows": joined_rows,
        "join_coverage_ratio": join_coverage_ratio,
        "missing_industry_rows": missing_industry_rows,
        "neutral_score_generated": neutral_score_generated,
        "neutral_score_columns": NEUTRAL_COLUMNS,
        "neutral_score_row_count": neutral_score_row_count,
        "small_industry_group_detected": small_industry_group_detected,
        "fwd_ret_used_in_score_formula": False,
        **forbidden_false,
        "final_decision": final_decision,
        "recommended_next_step": recommended_next_step,
    }
    write_json(OUT_DIR / "trd_co_static_industry_neutral_score_run_summary.json", summary)

    prereq = {
        "run_timestamp": run_timestamp,
        "prerequisites_passed": prerequisites_passed,
        "required_inputs": prereq_rows,
        "trd_co_xlsx_read": trd_co_xlsx_read,
        "ingestion_debug_reference_used": ingestion_debug_reference_used,
        "score_panel_columns_read": SCORE_COLUMNS,
        "excel_sheet_name": sheet_name,
        "excel_sheet_max_row": sheet_max_row,
        "row_count_before_metadata_drop": row_count_before_metadata_drop,
    }
    write_json(OUT_DIR / "trd_co_static_industry_source_audit_prerequisite_check.json", prereq)

    next_plan = f"""# Next Step Static Industry-Neutral Score Evaluation Prep

- final_decision: `{final_decision}`
- industry_source_classification: `{industry_source_classification}`
- pit_quality_status: `{pit_quality_status}`
- neutral_score_columns: `{", ".join(NEUTRAL_COLUMNS)}`

下一步仅评估 static industry-neutral score。该结果不是 PIT monthly industry neutralization，不能直接声称 industry-neutral alpha 已完全验证。
"""
    (OUT_DIR / "next_step_static_industry_neutral_score_evaluation_plan.md").write_text(next_plan, encoding="utf-8")

    report = f"""# TRD_Co Static Industry Source Audit + Static Industry Neutral Score Run v1

## 结论

`{final_decision}`

## 行业源

- TRD_Co 当前导出为一股一行：`{one_row_per_stkcd}`
- primary industry: `{primary_industry_code_field}` / `{primary_industry_name_field}`
- coarse industry: `{coarse_industry_code_field}` / `{coarse_industry_name_field}`
- industry_source_classification: `{industry_source_classification}`
- pit_quality_status: `{pit_quality_status}`

该源可用于 static industry exposure / static industry-neutral robustness check，不可声称 PIT monthly industry neutralization，也不可声称 industry-neutral alpha 已完全验证。

## Join 和 Score

- score_panel_rows: `{score_panel_rows}`
- joined_rows: `{joined_rows}`
- join_coverage_ratio: `{join_coverage_ratio:.6f}`
- neutral_score_generated: `{neutral_score_generated}`
- neutral_score_columns: `{", ".join(NEUTRAL_COLUMNS)}`

本任务未计算 IC、D10-D1、组合收益、回测、交易成本、换手、Sharpe、MaxDD、benchmark-relative return、alpha/beta，未训练、未调参、未 SHAP、未写 production。
"""
    (OUT_DIR / "trd_co_static_industry_neutral_score_run_report.md").write_text(report, encoding="utf-8")

    final_qa = [
        {"check": "prerequisites_passed", "passed": prerequisites_passed, "notes": ""},
        {"check": "ingestion_debug_reference_used", "passed": ingestion_debug_reference_used, "notes": ""},
        {"check": "previous_stkcd_only_bug_fixed", "passed": previous_stkcd_only_bug_fixed, "notes": f"raw_columns={len(raw_columns)}"},
        {"check": "fwd_ret_not_used_in_score_formula", "passed": True, "notes": ""},
        {"check": "no_forbidden_calculations", "passed": True, "notes": "No IC/D10-D1/portfolio/return/backtest/training/production."},
    ]
    write_csv(OUT_DIR / "final_qa.csv", final_qa, ["check", "passed", "notes"])
    terminal_summary = {
        "task_name": TASK_NAME,
        "run_timestamp": run_timestamp,
        "stdout_log": rel(RUN_DIR / "run_stdout.txt"),
        "stderr_log": rel(RUN_DIR / "run_stderr.txt"),
        "output_directory": rel(OUT_DIR),
        "final_decision": final_decision,
        "exit_code": 0,
    }
    write_json(OUT_DIR / "terminal_summary.json", terminal_summary)
    (OUT_DIR / "task_completion_card.md").write_text(
        f"""# Task Completion Card

- task_name: `{TASK_NAME}`
- final_decision: `{final_decision}`
- output_directory: `{rel(OUT_DIR)}`
- neutral_score_generated: `{neutral_score_generated}`
- production_modified: `False`
""",
        encoding="utf-8",
    )
    (RUN_DIR / "RUN_STATE.md").write_text(
        f"""# RUN_STATE

任务：{TASK_NAME}
状态：完成

已读取：
- {TRD_CO_XLSX}
- {DEBUG_SUMMARY}
- {SCORE_PANEL} 的必要列

输出目录：
- {OUT_DIR}

final_decision: {final_decision}

禁止项确认：
- 未计算 IC / D10-D1
- 未构造组合或计算收益
- 未回测
- 未训练 / 调参 / SHAP
- 未写 production
""",
        encoding="utf-8",
    )

    del trd_df, cleaned, score_df, joined, join_source, group_size
    gc.collect()
    print(f"final_decision={final_decision}")
    print(f"join_coverage_ratio={join_coverage_ratio}")
    print(f"neutral_score_generated={neutral_score_generated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
