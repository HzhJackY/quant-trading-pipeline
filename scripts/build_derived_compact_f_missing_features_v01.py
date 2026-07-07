from __future__ import annotations

import gc
import json
import math
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / "derived_compact_f_missing_features_v01"

V3_PANEL = ROOT / "output" / "csmar_pit_clean_core_financial_factors_v3" / "pit_clean_core_financial_factors_monthly_v3.parquet"
V0_SUMMARY = ROOT / "output" / "derived_compact_f_missing_features_v0" / "derived_compact_f_missing_features_summary.json"
V0_COVERAGE = ROOT / "output" / "derived_compact_f_missing_features_v0" / "derived_feature_coverage.csv"
V0_INVALID = ROOT / "output" / "derived_compact_f_missing_features_v0" / "derived_feature_invalid_flags.csv"
V0_JOIN_QA = ROOT / "output" / "derived_compact_f_missing_features_v0" / "derived_feature_source_join_qa.csv"
V0_REPORT = ROOT / "output" / "derived_compact_f_missing_features_v0" / "derived_compact_f_missing_features_report.md"
PATCH_SUMMARY = ROOT / "output" / "existing_fs_missing_feature_constructibility_patch_v0" / "existing_fs_missing_feature_constructibility_patch_summary.json"
FORMULAS = ROOT / "output" / "existing_fs_missing_feature_constructibility_patch_v0" / "derived_feature_formula_candidates.csv"
RECLASSIFIED = ROOT / "output" / "existing_fs_missing_feature_constructibility_patch_v0" / "compact_f_missing_feature_reclassified.csv"

COMINS = ROOT / "data" / "csmar_exports" / "FS_Comins.xlsx"
COMBAS = ROOT / "data" / "csmar_exports" / "FS_Combas.xlsx"
COMBAS2 = ROOT / "data" / "csmar_exports" / "FS_Combas2.xlsx"
COMSCFD = ROOT / "data" / "csmar_exports" / "FS_Comscfd.xlsx"
COMBAS2_DES = ROOT / "data" / "csmar_exports" / "FS_Combas2[DES][xlsx].txt"

REQUIRED_INPUTS = [
    V3_PANEL,
    V0_SUMMARY,
    V0_COVERAGE,
    V0_INVALID,
    V0_JOIN_QA,
    V0_REPORT,
    PATCH_SUMMARY,
    FORMULAS,
    RECLASSIFIED,
    COMINS,
    COMBAS,
    COMBAS2,
    COMSCFD,
    COMBAS2_DES,
]

V3_REQUIRED = [
    "symbol",
    "month_end",
    "selected_report_period",
    "selected_pit_date",
    "ttm_complete_flag",
    "ttm_quarters_available",
    "uses_pre_2017_buffer_flag",
    "factor_validity_flags",
    "net_profit_parent_ttm",
    "net_profit_ttm",
    "revenue_ttm",
    "total_assets",
    "total_liabilities",
    "equity_parent",
    "total_equity",
]
INCOME_COLS = ["Stkcd", "Accper", "Typrep", "DeclareDate", "B001101000", "B001100000", "B001300000", "B002000000", "B002000101", "B003000000", "B004000000"]
BALANCE_COLS = ["Stkcd", "Accper", "Typrep", "DeclareDate", "A001100000", "A002100000", "A001000000", "A003100000", "A003000000"]
BALANCE2_COLS = ["Stkcd", "Accper", "Typrep", "DeclareDate", "A001100000", "A001123000", "A002100000"]
CASHFLOW_COLS = ["Stkcd", "Accper", "Typrep", "DeclareDate", "C001000000"]
EPS = 1e-12


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def json_default(value: Any) -> Any:
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return None if math.isnan(float(value)) else float(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    return str(value)


def write_json(path: Path, payload: dict[str, Any] | list[Any]) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=json_default), encoding="utf-8")


def missing_input_report(missing: list[Path]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "missing_input_report.md").write_text(
        "# Missing Input Report\n\n" + "\n".join(f"- {p.as_posix()}" for p in missing) + "\n",
        encoding="utf-8",
    )
    write_json(
        OUT_DIR / "derived_compact_f_missing_features_v01_summary.json",
        {"run_timestamp": now_iso(), "final_decision": "DERIVED_COMPACT_F_MISSING_FEATURES_V01_FAIL_BLOCK_INTEGRATION", "missing_inputs": [str(p) for p in missing]},
    )


def symbol_norm(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    if "." in text:
        text = text.split(".", 1)[0]
    return text.zfill(6)


def date_norm(value: Any) -> pd.Timestamp | pd.NaT:
    return pd.to_datetime(value, errors="coerce")


def period_key(value: Any) -> str | None:
    dt = date_norm(value)
    return None if pd.isna(dt) else dt.strftime("%Y-%m-%d")


def lag4_period(value: Any) -> str | None:
    dt = date_norm(value)
    return None if pd.isna(dt) else (dt - pd.DateOffset(years=1)).strftime("%Y-%m-%d")


def read_excel_filtered(path: Path, wanted_cols: list[str], symbols: set[str], periods: set[str]) -> tuple[pd.DataFrame, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(name): idx for idx, name in enumerate(header) if name is not None}
    missing_cols = [col for col in wanted_cols if col not in header_map]
    present_cols = [col for col in wanted_cols if col in header_map]
    read_cols = sorted(set(present_cols + [c for c in ["Stkcd", "Accper"] if c in header_map]), key=lambda c: header_map[c])
    indices = [header_map[col] for col in read_cols]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sym = symbol_norm(row[header_map["Stkcd"]]) if "Stkcd" in header_map else None
        per = period_key(row[header_map["Accper"]]) if "Accper" in header_map else None
        if sym not in symbols or per not in periods:
            continue
        rows.append({col: row[idx] for col, idx in zip(read_cols, indices)})
    wb.close()
    df = pd.DataFrame(rows)
    for col in wanted_cols:
        if col not in df.columns:
            df[col] = np.nan
    return df[wanted_cols], missing_cols


def prepare_source(df: pd.DataFrame, prefix: str, value_cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    out["symbol"] = out["Stkcd"].apply(symbol_norm)
    out["report_period"] = out["Accper"].apply(period_key)
    out[f"{prefix}_declare_date"] = pd.to_datetime(out["DeclareDate"], errors="coerce") if "DeclareDate" in out else pd.NaT
    out[f"{prefix}_typrep"] = out["Typrep"].astype(str) if "Typrep" in out else ""
    for col in value_cols:
        out[col] = pd.to_numeric(out[col], errors="coerce") if col in out else np.nan
    out["_typrep_priority"] = np.where(out[f"{prefix}_typrep"].str.upper().eq("A"), 0, 1)
    out = out.sort_values(["symbol", "report_period", "_typrep_priority", f"{prefix}_declare_date"])
    out = out.drop_duplicates(["symbol", "report_period"], keep="first")
    return out[["symbol", "report_period", f"{prefix}_typrep", f"{prefix}_declare_date"] + value_cols]


def denominator_invalid(series: pd.Series, negative_invalid: bool = True) -> pd.Series:
    s = pd.to_numeric(series, errors="coerce")
    invalid = s.isna() | (s.abs() <= EPS)
    if negative_invalid:
        invalid = invalid | (s < 0)
    return invalid


def ratio(numer: pd.Series, denom: pd.Series) -> pd.Series:
    n = pd.to_numeric(numer, errors="coerce")
    d = pd.to_numeric(denom, errors="coerce")
    return (n / d.replace(0, np.nan)).replace([np.inf, -np.inf], np.nan)


def quantile_row(panel: pd.DataFrame, feature: str) -> dict[str, Any]:
    s = pd.to_numeric(panel[feature], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    qs = s.quantile([0.001, 0.01, 0.05, 0.5, 0.95, 0.99, 0.999]) if len(s) else pd.Series(dtype=float)
    return {
        "feature_name": feature,
        "non_null_count": int(len(s)),
        "min": float(s.min()) if len(s) else None,
        "p001": float(qs.get(0.001)) if len(s) else None,
        "p01": float(qs.get(0.01)) if len(s) else None,
        "p05": float(qs.get(0.05)) if len(s) else None,
        "p50": float(qs.get(0.5)) if len(s) else None,
        "p95": float(qs.get(0.95)) if len(s) else None,
        "p99": float(qs.get(0.99)) if len(s) else None,
        "p999": float(qs.get(0.999)) if len(s) else None,
        "max": float(s.max()) if len(s) else None,
    }


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    missing = [p for p in REQUIRED_INPUTS if not p.exists()]
    if missing:
        missing_input_report(missing)
        print(f"Missing required inputs: {len(missing)}")
        return 0

    des_text = COMBAS2_DES.read_text(encoding="utf-8", errors="replace")
    inventory_confirmed = "A001123000" in des_text and "存货净额" in des_text

    v3_cols = pq.ParquetFile(V3_PANEL).schema.names
    v3_read = [c for c in V3_REQUIRED if c in v3_cols]
    base = pd.read_parquet(V3_PANEL, columns=v3_read)
    base["symbol"] = base["symbol"].apply(symbol_norm)
    base["month_end"] = pd.to_datetime(base["month_end"], errors="coerce")
    base["selected_report_period"] = pd.to_datetime(base["selected_report_period"], errors="coerce")
    base["selected_pit_date"] = pd.to_datetime(base["selected_pit_date"], errors="coerce")
    base["report_period_key"] = base["selected_report_period"].apply(period_key)
    base["lag4_report_period_key"] = base["selected_report_period"].apply(lag4_period)

    symbols = set(base["symbol"].dropna().astype(str))
    periods = set(base["report_period_key"].dropna().astype(str))
    eps_periods = periods | set(base["lag4_report_period_key"].dropna().astype(str))

    income_raw, income_missing = read_excel_filtered(COMINS, INCOME_COLS, symbols, eps_periods)
    balance_raw, balance_missing = read_excel_filtered(COMBAS, BALANCE_COLS, symbols, periods)
    balance2_raw, balance2_missing = read_excel_filtered(COMBAS2, BALANCE2_COLS, symbols, periods)
    cash_raw, cash_missing = read_excel_filtered(COMSCFD, CASHFLOW_COLS, symbols, periods)

    income = prepare_source(income_raw, "income", ["B001101000", "B001100000", "B001300000", "B002000000", "B002000101", "B003000000", "B004000000"])
    balance = prepare_source(balance_raw, "balance", ["A001100000", "A002100000", "A001000000", "A003100000", "A003000000"])
    balance2 = prepare_source(balance2_raw, "balance2", ["A001100000", "A001123000", "A002100000"])
    cash = prepare_source(cash_raw, "cashflow", ["C001000000"])

    panel = base.merge(income, how="left", left_on=["symbol", "report_period_key"], right_on=["symbol", "report_period"]).drop(columns=["report_period"], errors="ignore")
    panel = panel.merge(balance, how="left", left_on=["symbol", "report_period_key"], right_on=["symbol", "report_period"]).drop(columns=["report_period"], errors="ignore")
    panel = panel.merge(balance2, how="left", left_on=["symbol", "report_period_key"], right_on=["symbol", "report_period"]).drop(columns=["report_period"], errors="ignore")
    panel = panel.merge(cash, how="left", left_on=["symbol", "report_period_key"], right_on=["symbol", "report_period"]).drop(columns=["report_period"], errors="ignore")
    eps_lag = income[["symbol", "report_period", "B003000000", "B004000000"]].rename(columns={"report_period": "lag4_report_period_key", "B003000000": "basic_eps_lag4", "B004000000": "diluted_eps_lag4"})
    panel = panel.merge(eps_lag, how="left", on=["symbol", "lag4_report_period_key"])

    rename = {
        "A001100000_x": "current_assets",
        "A002100000_x": "current_liabilities",
        "A001000000": "total_assets_component",
        "A003100000": "equity_parent_component",
        "A003000000": "total_equity_component",
        "A001100000_y": "quick_current_assets",
        "A001123000": "inventories",
        "A002100000_y": "quick_current_liabilities",
        "B001300000": "operating_profit",
        "B001101000": "operating_revenue",
        "B001100000": "total_revenue",
        "B002000101": "net_profit_parent",
        "B002000000": "net_profit",
        "B003000000": "basic_eps",
        "B004000000": "diluted_eps",
        "C001000000": "operating_cash_flow",
    }
    panel = panel.rename(columns=rename)
    for col in rename.values():
        if col not in panel.columns:
            panel[col] = np.nan

    panel["current_ratio_denominator_invalid"] = denominator_invalid(panel["current_liabilities"])
    panel["quick_ratio_denominator_invalid"] = denominator_invalid(panel["quick_current_liabilities"])
    panel["eps_yoy_denominator_invalid"] = denominator_invalid(panel["basic_eps_lag4"], negative_invalid=True)
    panel["equity_multiplier_parent_denominator_invalid"] = denominator_invalid(panel["equity_parent_component"])
    panel["equity_multiplier_total_denominator_invalid"] = denominator_invalid(panel["total_equity_component"])
    panel["operating_margin_denominator_invalid"] = denominator_invalid(panel["operating_revenue"])
    panel["cfo_to_earnings_parent_denominator_invalid"] = denominator_invalid(panel["net_profit_parent"], negative_invalid=True)
    panel["cfo_to_earnings_total_denominator_invalid"] = denominator_invalid(panel["net_profit"], negative_invalid=True)

    panel["current_ratio_raw"] = ratio(panel["current_assets"], panel["current_liabilities"])
    panel["quick_ratio_raw"] = ratio(panel["quick_current_assets"] - panel["inventories"], panel["quick_current_liabilities"])
    panel["eps_yoy_raw"] = ratio(panel["basic_eps"], panel["basic_eps_lag4"]) - 1
    panel["diluted_eps_yoy_raw"] = ratio(panel["diluted_eps"], panel["diluted_eps_lag4"]) - 1
    panel["equity_multiplier_parent_raw"] = ratio(panel["total_assets_component"], panel["equity_parent_component"])
    panel["equity_multiplier_total_raw"] = ratio(panel["total_assets_component"], panel["total_equity_component"])
    panel["operating_margin_raw"] = ratio(panel["operating_profit"], panel["operating_revenue"])
    panel["operating_margin_total_revenue_raw"] = ratio(panel["operating_profit"], panel["total_revenue"])
    panel["cfo_to_earnings_parent_raw"] = ratio(panel["operating_cash_flow"], panel["net_profit_parent"])
    panel["cfo_to_earnings_total_raw"] = ratio(panel["operating_cash_flow"], panel["net_profit"])

    publish_cols = ["income_declare_date", "balance_declare_date", "balance2_declare_date", "cashflow_declare_date"]
    typrep_cols = ["income_typrep", "balance_typrep", "balance2_typrep", "cashflow_typrep"]
    for col in publish_cols:
        if col not in panel.columns:
            panel[col] = pd.NaT
    for col in typrep_cols:
        if col not in panel.columns:
            panel[col] = ""
    panel["source_lacks_publish_date_flag"] = panel[publish_cols].isna().any(axis=1)
    panel["typrep_not_A_flag"] = ~panel[typrep_cols].fillna("").apply(lambda row: all(str(x).upper() == "A" for x in row if str(x) != ""), axis=1)
    source_publish_after = pd.Series(False, index=panel.index)
    for col in publish_cols:
        source_publish_after = source_publish_after | (pd.to_datetime(panel[col], errors="coerce") > panel["month_end"])
    panel["source_publish_date_after_month_end_flag"] = source_publish_after

    component_flags = {
        "current_ratio_component_missing_flag": ["current_assets", "current_liabilities"],
        "quick_ratio_component_missing": ["quick_current_assets", "inventories", "quick_current_liabilities"],
        "eps_yoy_component_missing_flag": ["basic_eps", "basic_eps_lag4"],
        "equity_multiplier_component_missing_flag": ["total_assets_component", "equity_parent_component"],
        "operating_margin_component_missing_flag": ["operating_profit", "operating_revenue"],
        "cfo_to_earnings_component_missing_flag": ["operating_cash_flow", "net_profit_parent"],
    }
    for flag, cols in component_flags.items():
        panel[flag] = panel[cols].isna().any(axis=1)

    panel["source_panel_version"] = "csmar_pit_clean_core_financial_factors_v3"
    panel["derived_feature_version"] = "derived_compact_f_missing_features_v01"
    out_cols = [
        "symbol", "month_end", "selected_report_period", "selected_pit_date", "source_panel_version", "derived_feature_version",
        "current_ratio_raw", "quick_ratio_raw", "eps_yoy_raw", "diluted_eps_yoy_raw", "equity_multiplier_parent_raw", "equity_multiplier_total_raw",
        "operating_margin_raw", "operating_margin_total_revenue_raw", "cfo_to_earnings_parent_raw", "cfo_to_earnings_total_raw",
        "current_ratio_denominator_invalid", "quick_ratio_denominator_invalid", "quick_ratio_component_missing", "eps_yoy_denominator_invalid",
        "equity_multiplier_parent_denominator_invalid", "equity_multiplier_total_denominator_invalid", "operating_margin_denominator_invalid",
        "cfo_to_earnings_parent_denominator_invalid", "cfo_to_earnings_total_denominator_invalid", "source_lacks_publish_date_flag",
        "source_publish_date_after_month_end_flag", "typrep_not_A_flag",
        "current_ratio_component_missing_flag", "eps_yoy_component_missing_flag", "equity_multiplier_component_missing_flag",
        "operating_margin_component_missing_flag", "cfo_to_earnings_component_missing_flag",
        "current_assets", "inventories", "current_liabilities", "quick_current_assets", "quick_current_liabilities",
        "total_assets_component", "equity_parent_component", "total_equity_component", "operating_profit", "operating_revenue", "total_revenue",
        "net_profit_parent", "net_profit", "operating_cash_flow", "basic_eps", "diluted_eps", "basic_eps_lag4", "diluted_eps_lag4",
        "income_declare_date", "balance_declare_date", "balance2_declare_date", "cashflow_declare_date", "income_typrep", "balance_typrep", "balance2_typrep", "cashflow_typrep",
    ]
    out = panel[out_cols].copy()
    out.to_parquet(OUT_DIR / "derived_compact_f_missing_features_v01.parquet", index=False)

    features = [
        "current_ratio_raw", "quick_ratio_raw", "eps_yoy_raw", "diluted_eps_yoy_raw", "equity_multiplier_parent_raw", "equity_multiplier_total_raw",
        "operating_margin_raw", "operating_margin_total_revenue_raw", "cfo_to_earnings_parent_raw", "cfo_to_earnings_total_raw",
    ]
    coverage = [{"feature_name": f, "non_null_count": int(out[f].notna().sum()), "missing_count": int(out[f].isna().sum()), "coverage": float(out[f].notna().mean())} for f in features]
    pd.DataFrame(coverage).to_csv(OUT_DIR / "derived_feature_coverage_v01.csv", index=False)
    pd.DataFrame([quantile_row(out, f) for f in features]).to_csv(OUT_DIR / "derived_feature_extreme_quantiles_v01.csv", index=False)

    flags = [
        "current_ratio_denominator_invalid", "quick_ratio_denominator_invalid", "quick_ratio_component_missing", "eps_yoy_denominator_invalid",
        "equity_multiplier_parent_denominator_invalid", "equity_multiplier_total_denominator_invalid", "operating_margin_denominator_invalid",
        "cfo_to_earnings_parent_denominator_invalid", "cfo_to_earnings_total_denominator_invalid", "source_lacks_publish_date_flag",
        "source_publish_date_after_month_end_flag", "typrep_not_A_flag", "current_ratio_component_missing_flag", "eps_yoy_component_missing_flag",
        "equity_multiplier_component_missing_flag", "operating_margin_component_missing_flag", "cfo_to_earnings_component_missing_flag",
    ]
    invalid_rows = [{"flag_name": f, "true_count": int(out[f].fillna(False).sum()), "true_rate": float(out[f].fillna(False).mean())} for f in flags]
    pd.DataFrame(invalid_rows).to_csv(OUT_DIR / "derived_feature_invalid_flags_v01.csv", index=False)

    source_join_qa = pd.DataFrame([
        {"source": "income_statement", "path": str(COMINS.relative_to(ROOT)), "rows_loaded": len(income), "missing_columns": ";".join(income_missing), "join_match_count": int(panel["operating_profit"].notna().sum()), "typrep_not_A_count": int((panel["income_typrep"].fillna("").str.upper() != "A").sum()), "publish_date_missing_count": int(panel["income_declare_date"].isna().sum()), "publish_after_month_end_count": int((panel["income_declare_date"] > panel["month_end"]).fillna(False).sum())},
        {"source": "balance_sheet", "path": str(COMBAS.relative_to(ROOT)), "rows_loaded": len(balance), "missing_columns": ";".join(balance_missing), "join_match_count": int(panel["current_assets"].notna().sum()), "typrep_not_A_count": int((panel["balance_typrep"].fillna("").str.upper() != "A").sum()), "publish_date_missing_count": int(panel["balance_declare_date"].isna().sum()), "publish_after_month_end_count": int((panel["balance_declare_date"] > panel["month_end"]).fillna(False).sum())},
        {"source": "balance_sheet_inventory_patch", "path": str(COMBAS2.relative_to(ROOT)), "rows_loaded": len(balance2), "missing_columns": ";".join(balance2_missing), "join_match_count": int(panel["inventories"].notna().sum()), "typrep_not_A_count": int((panel["balance2_typrep"].fillna("").str.upper() != "A").sum()), "publish_date_missing_count": int(panel["balance2_declare_date"].isna().sum()), "publish_after_month_end_count": int((panel["balance2_declare_date"] > panel["month_end"]).fillna(False).sum())},
        {"source": "cash_flow_statement", "path": str(COMSCFD.relative_to(ROOT)), "rows_loaded": len(cash), "missing_columns": ";".join(cash_missing), "join_match_count": int(panel["operating_cash_flow"].notna().sum()), "typrep_not_A_count": int((panel["cashflow_typrep"].fillna("").str.upper() != "A").sum()), "publish_date_missing_count": int(panel["cashflow_declare_date"].isna().sum()), "publish_after_month_end_count": int((panel["cashflow_declare_date"] > panel["month_end"]).fillna(False).sum())},
    ])
    source_join_qa.to_csv(OUT_DIR / "derived_feature_source_join_qa_v01.csv", index=False)
    fs_combas2_qa = source_join_qa[source_join_qa["source"].eq("balance_sheet_inventory_patch")].copy()
    fs_combas2_qa["inventory_field_code"] = "A001123000"
    fs_combas2_qa["inventory_field_name"] = "存货净额"
    fs_combas2_qa["inventory_field_confirmed_by_des"] = inventory_confirmed
    fs_combas2_qa["quick_ratio_coverage"] = float(out["quick_ratio_raw"].notna().mean())
    fs_combas2_qa["quick_ratio_component_missing_count"] = int(out["quick_ratio_component_missing"].sum())
    fs_combas2_qa.to_csv(OUT_DIR / "fs_combas2_inventory_join_qa_v01.csv", index=False)

    source_after_count = int(out["source_publish_date_after_month_end_flag"].sum())
    if source_after_count:
        out.loc[out["source_publish_date_after_month_end_flag"], ["symbol", "month_end", "selected_report_period", "selected_pit_date", "income_declare_date", "balance_declare_date", "balance2_declare_date", "cashflow_declare_date"]].head(100).to_csv(OUT_DIR / "source_publish_date_after_month_end_sample_v01.csv", index=False)

    sample_cols = ["symbol", "month_end", "selected_report_period"] + features + ["quick_current_assets", "inventories", "quick_current_liabilities", "operating_cash_flow", "basic_eps", "basic_eps_lag4"]
    out[sample_cols].head(200).to_csv(OUT_DIR / "derived_feature_component_audit_sample_v01.csv", index=False)

    rows = len(out)
    symbols_count = int(out["symbol"].nunique())
    months_count = int(out["month_end"].nunique())
    duplicate_count = int(out.duplicated(["symbol", "month_end"]).sum())
    selected_pit_violation = int((out["selected_pit_date"] > out["month_end"]).fillna(False).sum())
    future_report_violation = int((out["selected_report_period"] > out["month_end"]).fillna(False).sum())
    infinite_count = int(np.isinf(out[features].to_numpy(dtype=float, na_value=np.nan)).sum())
    built = {
        "current_ratio_built": bool(out["current_ratio_raw"].notna().any()),
        "quick_ratio_built": bool(out["quick_ratio_raw"].notna().any()),
        "eps_yoy_built": bool(out["eps_yoy_raw"].notna().any()),
        "equity_multiplier_built": bool(out["equity_multiplier_parent_raw"].notna().any()),
        "operating_margin_built": bool(out["operating_margin_raw"].notna().any()),
        "cfo_to_earnings_built": bool(out["cfo_to_earnings_parent_raw"].notna().any()),
    }
    quick_cov = float(out["quick_ratio_raw"].notna().mean())
    quick_missing = int(out["quick_ratio_component_missing"].sum())
    source_lacks = int(out["source_lacks_publish_date_flag"].sum())
    source_pub_available = bool(out[publish_cols := ["income_declare_date", "balance_declare_date", "balance2_declare_date", "cashflow_declare_date"]].notna().any().any())

    fatal = any([not inventory_confirmed, not all(built.values()), duplicate_count > 0, selected_pit_violation > 0, future_report_violation > 0, infinite_count > 0])
    severe_source_after = source_after_count > int(rows * 0.05)
    if fatal or severe_source_after:
        final_decision = "DERIVED_COMPACT_F_MISSING_FEATURES_V01_FAIL_BLOCK_INTEGRATION"
        recommended_next_step = "Fix source joins, inventory DES confirmation, PIT alignment, or source date violations before integration review."
    elif source_after_count > 0 or source_lacks > 0:
        final_decision = "DERIVED_COMPACT_F_MISSING_FEATURES_V01_WATCH_SOURCE_DATE_REVIEW_REQUIRED"
        recommended_next_step = "Review source publish date warnings, then proceed to Derived Feature Integration Review v0 if accepted."
    else:
        final_decision = "DERIVED_COMPACT_F_MISSING_FEATURES_V01_READY_FOR_INTEGRATION_REVIEW"
        recommended_next_step = "Derived Feature Integration Review v0."

    summary = {
        "run_timestamp": now_iso(),
        "base_v3_panel_used": str(V3_PANEL.relative_to(ROOT)),
        "source_files_used": [str(p.relative_to(ROOT)) for p in [COMINS, COMBAS, COMBAS2, COMSCFD]],
        "fs_combas2_used": str(COMBAS2.relative_to(ROOT)),
        "inventory_field_code": "A001123000",
        "inventory_field_name": "存货净额",
        "inventory_field_confirmed_by_des": inventory_confirmed,
        "rows": int(rows),
        "symbols": symbols_count,
        "months": months_count,
        "one_row_per_symbol_month": duplicate_count == 0,
        "duplicate_symbol_month_count": duplicate_count,
        "selected_pit_date_violation_count": selected_pit_violation,
        "future_report_period_violation_count": future_report_violation,
        "source_publish_date_after_month_end_violation_count": source_after_count,
        **built,
        "quick_ratio_coverage": quick_cov,
        "quick_ratio_component_missing_count": quick_missing,
        "source_publish_date_available": source_pub_available,
        "source_lacks_publish_date_count": source_lacks,
        "infinite_value_count": infinite_count,
        "production_modified": False,
        "v3_modified": False,
        "transformed_panel_modified": False,
        "training_run": False,
        "backtest_run": False,
        "ic_calculated": False,
        "final_decision": final_decision,
        "recommended_next_step": recommended_next_step,
    }
    write_json(OUT_DIR / "derived_compact_f_missing_features_v01_summary.json", summary)

    report = [
        "# Derived Compact-F Missing Features Candidate Panel v0.1",
        "",
        "## 1. Scope",
        "",
        "This run only builds a candidate derived feature panel. It does not train, backtest, calculate IC, or modify production, v3, or the transformed panel.",
        "",
        "## 2. Inputs",
        "",
        *[f"- {p}" for p in summary["source_files_used"]],
        "",
        "## 3. FS_Combas2 Inventory Field Confirmation",
        "",
        f"- A001123000 [存货净额] confirmed by DES: {inventory_confirmed}",
        "",
        "## 4. Join / PIT Alignment",
        "",
        "The v3 panel is the master table. Source statements are joined by symbol and v3 selected_report_period, with Typrep=A preferred and DeclareDate audited against month_end.",
        "",
        "## 5. Derived Feature Formulas",
        "",
        "- Current_Ratio = A001100000 / A002100000",
        "- Quick_Ratio = (A001100000 - A001123000) / A002100000",
        "- EPS_YoY = B003000000 / lag4(B003000000) - 1",
        "- Equity_Multiplier = A001000000 / A003100000; total equity alternative retained",
        "- Operating_Margin = B001300000 / B001101000; total revenue alternative retained",
        "- CFO_to_Earnings = C001000000 / B002000101; total net profit alternative retained",
        "",
        "## 6. Coverage and Invalid Flags",
        "",
        "See derived_feature_coverage_v01.csv and derived_feature_invalid_flags_v01.csv.",
        "",
        "## 7. Quick Ratio Patch Result",
        "",
        f"- quick_ratio_built: {built['quick_ratio_built']}",
        f"- quick_ratio_coverage: {quick_cov}",
        f"- quick_ratio_component_missing_count: {quick_missing}",
        "",
        "## 8. Source Publish Date Review",
        "",
        f"- source_publish_date_after_month_end_violation_count: {source_after_count}",
        f"- source_lacks_publish_date_count: {source_lacks}",
        "",
        "## 9. QA Results",
        "",
        f"- Rows / symbols / months: {rows} / {symbols_count} / {months_count}",
        f"- Duplicate count: {duplicate_count}",
        f"- Selected PIT violation count: {selected_pit_violation}",
        f"- Future report period violation count: {future_report_violation}",
        f"- Infinite value count: {infinite_count}",
        "",
        "## 10. Decision",
        "",
        final_decision,
        "",
        "## 11. Recommended Next Step",
        "",
        recommended_next_step,
        "",
    ]
    (OUT_DIR / "derived_compact_f_missing_features_v01_report.md").write_text("\n".join(report), encoding="utf-8")
    (OUT_DIR / "task_completion_card.md").write_text(
        f"# Task Completion Card\n\n- task_name: Derived Compact-F Missing Features Candidate Panel v0.1 Patch\n- completed_at: {now_iso()}\n- final_decision: {final_decision}\n- output_dir: {OUT_DIR.relative_to(ROOT).as_posix()}\n",
        encoding="utf-8",
    )
    write_json(OUT_DIR / "terminal_summary.json", {"script": "scripts/build_derived_compact_f_missing_features_v01.py", "status": "completed", "stdout_log": "output/_agent_runs/derived_compact_f_missing_features_v01/run_stdout.txt", "stderr_log": "output/_agent_runs/derived_compact_f_missing_features_v01/run_stderr.txt", "final_decision": final_decision})
    pd.DataFrame([summary]).to_csv(OUT_DIR / "final_qa.csv", index=False)

    del base, panel, out, income_raw, balance_raw, balance2_raw, cash_raw, income, balance, balance2, cash
    gc.collect()
    print(f"final_decision={final_decision}")
    print(f"output_dir={OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
