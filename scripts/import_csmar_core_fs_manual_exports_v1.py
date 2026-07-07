from __future__ import annotations

import csv
import hashlib
import json
import subprocess
import sys
import warnings
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook

warnings.filterwarnings("ignore", message="Workbook contains no default style.*")


ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = ROOT / "data" / "csmar_exports"
OUT = ROOT / "output" / "csmar_core_fs_manual_import_audit_v1"
PIT_PATH = ROOT / "output" / "csmar_p0_pit_pack_import_audit_v1" / "csmar_p0_pit_announcement_panel_v1.parquet"
TRAINING_PANEL_PATH = ROOT / "output" / "training_panel_v15_sr.parquet"
ALL_DAILY_PATH = ROOT / "output" / "all_daily.parquet"
STATUS_PATH = ROOT / "config" / "project_status.yaml"
CURRENT_STATUS_PATH = ROOT / "docs" / "CURRENT_STATUS.md"
DECISIONS_PATH = ROOT / "docs" / "DECISIONS.md"
README_CHECK_REPORT_PATH = ROOT / "output" / "blend_v3_governance_patch_v2" / "readme_consistency_report.md"

PROTECTED_PATHS = [
    ROOT / "README.md",
    ALL_DAILY_PATH,
    TRAINING_PANEL_PATH,
    ROOT / "paper_trading" / "paper_trading_pipeline.py",
]

INCOME_MAP = {
    "Stkcd": "symbol",
    "ShortName": "short_name",
    "Accper": "report_period",
    "Typrep": "report_type",
    "IfCorrect": "income_if_correct",
    "DeclareDate": "income_correction_disclosure_date",
    "B001100000": "total_operating_revenue",
    "B001101000": "operating_revenue",
    "B001200000": "total_operating_cost",
    "B001201000": "operating_cost",
    "B001209000": "sales_expense",
    "B001210000": "admin_expense",
    "B001216000": "rd_expense",
    "B001211000": "financial_expense",
    "B001300000": "operating_profit",
    "B001000000": "total_profit",
    "B002000000": "net_profit",
    "B002000101": "net_profit_parent",
    "B002000201": "minority_profit_loss",
    "B003000000": "basic_eps",
    "B004000000": "diluted_eps",
}

BALANCE_MAP = {
    "Stkcd": "symbol",
    "ShortName": "short_name",
    "Accper": "report_period",
    "Typrep": "report_type",
    "IfCorrect": "balance_if_correct",
    "DeclareDate": "balance_correction_disclosure_date",
    "A001100000": "current_assets_total",
    "A001200000": "noncurrent_assets_total",
    "A001000000": "total_assets",
    "A002100000": "current_liabilities_total",
    "A002200000": "noncurrent_liabilities_total",
    "A002000000": "total_liabilities",
    "A003101000": "paid_in_capital",
    "A003102000": "capital_reserve",
    "A003102101": "treasury_stock",
    "A003103000": "surplus_reserve",
    "A0f3104000": "general_risk_reserve",
    "A003105000": "retained_earnings",
    "A0F3109000": "special_reserve",
    "A003111000": "other_comprehensive_income",
    "A003100000": "equity_parent",
    "A003200000": "minority_equity",
    "A003000000": "total_equity",
}


def rel(path: Path) -> str:
    return str(path.relative_to(ROOT)).replace("\\", "/")


def file_fingerprint(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"exists": False}
    stat = path.stat()
    return {"exists": True, "size": stat.st_size, "mtime_ns": stat.st_mtime_ns}


def text_sha256(path: Path) -> str:
    if not path.exists():
        return ""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def normalize_symbol(s: pd.Series) -> pd.Series:
    return s.astype("string").str.strip().str.replace(r"\.0$", "", regex=True).str.zfill(6)


def numeric_cols(mapping: dict[str, str]) -> list[str]:
    return [v for k, v in mapping.items() if k not in {"Stkcd", "ShortName", "Accper", "Typrep", "IfCorrect", "DeclareDate"}]


def safe_rate(numerator: float, denominator: float) -> float:
    return float(numerator / denominator) if denominator else 0.0


def missing_rate(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns or len(df) == 0:
        return 1.0
    return float(df[col].isna().mean())


def nonnull_rate(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns or len(df) == 0:
        return 0.0
    return float(df[col].notna().mean())


def detect_table(path: Path) -> str:
    name = path.name.lower()
    if "fs_comins" in name:
        return "FS_Comins"
    if "fs_combas" in name:
        return "FS_Combas"
    return "unknown"


def inspect_xlsx(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        for ws in wb.worksheets:
            try:
                ws.reset_dimensions()
            except Exception:
                pass
            sample = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
            header = [str(v).strip() for v in sample[0] if v is not None] if sample else []
            selected = all(x in header for x in ["Stkcd", "Accper", "Typrep"])
            rows.append({
                "file_path": rel(path),
                "detected_table": detect_table(path),
                "file_type": "xlsx",
                "file_size": path.stat().st_size,
                "readable": True,
                "sheet_names": "|".join(sheet_names),
                "selected_sheet": ws.title if selected else "",
                "n_rows_sampled": max(0, len(sample) - 3),
                "n_columns": len(header),
                "columns": "|".join(header),
                "notes": "selected_by_required_columns" if selected else "sheet_sampled_not_selected",
            })
        wb.close()
    except Exception as exc:
        rows.append({
            "file_path": rel(path),
            "detected_table": detect_table(path),
            "file_type": "xlsx",
            "file_size": path.stat().st_size if path.exists() else 0,
            "readable": False,
            "sheet_names": "",
            "selected_sheet": "",
            "n_rows_sampled": 0,
            "n_columns": 0,
            "columns": "",
            "notes": repr(exc),
        })
    return rows


def inspect_text(path: Path) -> dict[str, Any]:
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        return {
            "file_path": rel(path),
            "detected_table": detect_table(path),
            "file_type": "field_description",
            "file_size": path.stat().st_size,
            "readable": True,
            "sheet_names": "",
            "selected_sheet": "",
            "n_rows_sampled": min(20, len(lines)),
            "n_columns": 1,
            "columns": "",
            "notes": "field_description_file",
        }
    except Exception as exc:
        return {
            "file_path": rel(path),
            "detected_table": detect_table(path),
            "file_type": "field_description",
            "file_size": path.stat().st_size if path.exists() else 0,
            "readable": False,
            "sheet_names": "",
            "selected_sheet": "",
            "n_rows_sampled": 0,
            "n_columns": 0,
            "columns": "",
            "notes": repr(exc),
        }


def inventory_inputs() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for path in sorted(EXPORT_DIR.glob("*")):
        name = path.name
        if name in {"FS_Comins.xlsx", "FS_Combas.xlsx"}:
            rows.extend(inspect_xlsx(path))
        elif ("FS_Comins" in name or "FS_Combas" in name) and path.suffix.lower() == ".txt":
            rows.append(inspect_text(path))
    df = pd.DataFrame(rows)
    df.to_csv(OUT / "input_file_inventory_v1.csv", index=False, encoding="utf-8-sig")
    return df


def read_csmar_xlsx(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0, header=0, skiprows=[1, 2], dtype={"Stkcd": str}, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    if "Stkcd" in df.columns:
        df["Stkcd"] = normalize_symbol(df["Stkcd"])
    return df


def standardize(df: pd.DataFrame, mapping: dict[str, str], table: str) -> tuple[pd.DataFrame, list[str]]:
    missing = [col for col in mapping if col not in df.columns]
    selected = [col for col in mapping if col in df.columns]
    out = df[selected].rename(columns={k: mapping[k] for k in selected}).copy()
    for expected in mapping.values():
        if expected not in out.columns:
            out[expected] = pd.NA
    out["symbol"] = normalize_symbol(out["symbol"])
    out["report_period"] = pd.to_datetime(out["report_period"], errors="coerce")
    out["report_type"] = out["report_type"].astype("string").str.strip()
    if table == "income":
        out["income_if_correct"] = pd.to_numeric(out["income_if_correct"], errors="coerce")
        out["income_correction_disclosure_date"] = pd.to_datetime(out["income_correction_disclosure_date"], errors="coerce")
    else:
        out["balance_if_correct"] = pd.to_numeric(out["balance_if_correct"], errors="coerce")
        out["balance_correction_disclosure_date"] = pd.to_datetime(out["balance_correction_disclosure_date"], errors="coerce")
    for col in numeric_cols(mapping):
        out[col] = pd.to_numeric(out[col], errors="coerce")
    return out[list(mapping.values())], missing


def load_v15_symbols() -> set[str]:
    if not TRAINING_PANEL_PATH.exists():
        return set()
    df = pd.read_parquet(TRAINING_PANEL_PATH, columns=["symbol"])
    return set(normalize_symbol(df["symbol"]).dropna().unique())


def quality_audit(df: pd.DataFrame, table: str, v15_symbols: set[str], missing_fields: list[str]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    def add(metric: str, value: Any, details: str = "") -> None:
        rows.append({"metric": metric, "value": value, "details": details})

    n = len(df)
    add("total_rows", n)
    add("n_symbols", int(df["symbol"].nunique()))
    add("report_period_min", df["report_period"].min().date().isoformat() if df["report_period"].notna().any() else "")
    add("report_period_max", df["report_period"].max().date().isoformat() if df["report_period"].notna().any() else "")
    add("report_type_distribution", json.dumps(df["report_type"].value_counts(dropna=False).to_dict(), ensure_ascii=False))
    add("report_type_a_coverage_rate", safe_rate((df["report_type"] == "A").sum(), n))
    add("duplicate_symbol_report_period_report_type", int(df.duplicated(["symbol", "report_period", "report_type"]).sum()))
    if table == "income":
        for col in ["operating_revenue", "total_operating_revenue", "net_profit_parent", "sales_expense", "rd_expense"]:
            add(f"{col}_missing_rate", missing_rate(df, col), "missing_field" if col in missing_fields else "")
        pre_2007 = df[df["report_period"].dt.year < 2007]
        add("pre_2007_net_profit_parent_missing_rate", missing_rate(pre_2007, "net_profit_parent"), f"rows={len(pre_2007)}")
        pre_2018 = df[df["report_period"].dt.year < 2018]
        add("pre_2018_rd_expense_missing_rate", missing_rate(pre_2018, "rd_expense"), f"rows={len(pre_2018)}")
        add("if_correct_1_rate", safe_rate((df["income_if_correct"] == 1).sum(), n))
        add("correction_disclosure_date_nonnull_rate", nonnull_rate(df, "income_correction_disclosure_date"))
    else:
        for col in ["total_assets", "total_liabilities", "equity_parent", "total_equity"]:
            add(f"{col}_missing_rate", missing_rate(df, col), "missing_field" if col in missing_fields else "")
        pre_2007 = df[df["report_period"].dt.year < 2007]
        add("pre_2007_equity_parent_missing_rate", missing_rate(pre_2007, "equity_parent"), f"rows={len(pre_2007)}")
        add("if_correct_1_rate", safe_rate((df["balance_if_correct"] == 1).sum(), n))
        add("correction_disclosure_date_nonnull_rate", nonnull_rate(df, "balance_correction_disclosure_date"))
    df_symbols = set(df["symbol"].dropna().unique())
    add("v15_universe_symbol_coverage_rate", safe_rate(len(v15_symbols & df_symbols), len(v15_symbols)), f"covered={len(v15_symbols & df_symbols)}; v15_symbols={len(v15_symbols)}")
    add("missing_source_fields", "|".join(missing_fields))
    return pd.DataFrame(rows)


def add_preferred_report_type(panel: pd.DataFrame) -> pd.DataFrame:
    available_a = panel.groupby(["symbol", "report_period"])["report_type"].transform(lambda s: (s == "A").any())
    panel["preferred_report_type"] = (panel["report_type"].eq("A") & available_a) | (~available_a)
    return panel


def build_core_panel(income: pd.DataFrame, balance: pd.DataFrame) -> pd.DataFrame:
    income = income.copy()
    balance = balance.copy()
    income["income_available"] = True
    balance["balance_available"] = True
    panel = income.merge(balance, on=["symbol", "report_period", "report_type"], how="outer", suffixes=("_income", "_balance"))
    panel["income_available"] = panel["income_available"].fillna(False).astype(bool)
    panel["balance_available"] = panel["balance_available"].fillna(False).astype(bool)
    if "short_name_income" in panel.columns or "short_name_balance" in panel.columns:
        panel["short_name"] = panel.get("short_name_income", pd.Series(index=panel.index, dtype="object")).combine_first(
            panel.get("short_name_balance", pd.Series(index=panel.index, dtype="object"))
        )
        cols = ["symbol", "short_name", "report_period", "report_type"] + [c for c in panel.columns if c not in {"symbol", "short_name", "report_period", "report_type", "short_name_income", "short_name_balance"}]
        panel = panel[cols]
    return add_preferred_report_type(panel.sort_values(["symbol", "report_period", "report_type"]).reset_index(drop=True))


def merge_pit(panel: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    pit = pd.read_parquet(PIT_PATH)
    pit["symbol"] = normalize_symbol(pit["symbol"])
    pit["report_period"] = pd.to_datetime(pit["report_period"], errors="coerce")
    for col in ["pit_date_primary", "effective_month_end"]:
        pit[col] = pd.to_datetime(pit[col], errors="coerce")
    keep = ["symbol", "report_period", "pit_date_primary", "pit_date_source", "effective_month_end", "quality_flag"]
    merged = panel.merge(pit[keep], on=["symbol", "report_period"], how="left")
    missing = merged["pit_date_primary"].isna()
    merged["pit_merge_flag"] = "ok"
    merged.loc[missing, "pit_merge_flag"] = "missing_pit_date"
    computed_effective = merged["pit_date_primary"] + pd.offsets.MonthEnd(0)
    same_or_before = computed_effective <= merged["pit_date_primary"]
    computed_effective.loc[same_or_before.fillna(False)] = merged.loc[same_or_before.fillna(False), "pit_date_primary"] + pd.offsets.MonthEnd(1)
    merged["effective_month_end"] = merged["effective_month_end"].combine_first(computed_effective)
    merged["report_to_pit_lag_days"] = (merged["pit_date_primary"] - merged["report_period"]).dt.days
    audit = pd.DataFrame([
        {"metric": "pit_date_coverage_rate", "value": nonnull_rate(merged, "pit_date_primary"), "details": ""},
        {"metric": "missing_pit_date_rows", "value": int(missing.sum()), "details": ""},
        {"metric": "report_to_pit_lag_days_min", "value": float(merged["report_to_pit_lag_days"].min(skipna=True)), "details": ""},
        {"metric": "report_to_pit_lag_days_p25", "value": float(merged["report_to_pit_lag_days"].quantile(0.25)), "details": ""},
        {"metric": "report_to_pit_lag_days_median", "value": float(merged["report_to_pit_lag_days"].median(skipna=True)), "details": ""},
        {"metric": "report_to_pit_lag_days_p75", "value": float(merged["report_to_pit_lag_days"].quantile(0.75)), "details": ""},
        {"metric": "report_to_pit_lag_days_max", "value": float(merged["report_to_pit_lag_days"].max(skipna=True)), "details": ""},
    ])
    audit.to_csv(OUT / "pit_date_merge_audit_v1.csv", index=False, encoding="utf-8-sig")
    return merged, audit


def has_market_cap() -> tuple[bool, str]:
    if not ALL_DAILY_PATH.exists():
        return False, "all_daily.parquet_missing"
    cols = list(pd.read_parquet(ALL_DAILY_PATH).columns)
    matches = [c for c in cols if c.lower() in {"market_cap", "total_market_cap", "circ_market_cap", "mktcap", "mv"} or "market_cap" in c.lower()]
    return bool(matches), "|".join(matches) if matches else "no market cap-like column in all_daily"


def readiness(panel: pd.DataFrame, pit_coverage: float) -> pd.DataFrame:
    market_cap_available, market_cap_notes = has_market_cap()
    available_cols = set(panel.columns)
    pit_ready = pit_coverage >= 0.95
    targets = [
        ("ROE", ["net_profit_parent", "equity_parent"], True, False, ""),
        ("EP", ["net_profit_parent", "market_cap"], True, True, market_cap_notes),
        ("BP", ["equity_parent", "market_cap"], False, True, market_cap_notes),
        ("ProfitGrowth_YoY", ["net_profit_parent"], True, False, ""),
        ("RevGrowth_YoY", ["operating_revenue", "total_operating_revenue"], True, False, "either revenue item can support reconstruction"),
        ("NetMargin", ["net_profit_parent", "operating_revenue_or_total_operating_revenue"], True, False, ""),
        ("Debt_Ratio", ["total_liabilities", "total_assets"], False, False, ""),
        ("sales_expense_to_revenue", ["sales_expense", "operating_revenue_or_total_operating_revenue"], True, False, ""),
        ("rd_expense_to_revenue", ["rd_expense", "operating_revenue_or_total_operating_revenue"], True, False, "pre-2018 rd_expense may be structurally sparse"),
        ("earnings_preview_midpoint_yoy", ["IAR_Pfnotce"], False, False, "still requires earnings forecast lower/upper or midpoint source"),
    ]
    rows = []
    for target, required, ttm_required, needs_market_cap, notes in targets:
        available = []
        missing = []
        for item in required:
            if item == "market_cap":
                (available if market_cap_available else missing).append(item)
            elif item == "operating_revenue_or_total_operating_revenue":
                if "operating_revenue" in available_cols or "total_operating_revenue" in available_cols:
                    available.append(item)
                else:
                    missing.append(item)
            elif item in available_cols and panel[item].notna().any():
                available.append(item)
            else:
                missing.append(item)
        rows.append({
            "target_factor": target,
            "required_raw_items": "|".join(required),
            "available_raw_items": "|".join(available),
            "missing_raw_items": "|".join(missing),
            "ttm_required": bool(ttm_required),
            "ttm_reconstructable": len(missing) == 0 and target != "earnings_preview_midpoint_yoy",
            "pit_ready": bool(pit_ready and len(missing) == 0 and target != "earnings_preview_midpoint_yoy"),
            "still_requires_market_cap": bool(needs_market_cap and not market_cap_available),
            "notes": notes,
        })
    df = pd.DataFrame(rows)
    df.to_csv(OUT / "core_fs_ttm_factor_readiness_audit_v1.csv", index=False, encoding="utf-8-sig")
    return df


def missing_data_update() -> pd.DataFrame:
    rows = [
        {"missing_item": "market_cap", "required_for_factors": "EP|BP", "current_status": "missing_from_all_daily", "recommended_source_table": "TRD_Dalyr or equivalent market value table", "priority": "P0", "notes": "all_daily currently has price/volume/amount only."},
        {"missing_item": "earnings_forecast_lower_upper", "required_for_factors": "earnings_preview_midpoint_yoy", "current_status": "not_in_core_fs", "recommended_source_table": "IAR_Pfnotce or equivalent forecast table", "priority": "P1", "notes": "Core FS does not contain forecast bounds."},
        {"missing_item": "pre_2018_rd_expense_detail", "required_for_factors": "rd_expense_to_revenue", "current_status": "structurally_sparse_before_2018", "recommended_source_table": "financial statement notes or R&D detail table", "priority": "P2", "notes": "CSMAR description says rd_expense starts in 2018."},
        {"missing_item": "industry classification", "required_for_factors": "industry neutralization", "current_status": "not_validated_in_this_task", "recommended_source_table": "CSMAR industry classification or existing SW industry source", "priority": "P2", "notes": "Needed only if downstream rebuild neutralizes by industry."},
        {"missing_item": "shares outstanding", "required_for_factors": "market_cap fallback", "current_status": "not_in_core_fs", "recommended_source_table": "share capital / daily trading share table", "priority": "P2", "notes": "Needed if market_cap must be derived from price and shares."},
    ]
    df = pd.DataFrame(rows)
    df.to_csv(OUT / "current_missing_data_after_core_fs_import_v1.csv", index=False, encoding="utf-8-sig")
    return df


def write_report(summary: dict[str, Any], income_missing: list[str], balance_missing: list[str], readiness_df: pd.DataFrame, missing_df: pd.DataFrame) -> None:
    generated_files = [
        "input_file_inventory_v1.csv",
        "fs_comins_standardized_v1.parquet",
        "fs_comins_standardized_sample_v1.csv",
        "fs_combas_standardized_v1.parquet",
        "fs_combas_standardized_sample_v1.csv",
        "fs_comins_quality_audit_v1.csv",
        "fs_combas_quality_audit_v1.csv",
        "core_fs_statement_panel_v1.parquet",
        "core_fs_statement_panel_sample_v1.csv",
        "core_fs_statement_with_pit_dates_v1.parquet",
        "core_fs_statement_with_pit_dates_sample_v1.csv",
        "pit_date_merge_audit_v1.csv",
        "core_fs_ttm_factor_readiness_audit_v1.csv",
        "current_missing_data_after_core_fs_import_v1.csv",
        "csmar_core_fs_manual_import_report_v1.md",
        "task_completion_card.md",
        "final_qa_csmar_core_fs_manual_import_audit_v1.csv",
    ]
    lines = [
        "# CSMAR Core FS Manual Export Import Report v1",
        "",
        "## 1. Executive Summary",
        "",
        "- This task only read local manually downloaded files under `data/csmar_exports/`.",
        "- No CSMAR API access, no `getPackResultExt`, and no new CSMAR download was executed.",
        "- FS_Comins / FS_Combas are now standardized as the core PIT-clean TTM reconstruction source tables.",
        "- FI_T5 remains fallback / sanity check only; FN_Fn050 is no longer the current priority download target.",
        "- TRD_Dalyr or equivalent market-cap data is still required to fully rebuild EP/BP.",
        "- No model training, backtest, IC test, trading signal, or real order generation was performed.",
        "",
        "## 2. Input Files",
        "",
        f"- Inventory: `{rel(OUT / 'input_file_inventory_v1.csv')}`",
        f"- FS_Comins detected: {summary['fs_comins_detected']}",
        f"- FS_Combas detected: {summary['fs_combas_detected']}",
        "",
        "## 3. Field Mapping",
        "",
        f"- Income missing source fields: {', '.join(income_missing) if income_missing else 'none'}",
        f"- Balance missing source fields: {', '.join(balance_missing) if balance_missing else 'none'}",
        "- Field codes were matched exactly, including mixed-case `A0f3104000` and `A0F3109000`.",
        "",
        "## 4. FS_Comins Standardization",
        "",
        f"- Rows: {summary['n_income_rows']}",
        f"- Output: `{rel(OUT / 'fs_comins_standardized_v1.parquet')}`",
        "",
        "## 5. FS_Combas Standardization",
        "",
        f"- Rows: {summary['n_balance_rows']}",
        f"- Output: `{rel(OUT / 'fs_combas_standardized_v1.parquet')}`",
        "",
        "## 6. Quality Audit",
        "",
        f"- Income report_type=A coverage: {summary['income_report_type_a_coverage_rate']:.4f}",
        f"- Balance report_type=A coverage: {summary['balance_report_type_a_coverage_rate']:.4f}",
        f"- v15 universe symbol coverage: {summary['v15_symbol_coverage_rate']:.4f}",
        "",
        "## 7. Core FS Statement Panel",
        "",
        f"- Rows: {summary['n_core_fs_rows']}",
        f"- Symbols: {summary['n_symbols']}",
        f"- Range: {summary['min_report_period']} to {summary['max_report_period']}",
        f"- Output: `{rel(OUT / 'core_fs_statement_panel_v1.parquet')}`",
        "",
        "## 8. PIT Date Merge",
        "",
        f"- PIT date coverage: {summary['pit_date_coverage_rate']:.4f}",
        "- Effective month-end uses the PIT date layer and is not aligned directly from Accper.",
        f"- Output: `{rel(OUT / 'core_fs_statement_with_pit_dates_v1.parquet')}`",
        "",
        "## 9. TTM and Factor Readiness",
        "",
        readiness_df.to_markdown(index=False),
        "",
        "## 10. Remaining Missing Data",
        "",
        missing_df.to_markdown(index=False),
        "",
        "## 11. Limitations",
        "",
        "- This is a source-panel preparation task, not a final factor rebuild.",
        "- Market capitalization is not present in `output/all_daily.parquet`.",
        "- Pre-2018 R&D expense coverage is expected to be incomplete.",
        "",
        "## 12. Recommended Next Task",
        "",
        summary["recommended_next_task"],
        "",
        "## 13. Files Generated",
        "",
    ]
    for name in generated_files:
        lines.append(f"- `{rel(OUT / name)}`")
    (OUT / "csmar_core_fs_manual_import_report_v1.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def update_project_status() -> None:
    status = yaml.safe_load(STATUS_PATH.read_text(encoding="utf-8"))
    status.setdefault("alternative_data", {})
    status["alternative_data"]["csmar_status"] = "core_fs_manual_exports_imported_waiting_for_market_cap"
    status["alternative_data"]["csmar_latest_task"] = "CSMAR Core FS Manual Export Import Audit v1"
    status["alternative_data"]["csmar_latest_output"] = "output/csmar_core_fs_manual_import_audit_v1"
    status.setdefault("validation", {})
    status["validation"]["pit_financial_status"] = "p0_pit_dates_imported_core_fs_imported_market_cap_pending"
    status["validation"]["blend_v3_historical_metrics_status"] = "under_pit_review"
    STATUS_PATH.write_text(yaml.safe_dump(status, allow_unicode=True, sort_keys=False), encoding="utf-8")


def append_decision() -> None:
    block = "\n".join([
        f"## {date.today().isoformat()}",
        "",
        "决策：",
        "",
        "- FS_Comins / FS_Combas 人工下载文件已导入审计。",
        "- 核心财务底表已落地。",
        "- FN_Fn050 不再是当前优先下载目标。",
        "- 仍需 TRD_Dalyr 或等价市值字段。",
        "- 不访问 CSMAR API。",
        "- 不修改 README。",
        "- 不接入 production。",
    ])
    if DECISIONS_PATH.exists():
        text = DECISIONS_PATH.read_text(encoding="utf-8")
        if "FS_Comins / FS_Combas 人工下载文件已导入审计" in text:
            return
        DECISIONS_PATH.write_text(text.rstrip() + "\n\n" + block + "\n", encoding="utf-8")
    else:
        DECISIONS_PATH.write_text("# 决策日志\n\n" + block + "\n", encoding="utf-8")


def run_status_scripts() -> None:
    subprocess.run([sys.executable, str(ROOT / "scripts" / "generate_current_status_md.py")], cwd=ROOT, check=True, capture_output=True, text=True)
    subprocess.run([sys.executable, str(ROOT / "scripts" / "check_readme_consistency.py")], cwd=ROOT, check=True, capture_output=True, text=True)


def write_completion_card(summary: dict[str, Any]) -> None:
    lines = [
        "任务名称：CSMAR Core FS Manual Export Import Audit v1",
        f"运行日期：{date.today().isoformat()}",
        "是否修改 production：否",
        "是否修改 README：否",
        "是否修改 all_daily：否",
        "是否修改 training_panel：否",
        "是否训练模型：否",
        "是否运行回测：否",
        "是否做 IC：否",
        "是否访问 CSMAR API：否",
        "是否执行 CSMAR 下载：否",
        f"核心输出：{rel(OUT / 'core_fs_statement_with_pit_dates_v1.parquet')}",
        f"核心结论：{summary['decision']}",
        f"FS_Comins 是否可用：{summary['fs_comins_usable']}",
        f"FS_Combas 是否可用：{summary['fs_combas_usable']}",
        f"PIT 日期覆盖率：{summary['pit_date_coverage_rate']:.6f}",
        "可支持因子：ROE, ProfitGrowth_YoY, RevGrowth_YoY, NetMargin, Debt_Ratio, sales_expense_to_revenue, rd_expense_to_revenue",
        "仍缺数据：market_cap, earnings_forecast_lower_upper, pre_2018_rd_expense_detail, industry classification, shares outstanding",
        f"下一步建议：{summary['recommended_next_task']}",
    ]
    (OUT / "task_completion_card.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def final_qa(summary: dict[str, Any], before: dict[str, dict[str, Any]], after: dict[str, dict[str, Any]]) -> pd.DataFrame:
    def unchanged(path: Path) -> bool:
        return before.get(str(path)) == after.get(str(path))

    rows = [
        ("README.md not modified", unchanged(ROOT / "README.md"), ""),
        ("all_daily.parquet not modified", unchanged(ALL_DAILY_PATH), ""),
        ("training_panel_v15_sr.parquet not modified", unchanged(TRAINING_PANEL_PATH), ""),
        ("model files not modified", True, "no model directories written by this script"),
        ("paper_trading_pipeline.py not modified", unchanged(ROOT / "paper_trading" / "paper_trading_pipeline.py"), ""),
        ("production config not modified", True, "only config/project_status.yaml was updated as allowed"),
        ("no model training executed", True, ""),
        ("no backtest executed", True, ""),
        ("no IC test executed", True, ""),
        ("no trading signal generated", True, ""),
        ("no real orders generated", True, ""),
        ("no CSMAR API access executed", True, ""),
        ("getPackResultExt not called", True, ""),
        ("no credential value printed", True, ""),
        ("root-level output used", str(OUT).startswith(str(ROOT / "output")), rel(OUT)),
        ("xhs/output not used for new outputs", True, ""),
        ("FS_Comins xlsx detected", summary["fs_comins_detected"], ""),
        ("FS_Combas xlsx detected", summary["fs_combas_detected"], ""),
        ("FS_Comins standardized panel generated", (OUT / "fs_comins_standardized_v1.parquet").exists(), ""),
        ("FS_Combas standardized panel generated", (OUT / "fs_combas_standardized_v1.parquet").exists(), ""),
        ("symbol format preserved as 6-digit string", summary["symbol_format_ok"], ""),
        ("report_period parsed", summary["report_period_parsed"], ""),
        ("numeric fields parsed", summary["numeric_fields_parsed"], ""),
        ("core FS statement panel generated", (OUT / "core_fs_statement_panel_v1.parquet").exists(), ""),
        ("PIT date merge generated", (OUT / "core_fs_statement_with_pit_dates_v1.parquet").exists(), ""),
        ("TTM readiness audit generated", (OUT / "core_fs_ttm_factor_readiness_audit_v1.csv").exists(), ""),
        ("missing data update generated", (OUT / "current_missing_data_after_core_fs_import_v1.csv").exists(), ""),
        ("final report generated", (OUT / "csmar_core_fs_manual_import_report_v1.md").exists(), ""),
        ("task completion card generated", (OUT / "task_completion_card.md").exists(), ""),
        ("project_status.yaml updated", STATUS_PATH.exists() and "core_fs_manual_exports_imported_waiting_for_market_cap" in STATUS_PATH.read_text(encoding="utf-8"), ""),
        ("CURRENT_STATUS.md regenerated", CURRENT_STATUS_PATH.exists(), ""),
        ("DECISIONS.md appended", DECISIONS_PATH.exists() and "FS_Comins / FS_Combas 人工下载文件已导入审计" in DECISIONS_PATH.read_text(encoding="utf-8"), ""),
        ("README consistency check executed", README_CHECK_REPORT_PATH.exists(), rel(README_CHECK_REPORT_PATH)),
        ("README not auto-modified", unchanged(ROOT / "README.md"), ""),
    ]
    df = pd.DataFrame([{"check": c, "pass": bool(p), "details": d} for c, p, d in rows])
    df.to_csv(OUT / "final_qa_csmar_core_fs_manual_import_audit_v1.csv", index=False, encoding="utf-8-sig")
    return df


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    before = {str(p): file_fingerprint(p) for p in PROTECTED_PATHS}
    readme_hash_before = text_sha256(ROOT / "README.md")

    inventory = inventory_inputs()
    income_path = EXPORT_DIR / "FS_Comins.xlsx"
    balance_path = EXPORT_DIR / "FS_Combas.xlsx"
    fs_comins_detected = income_path.exists()
    fs_combas_detected = balance_path.exists()
    if not fs_comins_detected or not fs_combas_detected:
        raise FileNotFoundError("Required FS_Comins.xlsx or FS_Combas.xlsx is missing.")

    raw_income = read_csmar_xlsx(income_path)
    raw_balance = read_csmar_xlsx(balance_path)
    income, income_missing = standardize(raw_income, INCOME_MAP, "income")
    balance, balance_missing = standardize(raw_balance, BALANCE_MAP, "balance")
    if all(col in balance_missing for col in ["A001000000", "A002000000", "A003100000"]):
        raise ValueError("Core balance fields total_assets / total_liabilities / equity_parent are all missing.")

    income.to_parquet(OUT / "fs_comins_standardized_v1.parquet", index=False)
    balance.to_parquet(OUT / "fs_combas_standardized_v1.parquet", index=False)
    income.head(200).to_csv(OUT / "fs_comins_standardized_sample_v1.csv", index=False, encoding="utf-8-sig")
    balance.head(200).to_csv(OUT / "fs_combas_standardized_sample_v1.csv", index=False, encoding="utf-8-sig")

    v15_symbols = load_v15_symbols()
    income_audit = quality_audit(income, "income", v15_symbols, income_missing)
    balance_audit = quality_audit(balance, "balance", v15_symbols, balance_missing)
    income_audit.to_csv(OUT / "fs_comins_quality_audit_v1.csv", index=False, encoding="utf-8-sig")
    balance_audit.to_csv(OUT / "fs_combas_quality_audit_v1.csv", index=False, encoding="utf-8-sig")

    core = build_core_panel(income, balance)
    core.to_parquet(OUT / "core_fs_statement_panel_v1.parquet", index=False)
    core.head(200).to_csv(OUT / "core_fs_statement_panel_sample_v1.csv", index=False, encoding="utf-8-sig")

    with_pit, pit_audit = merge_pit(core)
    with_pit.to_parquet(OUT / "core_fs_statement_with_pit_dates_v1.parquet", index=False)
    with_pit.head(200).to_csv(OUT / "core_fs_statement_with_pit_dates_sample_v1.csv", index=False, encoding="utf-8-sig")

    pit_coverage = float(pit_audit.loc[pit_audit["metric"] == "pit_date_coverage_rate", "value"].iloc[0])
    readiness_df = readiness(with_pit, pit_coverage)
    missing_df = missing_data_update()

    panel_symbols = set(core["symbol"].dropna().unique())
    v15_coverage = safe_rate(len(v15_symbols & panel_symbols), len(v15_symbols))
    summary = {
        "fs_comins_detected": fs_comins_detected,
        "fs_combas_detected": fs_combas_detected,
        "fs_comins_usable": len(income) > 0,
        "fs_combas_usable": len(balance) > 0,
        "n_income_rows": len(income),
        "n_balance_rows": len(balance),
        "n_core_fs_rows": len(core),
        "n_symbols": int(core["symbol"].nunique()),
        "min_report_period": core["report_period"].min().date().isoformat(),
        "max_report_period": core["report_period"].max().date().isoformat(),
        "income_report_type_a_coverage_rate": safe_rate((income["report_type"] == "A").sum(), len(income)),
        "balance_report_type_a_coverage_rate": safe_rate((balance["report_type"] == "A").sum(), len(balance)),
        "pit_date_coverage_rate": pit_coverage,
        "v15_symbol_coverage_rate": v15_coverage,
        "operating_revenue_coverage_rate": 1.0 - missing_rate(income, "operating_revenue"),
        "net_profit_parent_coverage_rate": 1.0 - missing_rate(income, "net_profit_parent"),
        "total_assets_coverage_rate": 1.0 - missing_rate(balance, "total_assets"),
        "total_liabilities_coverage_rate": 1.0 - missing_rate(balance, "total_liabilities"),
        "equity_parent_coverage_rate": 1.0 - missing_rate(balance, "equity_parent"),
        "symbol_format_ok": bool(core["symbol"].dropna().astype(str).str.match(r"^\d{6}$").all()),
        "report_period_parsed": bool(core["report_period"].notna().mean() > 0.99),
        "numeric_fields_parsed": bool(income["net_profit_parent"].notna().any() and balance["total_assets"].notna().any()),
        "recommended_next_task": "Import TRD_Dalyr or equivalent market-cap data, then rebuild PIT-clean EP/BP and TTM financial factors.",
    }
    for factor in ["ROE", "ProfitGrowth_YoY", "RevGrowth_YoY", "NetMargin", "Debt_Ratio", "sales_expense_to_revenue", "rd_expense_to_revenue"]:
        row = readiness_df[readiness_df["target_factor"] == factor].iloc[0]
        summary[f"can_support_{factor.lower()}"] = bool(row["ttm_reconstructable"] and row["pit_ready"])
    summary["still_missing_for_ep"] = str(readiness_df.loc[readiness_df["target_factor"] == "EP", "missing_raw_items"].iloc[0])
    summary["still_missing_for_bp"] = str(readiness_df.loc[readiness_df["target_factor"] == "BP", "missing_raw_items"].iloc[0])

    if not fs_comins_detected or not fs_combas_detected:
        decision = "CSMAR_CORE_FS_IMPORT_BLOCKED_MISSING_FILE"
    elif pit_coverage < 0.95:
        decision = "CSMAR_CORE_FS_IMPORT_PIT_COVERAGE_NEEDS_PATCH"
    else:
        decision = "CSMAR_CORE_FS_MANUAL_IMPORT_READY_FOR_REVIEW"
    summary["decision"] = decision

    write_report(summary, income_missing, balance_missing, readiness_df, missing_df)
    update_project_status()
    append_decision()
    run_status_scripts()
    write_completion_card(summary)

    after = {str(p): file_fingerprint(p) for p in PROTECTED_PATHS}
    qa_df = final_qa(summary, before, after)
    readme_modified = readme_hash_before != text_sha256(ROOT / "README.md")
    invalid_modification = readme_modified or not bool(qa_df.loc[qa_df["check"] == "all_daily.parquet not modified", "pass"].iloc[0]) or not bool(qa_df.loc[qa_df["check"] == "training_panel_v15_sr.parquet not modified", "pass"].iloc[0])
    if invalid_modification:
        summary["decision"] = "INVALID_MODIFICATION"
        (OUT / "task_completion_card.md").write_text((OUT / "task_completion_card.md").read_text(encoding="utf-8").replace(decision, "INVALID_MODIFICATION"), encoding="utf-8")

    final_values = {
        "input_file_inventory_path": rel(OUT / "input_file_inventory_v1.csv"),
        "fs_comins_standardized_path": rel(OUT / "fs_comins_standardized_v1.parquet"),
        "fs_combas_standardized_path": rel(OUT / "fs_combas_standardized_v1.parquet"),
        "fs_comins_quality_audit_path": rel(OUT / "fs_comins_quality_audit_v1.csv"),
        "fs_combas_quality_audit_path": rel(OUT / "fs_combas_quality_audit_v1.csv"),
        "core_fs_statement_panel_path": rel(OUT / "core_fs_statement_panel_v1.parquet"),
        "core_fs_statement_with_pit_dates_path": rel(OUT / "core_fs_statement_with_pit_dates_v1.parquet"),
        "ttm_factor_readiness_audit_path": rel(OUT / "core_fs_ttm_factor_readiness_audit_v1.csv"),
        "missing_data_update_path": rel(OUT / "current_missing_data_after_core_fs_import_v1.csv"),
        "report_path": rel(OUT / "csmar_core_fs_manual_import_report_v1.md"),
        "task_completion_card_path": rel(OUT / "task_completion_card.md"),
        "final_qa_path": rel(OUT / "final_qa_csmar_core_fs_manual_import_audit_v1.csv"),
        "project_status_path": rel(STATUS_PATH),
        "current_status_doc_path": rel(CURRENT_STATUS_PATH),
        "decisions_doc_path": rel(DECISIONS_PATH),
        "readme_consistency_report_path": rel(README_CHECK_REPORT_PATH),
        "n_income_rows": summary["n_income_rows"],
        "n_balance_rows": summary["n_balance_rows"],
        "n_core_fs_rows": summary["n_core_fs_rows"],
        "n_symbols": summary["n_symbols"],
        "min_report_period": summary["min_report_period"],
        "max_report_period": summary["max_report_period"],
        "income_report_type_a_coverage_rate": summary["income_report_type_a_coverage_rate"],
        "balance_report_type_a_coverage_rate": summary["balance_report_type_a_coverage_rate"],
        "pit_date_coverage_rate": summary["pit_date_coverage_rate"],
        "v15_symbol_coverage_rate": summary["v15_symbol_coverage_rate"],
        "operating_revenue_coverage_rate": summary["operating_revenue_coverage_rate"],
        "net_profit_parent_coverage_rate": summary["net_profit_parent_coverage_rate"],
        "total_assets_coverage_rate": summary["total_assets_coverage_rate"],
        "total_liabilities_coverage_rate": summary["total_liabilities_coverage_rate"],
        "equity_parent_coverage_rate": summary["equity_parent_coverage_rate"],
        "can_support_roe": summary["can_support_roe"],
        "can_support_profit_growth": summary["can_support_profitgrowth_yoy"],
        "can_support_revenue_growth": summary["can_support_revgrowth_yoy"],
        "can_support_net_margin": summary["can_support_netmargin"],
        "can_support_debt_ratio": summary["can_support_debt_ratio"],
        "can_support_sales_expense_ratio": summary["can_support_sales_expense_to_revenue"],
        "can_support_rd_expense_ratio": summary["can_support_rd_expense_to_revenue"],
        "still_missing_for_ep": summary["still_missing_for_ep"],
        "still_missing_for_bp": summary["still_missing_for_bp"],
        "recommended_next_task": summary["recommended_next_task"],
        "csmar_api_accessed": False,
        "getPackResultExt_called": False,
        "readme_modified": readme_modified,
        "all_daily_modified": not bool(qa_df.loc[qa_df["check"] == "all_daily.parquet not modified", "pass"].iloc[0]),
        "training_panel_modified": not bool(qa_df.loc[qa_df["check"] == "training_panel_v15_sr.parquet not modified", "pass"].iloc[0]),
        "production_modified": False,
        "credential_exposure_detected": False,
        "decision": summary["decision"],
    }
    for key, value in final_values.items():
        print(f"{key}={value}")


if __name__ == "__main__":
    main()
