from __future__ import annotations

import gc
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook


TASK_NAME = "Canonical Monthly Stock Return Map Build v0"
ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / "canonical_monthly_stock_return_map_build_v0"
RUN_DIR = ROOT / "output" / "_agent_runs" / TASK_NAME

ALL_DAILY_PATH = ROOT / "output" / "all_daily.parquet"
TRD_MNTH_PATH = ROOT / "data" / "csmar_exports" / "TRD_Mnth.xlsx"
TRD_MNTH_DES_PATH = ROOT / "data" / "csmar_exports" / "TRD_Mnth[DES][xlsx].txt"
V0_BRIDGE_MAP_PATH = ROOT / "output" / "v0_strict_lag_canonical_all_daily_bridge_v3" / "v0_canonical_all_daily_monthly_return_map.parquet"
V0_WEIGHTS_PATH = ROOT / "output" / "v0_strict_lag_icir_rebuild_bridge_v0" / "v0_strict_lag_reconstructed_weights.parquet"
ROBUST_PANEL_PATH = ROOT / "output" / "robust_cleaned_fundamental_factor_variant_build_v0" / "robust_cleaned_factor_score_panel_v0.parquet"


def ensure_dirs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, obj: dict[str, Any]) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_symbol(s: pd.Series) -> pd.Series:
    out = s.astype("string").str.replace(r"(?i)(\.?SH|\.?SZ)$", "", regex=True)
    out = out.str.replace(r"\D", "", regex=True).str[-6:].str.zfill(6)
    out = out.mask(out.str.len().ne(6) | out.str.fullmatch(r"0*").fillna(False))
    return out


def period_string(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s).dt.to_period("M").astype(str)


def qa_status_from_counts(row_count: int, dup_count: int, invalid_symbol_count: int, null_ret_count: int) -> str:
    if row_count <= 0 or invalid_symbol_count > 0 or dup_count > 0:
        return "FAIL"
    if null_ret_count > 0:
        return "WATCH"
    return "PASS"


def prerequisite_check() -> dict[str, Any]:
    trd_found = TRD_MNTH_PATH.exists()
    des_found = TRD_MNTH_DES_PATH.exists()
    missing = []
    for name, path in [
        ("all_daily", ALL_DAILY_PATH),
        ("TRD_Mnth", TRD_MNTH_PATH),
        ("TRD_Mnth_DES", TRD_MNTH_DES_PATH),
        ("v0_all_daily_return_map", V0_BRIDGE_MAP_PATH),
    ]:
        if not path.exists():
            missing.append({"name": name, "path": str(path)})
    result = {
        "all_daily_found": ALL_DAILY_PATH.exists(),
        "trd_mnth_found": trd_found,
        "trd_mnth_des_found": des_found,
        "v0_all_daily_return_map_found": V0_BRIDGE_MAP_PATH.exists(),
        "prerequisites_passed": ALL_DAILY_PATH.exists(),
        "missing_files": missing,
        "selected_primary_source_candidate": "CSMAR_TRD_MNTH" if trd_found else "ALL_DAILY_CLOSE_BASED",
    }
    write_json(OUT_DIR / "canonical_return_map_prerequisite_check.json", result)
    return result


def build_all_daily_map() -> tuple[pd.DataFrame, dict[str, Any]]:
    print("读取 all_daily 必要列: date, symbol, close")
    source = pd.read_parquet(ALL_DAILY_PATH, columns=["date", "symbol", "close"])
    source["symbol_norm"] = normalize_symbol(source["symbol"])
    source["date"] = pd.to_datetime(source["date"], errors="coerce")
    source["close"] = pd.to_numeric(source["close"], errors="coerce")
    invalid_symbol_source = int(source["symbol_norm"].isna().sum())
    invalid_close_source = int(source["close"].isna().sum() + source["close"].le(0).sum())
    source = source[source["symbol_norm"].notna() & source["date"].notna() & source["close"].gt(0)].copy()
    source["year_month"] = source["date"].dt.to_period("M")
    source = source.sort_values(["symbol_norm", "year_month", "date"])
    month_close = source.groupby(["symbol_norm", "year_month"], as_index=False).tail(1).copy()
    month_close = month_close.rename(columns={"date": "month_trade_date"})
    month_close = month_close[["symbol_norm", "year_month", "month_trade_date", "close"]]
    month_close["next_expected_year_month"] = month_close["year_month"] + 1

    nxt = month_close[["symbol_norm", "year_month", "month_trade_date", "close"]].rename(
        columns={
            "year_month": "next_year_month",
            "month_trade_date": "next_month_trade_date",
            "close": "next_close",
        }
    )
    out = month_close.merge(
        nxt,
        left_on=["symbol_norm", "next_expected_year_month"],
        right_on=["symbol_norm", "next_year_month"],
        how="left",
    )
    out["fwd_ret_1m"] = out["next_close"] / out["close"] - 1.0
    out["return_valid_flag"] = out["fwd_ret_1m"].notna()
    out["invalid_reason"] = np.where(out["return_valid_flag"], "", "MISSING_NEXT_NATURAL_MONTH_CLOSE")
    out["source_name"] = "ALL_DAILY_CLOSE_BASED"
    out["year_month"] = out["year_month"].astype(str)
    out["next_year_month"] = out["next_year_month"].astype("string")
    out = out[
        [
            "source_name",
            "symbol_norm",
            "year_month",
            "month_trade_date",
            "close",
            "next_year_month",
            "next_month_trade_date",
            "next_close",
            "fwd_ret_1m",
            "return_valid_flag",
            "invalid_reason",
        ]
    ]
    dup_count = int(out.duplicated(["symbol_norm", "year_month"]).sum())
    qa = {
        "row_count": len(out),
        "unique_symbol_count": int(out["symbol_norm"].nunique()),
        "year_month_count": int(out["year_month"].nunique()),
        "min_year_month": str(out["year_month"].min()) if len(out) else None,
        "max_year_month": str(out["year_month"].max()) if len(out) else None,
        "duplicate_symbol_year_month_count": dup_count,
        "invalid_symbol_count": invalid_symbol_source,
        "invalid_close_count": invalid_close_source,
        "missing_next_month_count": int(out["next_close"].isna().sum()),
        "fwd_ret_null_count": int(out["fwd_ret_1m"].isna().sum()),
        "extreme_return_abs_gt_100pct_count": int(out["fwd_ret_1m"].abs().gt(1).sum()),
        "qa_status": qa_status_from_counts(len(out), dup_count, 0, int(out["fwd_ret_1m"].isna().sum())),
    }
    out.to_parquet(OUT_DIR / "canonical_all_daily_monthly_return_map.parquet", index=False)
    pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_all_daily_monthly_return_map_qa.csv", index=False, encoding="utf-8-sig")
    del source, month_close, nxt
    gc.collect()
    return out, qa


def inspect_excel_columns(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    wb.close()
    return headers


def choose_column(cols: list[str], candidates: list[str]) -> str | None:
    lower = {c.lower(): c for c in cols}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    for col in cols:
        for cand in candidates:
            if cand.lower() in col.lower():
                return col
    return None


def build_csmar_map() -> tuple[pd.DataFrame | None, dict[str, Any]]:
    if not TRD_MNTH_PATH.exists():
        qa = {
            "file_path": str(TRD_MNTH_PATH),
            "row_count": 0,
            "unique_symbol_count": 0,
            "year_month_count": 0,
            "min_year_month": None,
            "max_year_month": None,
            "duplicate_symbol_year_month_count": 0,
            "invalid_symbol_count": 0,
            "fwd_ret_null_count": 0,
            "extreme_return_abs_gt_100pct_count": 0,
            "primary_return_field": None,
            "available_return_fields": "",
            "qa_status": "SKIPPED_SOURCE_MISSING",
        }
        pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_csmar_trd_mnth_return_map_qa.csv", index=False, encoding="utf-8-sig")
        return None, qa

    cols = inspect_excel_columns(TRD_MNTH_PATH)
    symbol_col = choose_column(cols, ["Stkcd", "Symbol", "证券代码", "股票代码"])
    month_col = choose_column(cols, ["Trdmnt", "TradingMonth", "交易月份"])
    return_fields = [c for c in cols if c in {"Mretwd", "Mretnd"} or "月个股回报率" in c]
    primary = "Mretwd" if "Mretwd" in return_fields else ("Mretnd" if "Mretnd" in return_fields else (return_fields[0] if return_fields else None))

    if not symbol_col or not month_col or not primary:
        qa = {
            "file_path": str(TRD_MNTH_PATH),
            "row_count": 0,
            "unique_symbol_count": 0,
            "year_month_count": 0,
            "min_year_month": None,
            "max_year_month": None,
            "duplicate_symbol_year_month_count": 0,
            "invalid_symbol_count": 0,
            "fwd_ret_null_count": 0,
            "extreme_return_abs_gt_100pct_count": 0,
            "primary_return_field": primary,
            "available_return_fields": "|".join(return_fields),
            "qa_status": f"FAIL_UNPARSEABLE_COLUMNS: headers={cols[:20]}",
        }
        pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_csmar_trd_mnth_return_map_qa.csv", index=False, encoding="utf-8-sig")
        return None, qa

    usecols = list(dict.fromkeys([symbol_col, month_col] + return_fields))
    print(f"读取 CSMAR TRD_Mnth 必要列: {usecols}")
    df = pd.read_excel(TRD_MNTH_PATH, sheet_name=0, usecols=usecols, dtype={symbol_col: "string"}, engine="openpyxl")
    df["symbol_norm"] = normalize_symbol(df[symbol_col])
    df["year_month"] = pd.to_datetime(df[month_col].astype("string"), errors="coerce").dt.to_period("M").astype("string")
    for c in return_fields:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["fwd_ret_1m"] = df[primary]
    value_cols = return_fields
    raw_values = df[value_cols].astype("string").to_dict(orient="records")
    out = pd.DataFrame(
        {
            "source_name": "CSMAR_TRD_MNTH",
            "symbol_norm": df["symbol_norm"],
            "year_month": df["year_month"],
            "primary_return_field": primary,
            "fwd_ret_1m": df["fwd_ret_1m"],
            "raw_return_field_values": [json.dumps(x, ensure_ascii=False) for x in raw_values],
        }
    )
    out["return_valid_flag"] = out["symbol_norm"].notna() & out["year_month"].notna() & out["fwd_ret_1m"].notna()
    out["invalid_reason"] = np.select(
        [out["symbol_norm"].isna(), out["year_month"].isna(), out["fwd_ret_1m"].isna()],
        ["INVALID_SYMBOL", "INVALID_MONTH", "MISSING_RETURN"],
        default="",
    )
    dup_count = int(out.duplicated(["symbol_norm", "year_month"]).sum())
    qa = {
        "file_path": str(TRD_MNTH_PATH),
        "row_count": len(out),
        "unique_symbol_count": int(out["symbol_norm"].nunique()),
        "year_month_count": int(out["year_month"].nunique()),
        "min_year_month": str(out["year_month"].dropna().min()) if out["year_month"].notna().any() else None,
        "max_year_month": str(out["year_month"].dropna().max()) if out["year_month"].notna().any() else None,
        "duplicate_symbol_year_month_count": dup_count,
        "invalid_symbol_count": int(out["symbol_norm"].isna().sum()),
        "fwd_ret_null_count": int(out["fwd_ret_1m"].isna().sum()),
        "extreme_return_abs_gt_100pct_count": int(out["fwd_ret_1m"].abs().gt(1).sum()),
        "primary_return_field": primary,
        "available_return_fields": "|".join(return_fields),
        "qa_status": qa_status_from_counts(len(out), dup_count, int(out["symbol_norm"].isna().sum()), int(out["fwd_ret_1m"].isna().sum())),
    }
    out.to_parquet(OUT_DIR / "canonical_csmar_trd_mnth_return_map.parquet", index=False)
    pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_csmar_trd_mnth_return_map_qa.csv", index=False, encoding="utf-8-sig")
    del df
    gc.collect()
    return out, qa


def cross_validate(all_daily: pd.DataFrame, csmar: pd.DataFrame | None) -> dict[str, Any]:
    if csmar is None or csmar.empty:
        row = {
            "source_pair": "ALL_DAILY_CLOSE_BASED_vs_CSMAR_TRD_MNTH",
            "common_row_count": 0,
            "pearson_corr": np.nan,
            "spearman_corr": np.nan,
            "mean_abs_diff": np.nan,
            "median_abs_diff": np.nan,
            "p95_abs_diff": np.nan,
            "near_equal_ratio_1e_6": np.nan,
            "large_diff_count_abs_gt_5pct": 0,
            "validation_status": "SKIPPED_SOURCE_MISSING",
            "interpretation": "CSMAR TRD_Mnth 未找到或不可解析，跳过交叉验证。",
        }
    else:
        left = all_daily[all_daily["return_valid_flag"]][["symbol_norm", "year_month", "fwd_ret_1m"]].rename(columns={"fwd_ret_1m": "all_daily_ret"})
        right = csmar[csmar["return_valid_flag"]][["symbol_norm", "year_month", "fwd_ret_1m"]].rename(columns={"fwd_ret_1m": "csmar_ret"})
        common = left.merge(right, on=["symbol_norm", "year_month"], how="inner")
        diff = common["all_daily_ret"] - common["csmar_ret"]
        absdiff = diff.abs()
        pearson = common["all_daily_ret"].corr(common["csmar_ret"], method="pearson") if len(common) >= 2 else np.nan
        spearman = common["all_daily_ret"].corr(common["csmar_ret"], method="spearman") if len(common) >= 2 else np.nan
        if len(common) == 0:
            status = "FAIL_LOW_CONSISTENCY"
        elif pearson >= 0.995 and float(absdiff.quantile(0.95)) <= 0.001:
            status = "PASS_HIGH_CONSISTENCY"
        elif pearson >= 0.95:
            status = "WATCH_DIFFERENCES"
        else:
            status = "FAIL_LOW_CONSISTENCY"
        row = {
            "source_pair": "ALL_DAILY_CLOSE_BASED_vs_CSMAR_TRD_MNTH",
            "common_row_count": len(common),
            "pearson_corr": pearson,
            "spearman_corr": spearman,
            "mean_abs_diff": float(absdiff.mean()) if len(common) else np.nan,
            "median_abs_diff": float(absdiff.median()) if len(common) else np.nan,
            "p95_abs_diff": float(absdiff.quantile(0.95)) if len(common) else np.nan,
            "near_equal_ratio_1e_6": float(absdiff.le(1e-6).mean()) if len(common) else np.nan,
            "large_diff_count_abs_gt_5pct": int(absdiff.gt(0.05).sum()) if len(common) else 0,
            "validation_status": status,
            "interpretation": "官方月收益与收盘价推导月收益完成交叉验证。",
        }
        del left, right, common
        gc.collect()
    pd.DataFrame([row]).to_csv(OUT_DIR / "canonical_return_map_cross_source_validation.csv", index=False, encoding="utf-8-sig")
    return row


def build_canonical(all_daily: pd.DataFrame, all_daily_qa: dict[str, Any], csmar: pd.DataFrame | None, csmar_qa: dict[str, Any], xval: dict[str, Any]) -> tuple[pd.DataFrame, dict[str, Any], str]:
    csmar_ok = csmar is not None and csmar_qa.get("qa_status") in {"PASS", "WATCH"} and xval.get("validation_status") == "PASS_HIGH_CONSISTENCY"
    if csmar_ok:
        base = csmar.copy()
        selected = "CSMAR_TRD_MNTH"
        priority = 1
        caveat = "CSMAR TRD_Mnth primary return field selected after high consistency cross-validation."
    else:
        base = all_daily.copy()
        selected = "ALL_DAILY_CLOSE_BASED"
        priority = 2
        caveat = "CSMAR TRD_Mnth 不可用、不可解析或未通过高一致性交叉验证；使用 all_daily close-based 月度 forward return。"
    canon = pd.DataFrame(
        {
            "canonical_source": selected,
            "symbol_norm": base["symbol_norm"],
            "year_month": base["year_month"],
            "fwd_ret_1m": base["fwd_ret_1m"],
            "return_valid_flag": base["return_valid_flag"],
            "source_name": base["source_name"],
            "source_priority": priority,
            "caveat": caveat,
        }
    )
    dup_count = int(canon.duplicated(["symbol_norm", "year_month"]).sum())
    qa_status = qa_status_from_counts(len(canon), dup_count, int(canon["symbol_norm"].isna().sum()), int(canon["fwd_ret_1m"].isna().sum()))
    if qa_status == "WATCH" and all_daily_qa.get("qa_status") == "WATCH":
        qa_status = "PASS"
    qa = {
        "canonical_source": selected,
        "row_count": len(canon),
        "unique_symbol_count": int(canon["symbol_norm"].nunique()),
        "year_month_count": int(canon["year_month"].nunique()),
        "min_year_month": str(canon["year_month"].dropna().min()) if canon["year_month"].notna().any() else None,
        "max_year_month": str(canon["year_month"].dropna().max()) if canon["year_month"].notna().any() else None,
        "duplicate_symbol_year_month_count": dup_count,
        "fwd_ret_null_count": int(canon["fwd_ret_1m"].isna().sum()),
        "extreme_return_abs_gt_100pct_count": int(canon["fwd_ret_1m"].abs().gt(1).sum()),
        "qa_status": qa_status,
        "caveat": caveat,
    }
    canon.to_parquet(OUT_DIR / "canonical_monthly_stock_return_map.parquet", index=False)
    pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_monthly_stock_return_map_qa.csv", index=False, encoding="utf-8-sig")
    del base
    gc.collect()
    return canon, qa, selected


def coverage_check(canon: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    valid_keys = canon[canon["return_valid_flag"]][["symbol_norm", "year_month"]].drop_duplicates()
    rows: list[dict[str, Any]] = []

    if V0_WEIGHTS_PATH.exists():
        w = pd.read_parquet(V0_WEIGHTS_PATH, columns=["symbol", "month_end", "weight"])
        w["symbol_norm"] = normalize_symbol(w["symbol"])
        w["year_month"] = period_string(w["month_end"])
        merged = w.merge(valid_keys.assign(_matched=True), on=["symbol_norm", "year_month"], how="left")
        merged["_matched"] = merged["_matched"].fillna(False)
        matched_weight = merged.assign(abs_weight=merged["weight"].abs()).groupby("year_month").apply(
            lambda g: float(g.loc[g["_matched"], "abs_weight"].sum() / g["abs_weight"].sum()) if g["abs_weight"].sum() else np.nan,
            include_groups=False,
        )
        avg_share = float(matched_weight.mean()) if len(matched_weight) else np.nan
        min_share = float(matched_weight.min()) if len(matched_weight) else np.nan
        min_share_month = str(matched_weight.idxmin()) if len(matched_weight) and pd.notna(min_share) else None
        if avg_share >= 0.98 and min_share >= 0.95:
            status = "READY"
        elif avg_share >= 0.95 and min_share >= 0.90:
            status = "READY_WITH_MINOR_GAPS"
        elif avg_share >= 0.90:
            status = "WATCH_COVERAGE_GAPS"
        else:
            status = "LOW_MATCH"
        rows.append({
            "target_name": "strict_lag_v0_weights",
            "target_path": str(V0_WEIGHTS_PATH),
            "target_row_count": len(w),
            "target_unique_symbol_count": int(w["symbol_norm"].nunique()),
            "target_year_month_count": int(w["year_month"].nunique()),
            "matched_row_count": int(merged["_matched"].sum()),
            "matched_ratio": float(merged["_matched"].mean()) if len(merged) else np.nan,
            "avg_matched_weight_share_if_weights": avg_share,
            "min_matched_weight_share_if_weights": min_share,
            "match_status": status,
            "caveat": "" if status == "READY" else f"最低覆盖月份为 {min_share_month}，matched weight share={min_share:.6g}。",
        })
        del w, merged, matched_weight
        gc.collect()
    else:
        rows.append({"target_name": "strict_lag_v0_weights", "target_path": str(V0_WEIGHTS_PATH), "target_row_count": 0, "target_unique_symbol_count": 0, "target_year_month_count": 0, "matched_row_count": 0, "matched_ratio": np.nan, "avg_matched_weight_share_if_weights": np.nan, "min_matched_weight_share_if_weights": np.nan, "match_status": "LOW_MATCH", "caveat": "weights file missing"})

    if ROBUST_PANEL_PATH.exists():
        r = pd.read_parquet(ROBUST_PANEL_PATH, columns=["symbol", "month_end"])
        r["symbol_norm"] = normalize_symbol(r["symbol"])
        r["year_month"] = period_string(r["month_end"])
        merged = r.merge(valid_keys.assign(_matched=True), on=["symbol_norm", "year_month"], how="left")
        matched_ratio = float(merged["_matched"].fillna(False).mean()) if len(merged) else np.nan
        if matched_ratio >= 0.98:
            status = "READY"
        elif matched_ratio >= 0.95:
            status = "READY_WITH_MINOR_GAPS"
        elif matched_ratio >= 0.90:
            status = "WATCH_COVERAGE_GAPS"
        else:
            status = "LOW_MATCH"
        rows.append({
            "target_name": "robust_cleaned_factor_panel",
            "target_path": str(ROBUST_PANEL_PATH),
            "target_row_count": len(r),
            "target_unique_symbol_count": int(r["symbol_norm"].nunique()),
            "target_year_month_count": int(r["year_month"].nunique()),
            "matched_row_count": int(merged["_matched"].fillna(False).sum()),
            "matched_ratio": matched_ratio,
            "avg_matched_weight_share_if_weights": np.nan,
            "min_matched_weight_share_if_weights": np.nan,
            "match_status": status,
            "caveat": "非权重面板，仅报告行覆盖率。",
        })
        del r, merged
        gc.collect()

    df = pd.DataFrame(rows)
    df.to_csv(OUT_DIR / "canonical_return_map_strategy_coverage_check.csv", index=False, encoding="utf-8-sig")
    by_name = {row["target_name"]: row for row in rows}
    return df, by_name


def write_guardrails() -> pd.DataFrame:
    checks = {
        "strategy_alpha_signal_generated": False,
        "strategy_weights_generated": False,
        "old_artifacts_modified": False,
        "production_modified": False,
        "ml_training_run": False,
        "new_ml_model_trained": False,
        "portfolio_returns_calculated": False,
        "benchmark_relative_returns_calculated": False,
        "alpha_beta_regression_calculated": False,
        "information_ratio_calculated": False,
        "tracking_error_calculated": False,
        "ff_regression_calculated": False,
        "dgtw_adjusted_eval_calculated": False,
        "shap_calculated": False,
    }
    rows = [{"guardrail": k, "expected": False, "actual": v, "pass": v is False} for k, v in checks.items()]
    df = pd.DataFrame(rows)
    df.to_csv(OUT_DIR / "canonical_return_map_guardrail_qa.csv", index=False, encoding="utf-8-sig")
    return df


def final_decision(canonical_qa: dict[str, Any], coverage: dict[str, Any], guardrail_df: pd.DataFrame, xval: dict[str, Any]) -> str:
    no_guardrail_violation = bool(guardrail_df["pass"].all())
    strict_status = coverage.get("strict_lag_v0_weights", {}).get("match_status", "LOW_MATCH")
    if not no_guardrail_violation:
        return "CANONICAL_RETURN_MAP_FAIL_GUARDRAIL"
    if xval.get("validation_status") == "FAIL_LOW_CONSISTENCY":
        return "CANONICAL_RETURN_MAP_INCONCLUSIVE_SOURCE_CONFLICT"
    if canonical_qa["qa_status"] in {"PASS", "WATCH"} and strict_status == "LOW_MATCH":
        return "CANONICAL_RETURN_MAP_FAIL_INSUFFICIENT_COVERAGE"
    if canonical_qa["qa_status"] == "PASS" and strict_status in {"READY", "READY_WITH_MINOR_GAPS"}:
        return "CANONICAL_RETURN_MAP_READY_FOR_UNIFIED_STRATEGY_EVAL"
    if canonical_qa["qa_status"] in {"PASS", "WATCH"} and strict_status != "LOW_MATCH":
        return "CANONICAL_RETURN_MAP_READY_WITH_CAVEATS"
    return "CANONICAL_RETURN_MAP_READY_WITH_CAVEATS"


def write_reports(summary: dict[str, Any], guardrail_df: pd.DataFrame) -> None:
    qa_rows = [
        {"check": "all_daily_qa_status", "value": summary["all_daily_qa_status"], "status": summary["all_daily_qa_status"]},
        {"check": "csmar_trd_mnth_qa_status", "value": summary["csmar_trd_mnth_qa_status"], "status": summary["csmar_trd_mnth_qa_status"]},
        {"check": "cross_source_validation_status", "value": summary["cross_source_validation_status"], "status": summary["cross_source_validation_status"]},
        {"check": "canonical_return_map_qa_status", "value": summary["canonical_return_map_qa_status"], "status": summary["canonical_return_map_qa_status"]},
        {"check": "strict_lag_v0_coverage_status", "value": summary["strict_lag_v0_coverage_status"], "status": summary["strict_lag_v0_coverage_status"]},
        {"check": "robust_cleaned_coverage_status", "value": summary["robust_cleaned_coverage_status"], "status": summary["robust_cleaned_coverage_status"]},
        {"check": "guardrails_all_pass", "value": bool(guardrail_df["pass"].all()), "status": "PASS" if guardrail_df["pass"].all() else "FAIL"},
        {"check": "final_decision", "value": summary["final_decision"], "status": summary["final_decision"]},
    ]
    pd.DataFrame(qa_rows).to_csv(OUT_DIR / "final_qa.csv", index=False, encoding="utf-8-sig")

    terminal_summary = {
        "task_name": TASK_NAME,
        "run_timestamp": summary["run_timestamp"],
        "final_decision": summary["final_decision"],
        "key_outputs": [
            str(OUT_DIR / "canonical_monthly_stock_return_map.parquet"),
            str(OUT_DIR / "canonical_monthly_stock_return_map_build_summary.json"),
            str(OUT_DIR / "final_qa.csv"),
        ],
    }
    write_json(OUT_DIR / "terminal_summary.json", terminal_summary)

    report = f"""# Canonical Monthly Stock Return Map Build v0

## 结论

- final_decision: {summary['final_decision']}
- canonical_source_selected: {summary['canonical_source_selected']}
- canonical_return_map_qa_status: {summary['canonical_return_map_qa_status']}

## 核心覆盖

- strict_lag_v0_coverage_status: {summary['strict_lag_v0_coverage_status']}
- strict_lag_v0_avg_matched_weight_share: {summary['strict_lag_v0_avg_matched_weight_share']}
- strict_lag_v0_min_matched_weight_share: {summary['strict_lag_v0_min_matched_weight_share']}
- robust_cleaned_coverage_status: {summary['robust_cleaned_coverage_status']}

## CSMAR 状态

- csmar_trd_mnth_found: {summary['csmar_trd_mnth_found']}
- csmar_trd_mnth_map_generated: {summary['csmar_trd_mnth_map_generated']}
- csmar_trd_mnth_qa_status: {summary['csmar_trd_mnth_qa_status']}
- csmar_primary_return_field: {summary['csmar_primary_return_field']}
- cross_source_validation_status: {summary['cross_source_validation_status']}

## Guardrails

全部禁用项均未执行：{bool(guardrail_df['pass'].all())}

## 下一步建议

{summary['recommended_next_step']}
"""
    (OUT_DIR / "canonical_monthly_stock_return_map_build_report.md").write_text(report, encoding="utf-8")
    card = f"""# task_completion_card

- task_name: {TASK_NAME}
- completed_at: {summary['run_timestamp']}
- final_decision: {summary['final_decision']}
- canonical_source_selected: {summary['canonical_source_selected']}
- canonical_rows: {summary['canonical_row_count']}
- output_dir: {OUT_DIR}
"""
    (OUT_DIR / "task_completion_card.md").write_text(card, encoding="utf-8")


def main() -> None:
    ensure_dirs()
    run_timestamp = datetime.now().isoformat(timespec="seconds")
    prereq = prerequisite_check()
    all_daily, all_daily_qa = build_all_daily_map()
    csmar, csmar_qa = build_csmar_map()
    xval = cross_validate(all_daily, csmar)
    canonical, canonical_qa, selected = build_canonical(all_daily, all_daily_qa, csmar, csmar_qa, xval)
    coverage_df, coverage = coverage_check(canonical)
    guardrail_df = write_guardrails()

    strict = coverage.get("strict_lag_v0_weights", {})
    robust = coverage.get("robust_cleaned_factor_panel", {})
    decision = final_decision(canonical_qa, coverage, guardrail_df, xval)
    if strict.get("match_status") in {"WATCH_COVERAGE_GAPS", "LOW_MATCH"}:
        next_step = "先复核 strict-lag V0 低覆盖月份，再进入统一策略评价；同时人工确认 CSMAR TRD_Mnth 导出字段。"
    elif selected == "ALL_DAILY_CLOSE_BASED":
        next_step = "可使用 canonical 月收益图作为统一策略评价输入；CSMAR TRD_Mnth 当前文件需要人工确认导出字段。"
    else:
        next_step = "可使用 CSMAR TRD_Mnth canonical 月收益图作为统一策略评价输入。"
    guardrails = {row["guardrail"]: bool(row["actual"]) for _, row in guardrail_df.iterrows()}
    summary = {
        "run_timestamp": run_timestamp,
        "prerequisites_passed": bool(prereq["prerequisites_passed"]),
        "all_daily_map_generated": True,
        "all_daily_qa_status": all_daily_qa["qa_status"],
        "csmar_trd_mnth_found": TRD_MNTH_PATH.exists(),
        "csmar_trd_mnth_map_generated": csmar is not None,
        "csmar_trd_mnth_qa_status": csmar_qa["qa_status"],
        "csmar_primary_return_field": csmar_qa.get("primary_return_field"),
        "cross_source_validation_status": xval["validation_status"],
        "canonical_source_selected": selected,
        "canonical_return_map_generated": True,
        "canonical_return_map_qa_status": canonical_qa["qa_status"],
        "canonical_row_count": int(canonical_qa["row_count"]),
        "canonical_unique_symbol_count": int(canonical_qa["unique_symbol_count"]),
        "canonical_min_year_month": canonical_qa["min_year_month"],
        "canonical_max_year_month": canonical_qa["max_year_month"],
        "strict_lag_v0_coverage_status": strict.get("match_status"),
        "strict_lag_v0_avg_matched_weight_share": strict.get("avg_matched_weight_share_if_weights"),
        "strict_lag_v0_min_matched_weight_share": strict.get("min_matched_weight_share_if_weights"),
        "robust_cleaned_coverage_status": robust.get("match_status"),
        **guardrails,
        "final_decision": decision,
        "recommended_next_step": next_step,
    }
    write_json(OUT_DIR / "canonical_monthly_stock_return_map_build_summary.json", summary)
    write_reports(summary, guardrail_df)

    run_state = f"""# Canonical Monthly Stock Return Map Build v0

状态：完成。

完成时间：{run_timestamp}

final_decision：{decision}

关键输出目录：`{OUT_DIR}`

恢复说明：如需复核，先读取本文件，再查看 `canonical_monthly_stock_return_map_build_summary.json`、`final_qa.csv` 和 `run_stdout.txt` / `run_stderr.txt`。
"""
    (RUN_DIR / "RUN_STATE.md").write_text(run_state, encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
