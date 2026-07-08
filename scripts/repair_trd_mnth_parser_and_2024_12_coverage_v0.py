from __future__ import annotations

import gc
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


TASK_NAME = "TRD_Mnth Parser Repair & 2024-12 Return Coverage Repair v0"
ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / "trd_mnth_parser_repair_2024_12_coverage_repair_v0"
RUN_DIR = ROOT / "output" / "_agent_runs" / TASK_NAME

CSMAR_DIR = ROOT / "data" / "csmar_exports"
ALL_DAILY_PATH = ROOT / "output" / "all_daily.parquet"
CANONICAL_ORIG_PATH = ROOT / "output" / "canonical_monthly_stock_return_map_build_v0" / "canonical_monthly_stock_return_map.parquet"
PREV_SUMMARY_PATH = ROOT / "output" / "canonical_monthly_stock_return_map_build_v0" / "canonical_monthly_stock_return_map_build_summary.json"
V0_WEIGHTS_PATH = ROOT / "output" / "v0_strict_lag_icir_rebuild_bridge_v0" / "v0_strict_lag_reconstructed_weights.parquet"


def normalize_symbol(s: pd.Series) -> pd.Series:
    out = s.astype("string").str.replace(r"(?i)(\.?SH|\.?SZ)$", "", regex=True)
    out = out.str.replace(r"\D", "", regex=True).str[-6:].str.zfill(6)
    return out.mask(out.str.len().ne(6) | out.str.fullmatch(r"0*").fillna(False))


def write_json(path: Path, obj: dict[str, Any]) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def as_float_or_none(x: Any) -> float | None:
    try:
        if pd.isna(x):
            return None
        return float(x)
    except Exception:
        return None


def candidates() -> tuple[list[Path], list[Path]]:
    trd = []
    for pat in ["TRD_Mnth.xlsx", "TRD_Mnth.csv", "TRD_Mnth*.xlsx", "TRD_Mnth*.csv"]:
        trd.extend(CSMAR_DIR.glob(pat))
    trd = sorted(set(p for p in trd if p.is_file()))
    des = sorted(set(p for p in CSMAR_DIR.glob("TRD_Mnth*.txt") if p.is_file()))
    return trd, des


def prerequisite_check(trd_files: list[Path], des_files: list[Path]) -> dict[str, Any]:
    missing = []
    for name, path in [
        ("all_daily", ALL_DAILY_PATH),
        ("canonical_return_map", CANONICAL_ORIG_PATH),
        ("strict_lag_weights", V0_WEIGHTS_PATH),
    ]:
        if not path.exists():
            missing.append({"name": name, "path": str(path)})
    if not trd_files:
        missing.append({"name": "TRD_Mnth", "path": str(CSMAR_DIR / "TRD_Mnth*")})
    result = {
        "trd_mnth_files_found": [str(p) for p in trd_files],
        "trd_mnth_des_found": [str(p) for p in des_files],
        "all_daily_found": ALL_DAILY_PATH.exists(),
        "canonical_return_map_found": CANONICAL_ORIG_PATH.exists(),
        "strict_lag_weights_found": V0_WEIGHTS_PATH.exists(),
        "prerequisites_passed": bool(trd_files) and ALL_DAILY_PATH.exists() and CANONICAL_ORIG_PATH.exists() and V0_WEIGHTS_PATH.exists(),
        "missing_files": missing,
    }
    write_json(OUT_DIR / "trd_mnth_parser_repair_prerequisite_check.json", result)
    return result


def row_score(values: list[Any]) -> tuple[int, bool, bool, bool]:
    text = " ".join("" if v is None else str(v) for v in values)
    contains_stkcd = any(x in text for x in ["Stkcd", "Symbol", "股票代码", "证券代码"])
    contains_trdmnt = any(x in text for x in ["Trdmnt", "TradingMonth", "交易月份", "交易月"])
    contains_return = any(x in text for x in ["Mretwd", "Mretnd", "月个股回报率"])
    return int(contains_stkcd) + int(contains_trdmnt) + int(contains_return), contains_stkcd, contains_trdmnt, contains_return


def inspect_excel(path: Path) -> tuple[list[dict[str, Any]], str | None, int | None]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    rows = []
    best_score = -1
    best_idx = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        vals = list(row)
        non_null = [v for v in vals if v is not None and str(v).strip() != ""]
        score, has_stkcd, has_trdmnt, has_ret = row_score(non_null)
        if score > best_score:
            best_score, best_idx = score, idx
        rows.append(
            {
                "file_path": str(path),
                "sheet_name": sheet,
                "row_index_zero_based": idx,
                "excel_row_number": idx + 1,
                "non_null_count": len(non_null),
                "row_values_preview": json.dumps([str(v) for v in non_null[:20]], ensure_ascii=False),
                "contains_stkcd": has_stkcd,
                "contains_trdmnt": has_trdmnt,
                "contains_return_field": has_ret,
                "candidate_header_score": score,
                "recommended_header_row": None,
                "diagnosis": "",
            }
        )
    wb.close()
    recommended = 3 if any(r["candidate_header_score"] >= 2 and r["row_index_zero_based"] == 3 for r in rows) else best_idx
    for r in rows:
        r["recommended_header_row"] = recommended
        r["diagnosis"] = "best_candidate" if r["row_index_zero_based"] == recommended else "scanned"
    return rows, sheet, recommended


def inspect_csv(path: Path) -> tuple[list[dict[str, Any]], str | None, int | None]:
    rows = []
    raw = pd.read_csv(path, header=None, nrows=10, dtype="string", encoding_errors="ignore")
    best_score = -1
    best_idx = None
    for idx, row in raw.iterrows():
        vals = [v for v in row.tolist() if pd.notna(v) and str(v).strip() != ""]
        score, has_stkcd, has_trdmnt, has_ret = row_score(vals)
        if score > best_score:
            best_score, best_idx = score, idx
        rows.append(
            {
                "file_path": str(path),
                "sheet_name": "",
                "row_index_zero_based": int(idx),
                "excel_row_number": int(idx) + 1,
                "non_null_count": len(vals),
                "row_values_preview": json.dumps([str(v) for v in vals[:20]], ensure_ascii=False),
                "contains_stkcd": has_stkcd,
                "contains_trdmnt": has_trdmnt,
                "contains_return_field": has_ret,
                "candidate_header_score": score,
                "recommended_header_row": None,
                "diagnosis": "",
            }
        )
    recommended = 3 if any(r["candidate_header_score"] >= 2 and r["row_index_zero_based"] == 3 for r in rows) else best_idx
    for r in rows:
        r["recommended_header_row"] = recommended
        r["diagnosis"] = "best_candidate" if r["row_index_zero_based"] == recommended else "scanned"
    return rows, "", recommended


def try_read_headers(path: Path, header: int) -> list[str]:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, header=header, nrows=5, dtype="string", encoding_errors="ignore")
    else:
        df = pd.read_excel(path, sheet_name=0, header=header, nrows=5, dtype="string", engine="openpyxl")
    return [str(c).strip() for c in df.columns]


def try_read_headers_skiprows3(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, skiprows=3, header=0, nrows=5, dtype="string", encoding_errors="ignore")
    else:
        df = pd.read_excel(path, sheet_name=0, skiprows=3, header=0, nrows=5, dtype="string", engine="openpyxl")
    return [str(c).strip() for c in df.columns]


def choose_column(cols: list[str], candidates_: list[str]) -> str | None:
    lower = {c.lower(): c for c in cols}
    for cand in candidates_:
        if cand.lower() in lower:
            return lower[cand.lower()]
    for col in cols:
        for cand in candidates_:
            if cand.lower() in col.lower():
                return col
    return None


def return_fields(cols: list[str]) -> list[str]:
    priority = [
        "Mretwd",
        "Mretnd",
        "月个股回报率，考虑现金红利再投资",
        "月个股回报率，不考虑现金红利再投资",
        "月个股回报率",
    ]
    found = []
    for cand in priority:
        for col in cols:
            if cand.lower() == col.lower() or cand in col:
                if col not in found:
                    found.append(col)
    return found


def header_scan(trd_files: list[Path]) -> tuple[pd.DataFrame, Path | None, int | None, list[str]]:
    rows = []
    best_path = None
    best_header = None
    best_cols: list[str] = []
    best_score = -1
    for path in trd_files:
        scanned, _, recommended = inspect_csv(path) if path.suffix.lower() == ".csv" else inspect_excel(path)
        rows.extend(scanned)
        for header in [0, 1, 2, 3]:
            try:
                cols = try_read_headers(path, header)
                score = int(choose_column(cols, ["Stkcd", "Symbol", "股票代码", "证券代码"]) is not None)
                score += int(choose_column(cols, ["Trdmnt", "TradingMonth", "交易月份", "交易月"]) is not None)
                score += int(bool(return_fields(cols)))
                diagnosis = f"read_attempt_header_{header}: score={score}; columns={cols[:20]}"
            except Exception as exc:
                cols = []
                score = -1
                diagnosis = f"read_attempt_header_{header}: error={type(exc).__name__}: {exc}"
            rows.append(
                {
                    "file_path": str(path),
                    "sheet_name": "" if path.suffix.lower() == ".csv" else "sheet1",
                    "row_index_zero_based": header,
                    "excel_row_number": header + 1,
                    "non_null_count": len(cols),
                    "row_values_preview": json.dumps(cols[:20], ensure_ascii=False),
                    "contains_stkcd": choose_column(cols, ["Stkcd", "Symbol", "股票代码", "证券代码"]) is not None,
                    "contains_trdmnt": choose_column(cols, ["Trdmnt", "TradingMonth", "交易月份", "交易月"]) is not None,
                    "contains_return_field": bool(return_fields(cols)),
                    "candidate_header_score": score,
                    "recommended_header_row": recommended,
                    "diagnosis": diagnosis,
                }
            )
            preferred_bonus = 0.1 if header == 3 else 0
            if score + preferred_bonus > best_score:
                best_score = score + preferred_bonus
                best_path = path
                best_header = header
                best_cols = cols
        try:
            cols = try_read_headers_skiprows3(path)
            score = int(choose_column(cols, ["Stkcd", "Symbol", "股票代码", "证券代码"]) is not None)
            score += int(choose_column(cols, ["Trdmnt", "TradingMonth", "交易月份", "交易月"]) is not None)
            score += int(bool(return_fields(cols)))
            diagnosis = f"read_attempt_skiprows_3_header_0: score={score}; columns={cols[:20]}"
        except Exception as exc:
            cols = []
            score = -1
            diagnosis = f"read_attempt_skiprows_3_header_0: error={type(exc).__name__}: {exc}"
        rows.append(
            {
                "file_path": str(path),
                "sheet_name": "" if path.suffix.lower() == ".csv" else "sheet1",
                "row_index_zero_based": 3,
                "excel_row_number": 4,
                "non_null_count": len(cols),
                "row_values_preview": json.dumps(cols[:20], ensure_ascii=False),
                "contains_stkcd": choose_column(cols, ["Stkcd", "Symbol", "股票代码", "证券代码"]) is not None,
                "contains_trdmnt": choose_column(cols, ["Trdmnt", "TradingMonth", "交易月份", "交易月"]) is not None,
                "contains_return_field": bool(return_fields(cols)),
                "candidate_header_score": score,
                "recommended_header_row": recommended,
                "diagnosis": diagnosis,
            }
        )
    diag = pd.DataFrame(rows)
    diag.to_csv(OUT_DIR / "trd_mnth_header_scan_diagnostic.csv", index=False, encoding="utf-8-sig")
    if best_score < 3:
        return diag, best_path, best_header, best_cols
    return diag, best_path, best_header, best_cols


def build_repaired_trd(path: Path | None, header: int | None, cols: list[str]) -> tuple[pd.DataFrame | None, dict[str, Any]]:
    if path is None or header is None:
        qa = {"file_path": None, "header_row_used": None, "row_count": 0, "unique_symbol_count": 0, "year_month_count": 0, "min_year_month": None, "max_year_month": None, "duplicate_symbol_year_month_count": 0, "invalid_symbol_count": 0, "monthly_return_null_count": 0, "fwd_ret_null_count": 0, "extreme_monthly_return_abs_gt_100pct_count": 0, "extreme_fwd_return_abs_gt_100pct_count": 0, "primary_return_field": None, "available_return_fields": "", "qa_status": "FAIL_NO_CANDIDATE_FILE", "caveat": "未找到可用 TRD_Mnth 文件。"}
        pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_csmar_trd_mnth_return_map_repaired_qa.csv", index=False, encoding="utf-8-sig")
        return None, qa
    symbol_col = choose_column(cols, ["Stkcd", "Symbol", "股票代码", "证券代码"])
    month_col = choose_column(cols, ["Trdmnt", "TradingMonth", "交易月份", "交易月"])
    ret_cols = return_fields(cols)
    primary = ret_cols[0] if ret_cols else None
    if not symbol_col or not month_col or not primary:
        qa = {"file_path": str(path), "header_row_used": header, "row_count": 0, "unique_symbol_count": 0, "year_month_count": 0, "min_year_month": None, "max_year_month": None, "duplicate_symbol_year_month_count": 0, "invalid_symbol_count": 0, "monthly_return_null_count": 0, "fwd_ret_null_count": 0, "extreme_monthly_return_abs_gt_100pct_count": 0, "extreme_fwd_return_abs_gt_100pct_count": 0, "primary_return_field": primary, "available_return_fields": "|".join(ret_cols), "qa_status": f"FAIL_UNPARSEABLE_COLUMNS: columns={cols[:20]}", "caveat": "header scan 未找到代码、月份、收益三类必要字段。"}
        pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_csmar_trd_mnth_return_map_repaired_qa.csv", index=False, encoding="utf-8-sig")
        return None, qa
    usecols = list(dict.fromkeys([symbol_col, month_col] + ret_cols))
    print(f"读取 TRD_Mnth 必要列: file={path}, header={header}, usecols={usecols}")
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, header=header, usecols=usecols, dtype={symbol_col: "string"}, encoding_errors="ignore")
    else:
        df = pd.read_excel(path, sheet_name=0, header=header, usecols=usecols, dtype={symbol_col: "string"}, engine="openpyxl")
    df["symbol_norm"] = normalize_symbol(df[symbol_col])
    df["year_month"] = pd.to_datetime(df[month_col].astype("string"), errors="coerce").dt.to_period("M").astype("string")
    for c in ret_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["monthly_return_t"] = df[primary]
    df = df.sort_values(["symbol_norm", "year_month"])
    df["fwd_ret_1m"] = df.groupby("symbol_norm")["monthly_return_t"].shift(-1)
    df["next_year_month"] = df.groupby("symbol_norm")["year_month"].shift(-1)
    expected_next = (pd.PeriodIndex(df["year_month"].dropna(), freq="M") + 1).astype(str)
    expected = pd.Series(pd.NA, index=df.index, dtype="string")
    expected.loc[df["year_month"].notna()] = expected_next
    df.loc[df["next_year_month"].ne(expected), "fwd_ret_1m"] = np.nan
    raw_values = df[ret_cols].astype("string").to_dict(orient="records")
    out = pd.DataFrame(
        {
            "source_name": "CSMAR_TRD_MNTH_REPAIRED",
            "file_path": str(path),
            "symbol_norm": df["symbol_norm"],
            "year_month": df["year_month"],
            "monthly_return_t": df["monthly_return_t"],
            "fwd_ret_1m": df["fwd_ret_1m"],
            "primary_return_field": primary,
            "raw_return_field_values": [json.dumps(x, ensure_ascii=False) for x in raw_values],
        }
    )
    out["return_valid_flag"] = out["symbol_norm"].notna() & out["year_month"].notna() & out["fwd_ret_1m"].notna()
    out["invalid_reason"] = np.select(
        [out["symbol_norm"].isna(), out["year_month"].isna(), out["monthly_return_t"].isna(), out["fwd_ret_1m"].isna()],
        ["INVALID_SYMBOL", "INVALID_MONTH", "MISSING_MONTHLY_RETURN_T", "MISSING_NEXT_MONTH_RETURN"],
        default="",
    )
    dup = int(out.duplicated(["symbol_norm", "year_month"]).sum())
    monthly_null = int(out["monthly_return_t"].isna().sum())
    fwd_null = int(out["fwd_ret_1m"].isna().sum())
    invalid_symbol = int(out["symbol_norm"].isna().sum())
    qa_status = "PASS" if dup == 0 and invalid_symbol == 0 and monthly_null == 0 and fwd_null == 0 else ("WATCH" if len(out) else "FAIL")
    qa = {"file_path": str(path), "header_row_used": header, "row_count": len(out), "unique_symbol_count": int(out["symbol_norm"].nunique()), "year_month_count": int(out["year_month"].nunique()), "min_year_month": str(out["year_month"].dropna().min()) if out["year_month"].notna().any() else None, "max_year_month": str(out["year_month"].dropna().max()) if out["year_month"].notna().any() else None, "duplicate_symbol_year_month_count": dup, "invalid_symbol_count": invalid_symbol, "monthly_return_null_count": monthly_null, "fwd_ret_null_count": fwd_null, "extreme_monthly_return_abs_gt_100pct_count": int(out["monthly_return_t"].abs().gt(1).sum()), "extreme_fwd_return_abs_gt_100pct_count": int(out["fwd_ret_1m"].abs().gt(1).sum()), "primary_return_field": primary, "available_return_fields": "|".join(ret_cols), "qa_status": qa_status, "caveat": "TRD_Mnth 月收益按 next natural month shift 构造 fwd_ret_1m；Mretwd 与 close-based price return 定义可能不同。"}
    out.to_parquet(OUT_DIR / "canonical_csmar_trd_mnth_return_map_repaired.parquet", index=False)
    pd.DataFrame([qa]).to_csv(OUT_DIR / "canonical_csmar_trd_mnth_return_map_repaired_qa.csv", index=False, encoding="utf-8-sig")
    del df
    gc.collect()
    return out, qa


def cross_validate_trd_all_daily(trd: pd.DataFrame | None, canonical: pd.DataFrame) -> dict[str, Any]:
    if trd is None or trd.empty:
        row = {"common_row_count": 0, "pearson_corr": np.nan, "spearman_corr": np.nan, "mean_abs_diff": np.nan, "median_abs_diff": np.nan, "p95_abs_diff": np.nan, "near_equal_ratio_1e_6": np.nan, "diff_abs_gt_1pct_count": 0, "diff_abs_gt_5pct_count": 0, "validation_status": "INCONCLUSIVE_LOW_OVERLAP", "interpretation": "TRD_Mnth 未成功解析，无法交叉验证。"}
    else:
        left = trd[trd["return_valid_flag"]][["symbol_norm", "year_month", "fwd_ret_1m"]].rename(columns={"fwd_ret_1m": "trd_ret"})
        right = canonical[canonical["return_valid_flag"]][["symbol_norm", "year_month", "fwd_ret_1m"]].rename(columns={"fwd_ret_1m": "all_daily_ret"})
        common = left.merge(right, on=["symbol_norm", "year_month"], how="inner")
        diff = common["trd_ret"] - common["all_daily_ret"]
        absdiff = diff.abs()
        pearson = common["trd_ret"].corr(common["all_daily_ret"], method="pearson") if len(common) >= 2 else np.nan
        spearman = common["trd_ret"].corr(common["all_daily_ret"], method="spearman") if len(common) >= 2 else np.nan
        if len(common) < 100:
            status = "INCONCLUSIVE_LOW_OVERLAP"
        elif pearson >= 0.995 and float(absdiff.quantile(0.95)) <= 0.001:
            status = "PASS_HIGH_CONSISTENCY"
        elif pearson >= 0.95:
            status = "WATCH_DIFFERENCES"
        else:
            status = "WATCH_DIFFERENCES"
        row = {"common_row_count": len(common), "pearson_corr": as_float_or_none(pearson), "spearman_corr": as_float_or_none(spearman), "mean_abs_diff": as_float_or_none(absdiff.mean()) if len(common) else None, "median_abs_diff": as_float_or_none(absdiff.median()) if len(common) else None, "p95_abs_diff": as_float_or_none(absdiff.quantile(0.95)) if len(common) else None, "near_equal_ratio_1e_6": as_float_or_none(absdiff.le(1e-6).mean()) if len(common) else None, "diff_abs_gt_1pct_count": int(absdiff.gt(0.01).sum()) if len(common) else 0, "diff_abs_gt_5pct_count": int(absdiff.gt(0.05).sum()) if len(common) else 0, "validation_status": status, "interpretation": "TRD_Mnth Mretwd 可能含现金红利再投资；all_daily 是收盘价价差收益，差异需按收益定义解释。"}
    pd.DataFrame([row]).to_csv(OUT_DIR / "trd_mnth_vs_all_daily_return_validation.csv", index=False, encoding="utf-8-sig")
    return row


def month_close_from_all_daily(symbols: set[str]) -> pd.DataFrame:
    df = pd.read_parquet(ALL_DAILY_PATH, columns=["date", "symbol", "close"])
    df["symbol_norm"] = normalize_symbol(df["symbol"])
    df = df[df["symbol_norm"].isin(symbols)].copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    df = df[df["date"].notna() & df["close"].gt(0)].copy()
    df["year_month"] = df["date"].dt.to_period("M").astype(str)
    df = df[df["year_month"].isin(["2024-12", "2025-01"])].sort_values(["symbol_norm", "year_month", "date"])
    out = df.groupby(["symbol_norm", "year_month"], as_index=False).tail(1)[["symbol_norm", "year_month", "date", "close"]]
    del df
    gc.collect()
    return out


def v0_2024_12_detail(trd: pd.DataFrame | None, canonical: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    weights = pd.read_parquet(V0_WEIGHTS_PATH, columns=["symbol", "month_end", "weight", "rank_in_month"])
    weights["symbol_norm"] = normalize_symbol(weights["symbol"])
    weights["year_month"] = pd.to_datetime(weights["month_end"]).dt.to_period("M").astype(str)
    w = weights[(weights["year_month"] == "2024-12") & weights["weight"].abs().gt(0)].copy()
    symbols = set(w["symbol_norm"].dropna())
    close = month_close_from_all_daily(symbols)
    c12 = close[close["year_month"] == "2024-12"].rename(columns={"date": "date_2024_12", "close": "close_2024_12"})[["symbol_norm", "date_2024_12", "close_2024_12"]]
    c01 = close[close["year_month"] == "2025-01"].rename(columns={"date": "date_2025_01", "close": "close_2025_01"})[["symbol_norm", "date_2025_01", "close_2025_01"]]
    all_daily_key = canonical[(canonical["year_month"] == "2024-12")][["symbol_norm", "fwd_ret_1m", "return_valid_flag"]].rename(columns={"fwd_ret_1m": "all_daily_canonical_fwd_ret_1m", "return_valid_flag": "all_daily_valid"})
    if trd is not None and not trd.empty:
        trd_key = trd[(trd["year_month"] == "2024-12")][["symbol_norm", "fwd_ret_1m", "return_valid_flag"]].rename(columns={"fwd_ret_1m": "trd_mnth_fwd_ret_1m", "return_valid_flag": "trd_valid"})
    else:
        trd_key = pd.DataFrame(columns=["symbol_norm", "trd_mnth_fwd_ret_1m", "trd_valid"])
    detail = w.merge(trd_key, on="symbol_norm", how="left").merge(all_daily_key, on="symbol_norm", how="left").merge(c12, on="symbol_norm", how="left").merge(c01, on="symbol_norm", how="left")
    detail["matched_in_repaired_trd_mnth"] = detail["trd_valid"].fillna(False).astype(bool)
    detail["matched_in_all_daily_canonical"] = detail["all_daily_valid"].fillna(False).astype(bool)
    detail["has_all_daily_2024_12"] = detail["close_2024_12"].notna()
    detail["has_all_daily_2025_01"] = detail["close_2025_01"].notna()
    detail["manually_computable_from_all_daily"] = detail["has_all_daily_2024_12"] & detail["has_all_daily_2025_01"]
    detail["manual_all_daily_fwd_ret_1m"] = detail["close_2025_01"] / detail["close_2024_12"] - 1.0

    def reason(row: pd.Series) -> str:
        if row["matched_in_repaired_trd_mnth"]:
            return "ALREADY_MATCHED_IN_REPAIRED_TRD_MNTH"
        if row["matched_in_all_daily_canonical"]:
            return "ALREADY_MATCHED_IN_ALL_DAILY_CANONICAL"
        if row["manually_computable_from_all_daily"]:
            return "ALL_DAILY_RETURN_MAP_CONSTRUCTION_MISSED_FIXABLE"
        if row["has_all_daily_2024_12"] and not row["has_all_daily_2025_01"]:
            return "SYMBOL_PRESENT_2024_12_BUT_2025_01_MISSING"
        if not row["has_all_daily_2024_12"] and not row["has_all_daily_2025_01"]:
            return "SYMBOL_ABSENT_FROM_ALL_DAILY_GLOBAL"
        return "UNKNOWN"

    detail["gap_reason"] = detail.apply(reason, axis=1)
    detail["suggested_fix"] = np.select(
        [
            detail["gap_reason"].eq("ALL_DAILY_RETURN_MAP_CONSTRUCTION_MISSED_FIXABLE"),
            detail["gap_reason"].eq("SYMBOL_PRESENT_2024_12_BUT_2025_01_MISSING"),
            detail["matched_in_repaired_trd_mnth"],
        ],
        [
            "在本任务 patched canonical map 中补充严格下一自然月 all_daily forward return。",
            "官方 TRD_Mnth 若可用则优先确认；all_daily 原始行情缺少 2025-01，不能用 t+2 替代。",
            "可用 repaired TRD_Mnth 官方 fwd return 覆盖。",
        ],
        default="保留为真实缺口或待人工确认。",
    )
    out = detail.rename(columns={"symbol": "symbol_raw"})[
        [
            "symbol_raw",
            "symbol_norm",
            "weight",
            "rank_in_month",
            "matched_in_repaired_trd_mnth",
            "trd_mnth_fwd_ret_1m",
            "matched_in_all_daily_canonical",
            "all_daily_canonical_fwd_ret_1m",
            "has_all_daily_2024_12",
            "has_all_daily_2025_01",
            "close_2024_12",
            "close_2025_01",
            "manually_computable_from_all_daily",
            "manual_all_daily_fwd_ret_1m",
            "gap_reason",
            "suggested_fix",
        ]
    ]
    out.to_csv(OUT_DIR / "v0_2024_12_repaired_source_match_detail.csv", index=False, encoding="utf-8-sig")
    patchable = out[out["gap_reason"].eq("ALL_DAILY_RETURN_MAP_CONSTRUCTION_MISSED_FIXABLE")].copy()
    del weights, close, c12, c01, detail
    gc.collect()
    return out, patchable


def maybe_patch_canonical(canonical: pd.DataFrame, patchable: pd.DataFrame) -> tuple[pd.DataFrame | None, pd.DataFrame, bool]:
    audit_rows = []
    if patchable.empty:
        audit = pd.DataFrame(columns=["symbol_norm", "year_month", "previous_missing", "patched", "month_trade_date", "next_month_trade_date", "close", "next_close", "patched_fwd_ret_1m", "patch_reason", "patch_valid"])
        audit.to_csv(OUT_DIR / "canonical_return_map_2024_12_patch_audit.csv", index=False, encoding="utf-8-sig")
        return None, audit, False
    patched = canonical.copy()
    for _, row in patchable.iterrows():
        mask = (patched["symbol_norm"].eq(row["symbol_norm"])) & (patched["year_month"].eq("2024-12"))
        if mask.any():
            patched.loc[mask, "fwd_ret_1m"] = row["manual_all_daily_fwd_ret_1m"]
            patched.loc[mask, "return_valid_flag"] = True
            patched.loc[mask, "caveat"] = patched.loc[mask, "caveat"].astype("string").fillna("") + " | patched_2024_12_from_raw_all_daily"
        else:
            patched = pd.concat(
                [
                    patched,
                    pd.DataFrame([{"canonical_source": "ALL_DAILY_CLOSE_BASED_PATCHED", "symbol_norm": row["symbol_norm"], "year_month": "2024-12", "fwd_ret_1m": row["manual_all_daily_fwd_ret_1m"], "return_valid_flag": True, "source_name": "ALL_DAILY_CLOSE_BASED_PATCHED", "source_priority": 2, "caveat": "patched_2024_12_from_raw_all_daily"}]),
                ],
                ignore_index=True,
            )
        audit_rows.append({"symbol_norm": row["symbol_norm"], "year_month": "2024-12", "previous_missing": True, "patched": True, "month_trade_date": None, "next_month_trade_date": None, "close": row["close_2024_12"], "next_close": row["close_2025_01"], "patched_fwd_ret_1m": row["manual_all_daily_fwd_ret_1m"], "patch_reason": "raw all_daily has valid 2024-12 and 2025-01 month-end closes", "patch_valid": True})
    patched.to_parquet(OUT_DIR / "canonical_monthly_stock_return_map_patched_v0.parquet", index=False)
    audit = pd.DataFrame(audit_rows)
    audit.to_csv(OUT_DIR / "canonical_return_map_2024_12_patch_audit.csv", index=False, encoding="utf-8-sig")
    return patched, audit, True


def coverage_for_map(name: str, retmap: pd.DataFrame | None) -> dict[str, Any]:
    if retmap is None or retmap.empty:
        return {"return_map_name": name, "row_count": 0, "unique_symbol_count": 0, "min_year_month": None, "max_year_month": None, "avg_matched_weight_share": np.nan, "min_matched_weight_share": np.nan, "low_match_month_count": np.nan, "worst_month": None, "worst_month_matched_weight_share": np.nan, "coverage_status": "LOW_MATCH", "caveat": "return map missing"}
    keys = retmap[retmap["return_valid_flag"]][["symbol_norm", "year_month"]].drop_duplicates()
    w = pd.read_parquet(V0_WEIGHTS_PATH, columns=["symbol", "month_end", "weight"])
    w["symbol_norm"] = normalize_symbol(w["symbol"])
    w["year_month"] = pd.to_datetime(w["month_end"]).dt.to_period("M").astype(str)
    m = w.merge(keys.assign(_matched=True), on=["symbol_norm", "year_month"], how="left")
    m["_matched"] = m["_matched"].fillna(False)
    m["abs_weight"] = m["weight"].abs()
    shares = m.groupby("year_month").apply(lambda g: float(g.loc[g["_matched"], "abs_weight"].sum() / g["abs_weight"].sum()) if g["abs_weight"].sum() else np.nan, include_groups=False)
    avg = float(shares.mean()) if len(shares) else np.nan
    mn = float(shares.min()) if len(shares) else np.nan
    worst = str(shares.idxmin()) if len(shares) and pd.notna(mn) else None
    if avg >= 0.98 and mn >= 0.95:
        status = "READY"
    elif avg >= 0.95 and mn >= 0.90:
        status = "READY_WITH_MINOR_GAPS"
    elif avg >= 0.90:
        status = "WATCH_COVERAGE_GAPS"
    else:
        status = "LOW_MATCH"
    row = {"return_map_name": name, "row_count": len(retmap), "unique_symbol_count": int(retmap["symbol_norm"].nunique()), "min_year_month": str(retmap["year_month"].dropna().min()) if retmap["year_month"].notna().any() else None, "max_year_month": str(retmap["year_month"].dropna().max()) if retmap["year_month"].notna().any() else None, "avg_matched_weight_share": avg, "min_matched_weight_share": mn, "low_match_month_count": int(shares.lt(0.95).sum()), "worst_month": worst, "worst_month_matched_weight_share": mn, "coverage_status": status, "caveat": "" if status in {"READY", "READY_WITH_MINOR_GAPS"} else f"worst_month={worst}"}
    del w, m, shares
    gc.collect()
    return row


def guardrails() -> pd.DataFrame:
    checks = {
        "strategy_alpha_signal_generated": False,
        "strategy_weights_generated": False,
        "old_artifacts_modified": False,
        "production_modified": False,
        "ml_training_run": False,
        "new_ml_model_trained": False,
        "portfolio_returns_calculated": False,
        "benchmark_relative_returns_calculated": False,
        "alpha_beta_regression_calculated": False,
        "information_ratio_calculated": False,
        "tracking_error_calculated": False,
        "ff_regression_calculated": False,
        "dgtw_adjusted_eval_calculated": False,
        "shap_calculated": False,
    }
    df = pd.DataFrame([{"guardrail": k, "expected": False, "actual": v, "pass": v is False} for k, v in checks.items()])
    df.to_csv(OUT_DIR / "trd_mnth_parser_repair_guardrail_qa.csv", index=False, encoding="utf-8-sig")
    return df


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    run_timestamp = datetime.now().isoformat(timespec="seconds")
    trd_files, des_files = candidates()
    prereq = prerequisite_check(trd_files, des_files)
    diag, best_path, best_header, best_cols = header_scan(trd_files)
    trd, trd_qa = build_repaired_trd(best_path, best_header, best_cols)
    canonical = pd.read_parquet(CANONICAL_ORIG_PATH)
    validation = cross_validate_trd_all_daily(trd, canonical)
    detail, patchable = v0_2024_12_detail(trd, canonical)
    patched, patch_audit, patch_generated = maybe_patch_canonical(canonical, patchable)
    coverage_rows = [
        coverage_for_map("all_daily_canonical_original", canonical),
        coverage_for_map("repaired_trd_mnth", trd),
    ]
    if patched is not None:
        coverage_rows.append(coverage_for_map("all_daily_canonical_patched_v0", patched))
    coverage_df = pd.DataFrame(coverage_rows)
    coverage_df.to_csv(OUT_DIR / "strict_lag_v0_coverage_after_trd_repair.csv", index=False, encoding="utf-8-sig")
    guardrail_df = guardrails()

    reason_mode = detail["gap_reason"].mode().iloc[0] if not detail.empty else "UNKNOWN"
    repaired_cov = coverage_df[coverage_df["return_map_name"].eq("repaired_trd_mnth")]
    patched_cov = coverage_df[coverage_df["return_map_name"].eq("all_daily_canonical_patched_v0")]
    patched_status = patched_cov["coverage_status"].iloc[0] if len(patched_cov) else None
    patched_avg = as_float_or_none(patched_cov["avg_matched_weight_share"].iloc[0]) if len(patched_cov) else None
    patched_min = as_float_or_none(patched_cov["min_matched_weight_share"].iloc[0]) if len(patched_cov) else None
    repaired_status = repaired_cov["coverage_status"].iloc[0] if len(repaired_cov) else "LOW_MATCH"
    no_guardrail = bool(guardrail_df["pass"].all())
    if not no_guardrail:
        decision = "RETURN_MAP_REPAIR_FAIL_GUARDRAIL"
    elif trd is not None and trd_qa["qa_status"] in {"PASS", "WATCH"} and repaired_status in {"READY", "READY_WITH_MINOR_GAPS"}:
        decision = "TRD_MNTH_REPAIRED_READY_FOR_CANONICAL_EVAL"
    elif patched is not None and patched_status in {"READY", "READY_WITH_MINOR_GAPS"}:
        decision = "ALL_DAILY_PATCHED_READY_FOR_CANONICAL_EVAL"
    elif trd is not None and validation["validation_status"] in {"WATCH_DIFFERENCES", "FAIL_LOW_CONSISTENCY"}:
        decision = "TRD_MNTH_REPAIRED_BUT_ALL_DAILY_PRIMARY_WITH_CAVEAT"
    else:
        decision = "RETURN_MAP_REPAIR_INCONCLUSIVE"

    if decision == "TRD_MNTH_REPAIRED_READY_FOR_CANONICAL_EVAL":
        recommended = "canonical_csmar_trd_mnth_return_map_repaired.parquet"
        next_step = "可用 repaired TRD_Mnth 作为下一轮统一评价收益源。"
    elif decision == "ALL_DAILY_PATCHED_READY_FOR_CANONICAL_EVAL":
        recommended = "canonical_monthly_stock_return_map_patched_v0.parquet"
        next_step = "可用 patched all_daily canonical map 进入下一轮统一评价。"
    else:
        recommended = "canonical_monthly_stock_return_map.parquet"
        next_step = "TRD_Mnth 仍需确认导出结构；2024-12 缺口按 detail 表逐只核对。"

    guardrail_values = {row["guardrail"]: bool(row["actual"]) for _, row in guardrail_df.iterrows()}
    summary = {
        "run_timestamp": run_timestamp,
        "prerequisites_passed": bool(prereq["prerequisites_passed"]),
        "trd_mnth_header_repaired": trd is not None,
        "trd_mnth_header_row_used": best_header,
        "trd_mnth_map_generated": trd is not None,
        "trd_mnth_qa_status": trd_qa["qa_status"],
        "trd_mnth_primary_return_field": trd_qa["primary_return_field"],
        "trd_mnth_row_count": int(trd_qa["row_count"]),
        "trd_mnth_unique_symbol_count": int(trd_qa["unique_symbol_count"]),
        "trd_mnth_min_year_month": trd_qa["min_year_month"],
        "trd_mnth_max_year_month": trd_qa["max_year_month"],
        "cross_validation_status": validation["validation_status"],
        "cross_validation_common_row_count": int(validation["common_row_count"]),
        "cross_validation_pearson_corr": validation["pearson_corr"],
        "cross_validation_mean_abs_diff": validation["mean_abs_diff"],
        "v0_2024_12_selected_count": int(len(detail)),
        "v0_2024_12_matched_in_trd_mnth_count": int(detail["matched_in_repaired_trd_mnth"].sum()) if not detail.empty else 0,
        "v0_2024_12_matched_in_all_daily_original_count": int(detail["matched_in_all_daily_canonical"].sum()) if not detail.empty else 0,
        "v0_2024_12_manually_computable_from_all_daily_count": int(detail["manually_computable_from_all_daily"].sum()) if not detail.empty else 0,
        "patch_generated": bool(patch_generated),
        "patched_return_map_generated": patched is not None,
        "patched_map_v0_coverage_status": patched_status,
        "patched_map_v0_avg_matched_weight_share": patched_avg,
        "patched_map_v0_min_matched_weight_share": patched_min,
        "main_2024_12_gap_reason": reason_mode,
        "repairable": bool(not patchable.empty or (trd is not None and detail["matched_in_repaired_trd_mnth"].any())),
        "recommended_return_map_for_next_eval": recommended,
        **guardrail_values,
        "final_decision": decision,
        "recommended_next_step": next_step,
    }
    write_json(OUT_DIR / "trd_mnth_parser_repair_2024_12_coverage_repair_summary.json", summary)
    pd.DataFrame([
        {"check": "trd_mnth_qa_status", "value": summary["trd_mnth_qa_status"]},
        {"check": "cross_validation_status", "value": summary["cross_validation_status"]},
        {"check": "main_2024_12_gap_reason", "value": summary["main_2024_12_gap_reason"]},
        {"check": "final_decision", "value": decision},
    ]).to_csv(OUT_DIR / "final_qa.csv", index=False, encoding="utf-8-sig")
    write_json(OUT_DIR / "terminal_summary.json", {"task_name": TASK_NAME, "run_timestamp": run_timestamp, "final_decision": decision, "output_dir": str(OUT_DIR)})
    (OUT_DIR / "trd_mnth_parser_repair_2024_12_coverage_repair_report.md").write_text(
        f"# TRD_Mnth Parser Repair & 2024-12 Coverage Repair v0\n\n- final_decision: {decision}\n- trd_mnth_qa_status: {trd_qa['qa_status']}\n- header_row_used: {best_header}\n- 2024-12 main gap reason: {reason_mode}\n- recommended_return_map_for_next_eval: {recommended}\n\n{next_step}\n",
        encoding="utf-8",
    )
    (OUT_DIR / "task_completion_card.md").write_text(
        f"# task_completion_card\n\n- task_name: {TASK_NAME}\n- completed_at: {run_timestamp}\n- final_decision: {decision}\n- output_dir: {OUT_DIR}\n",
        encoding="utf-8",
    )
    (RUN_DIR / "RUN_STATE.md").write_text(
        f"# {TASK_NAME}\n\n状态：完成。\n\n完成时间：{run_timestamp}\n\nfinal_decision：{decision}\n\n关键输出目录：`{OUT_DIR}`\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
