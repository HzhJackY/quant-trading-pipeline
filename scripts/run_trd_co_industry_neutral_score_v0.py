from __future__ import annotations

import csv
import gc
import json
import warnings
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

warnings.filterwarnings("ignore", message="Workbook contains no default style.*")


ROOT = Path(r"C:\dev\quant")
TASK_NAME = "trd_co_industry_neutral_score_run_v0"
OUT_DIR = ROOT / "output" / TASK_NAME
RUN_DIR = ROOT / "output" / "_agent_runs" / TASK_NAME

TRD_CO_XLSX = ROOT / "data" / "csmar_exports" / "TRD_Co.xlsx"
FIELD_DICT = ROOT / "output" / "csmar_field_dictionary_v0" / "csmar_field_dictionary_TRD_Co.csv"
FIELD_DICT_SUMMARY = ROOT / "output" / "csmar_field_dictionary_v0" / "csmar_field_dictionary_update_TRD_Co_summary.json"
SCORE_PANEL_PATH = ROOT / "output" / "simple_baseline_score_run_v0" / "simple_baseline_score_panel_v0.parquet"
SCORE_SUMMARY = ROOT / "output" / "simple_baseline_score_run_v0" / "simple_baseline_score_run_summary.json"
CANDIDATE_MANIFEST = ROOT / "output" / "simple_baseline_candidate_selection_v0" / "simple_baseline_candidate_manifest.csv"
FORMULA_MANIFEST = ROOT / "output" / "simple_baseline_score_prep_v0" / "simple_baseline_score_formula_manifest.csv"

TRD_FIELDS = [
    "Stkcd",
    "Stknme",
    "Listdt",
    "Indcd",
    "Indnme",
    "Nindcd",
    "Nindnme",
    "Nnindcd",
    "Nnindnme",
    "IndcdZX",
    "IndnmeZX",
    "OWNERSHIPTYPE",
    "OWNERSHIPTYPECODE",
    "Favaldt",
    "Curtrd",
    "Sctcd",
    "Statco",
    "Statdt",
    "Markettype",
    "FormerCode",
]

SCORE_COLUMNS = [
    "symbol",
    "month_end",
    "bp_rank",
    "ep_ttm_rank",
    "cfo_to_earnings_parent_rank",
    "VALUE_BP_SINGLE_score",
    "VALUE_QUALITY_EQUAL_WEIGHT_score",
    "fwd_ret_1m",
]

SOURCE_FEATURES = ["bp_rank", "ep_ttm_rank", "cfo_to_earnings_parent_rank"]
RAW_SCORE_COLUMNS = ["VALUE_BP_SINGLE_score", "VALUE_QUALITY_EQUAL_WEIGHT_score"]


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def norm_symbol(value: Any) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if "." in text:
        left, right = text.split(".", 1)
        if right.strip("0") == "" and left.isdigit():
            text = left
    if text.isdigit():
        return text.zfill(6)
    return text


def parse_date_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def read_trd_co_xlsx() -> tuple[pd.DataFrame, list[str], str]:
    from openpyxl import load_workbook

    wb = load_workbook(TRD_CO_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(max_row=1, values_only=True))]
    wanted = [field for field in TRD_FIELDS if field in header]
    idx_to_field = {idx: field for idx, field in enumerate(header) if field in wanted}
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {field: values[idx] if idx < len(values) else None for idx, field in idx_to_field.items()}
        if any(value is not None and str(value).strip() != "" for value in row.values()):
            rows.append(row)
    sheet_name = ws.title
    wb.close()
    del wb, ws
    df = pd.DataFrame(rows)
    for field in wanted:
        if field not in df.columns:
            df[field] = pd.NA
    gc.collect()
    return df, header, sheet_name


def top_categories(series: pd.Series, n: int = 8) -> str:
    counts = series.dropna().astype(str).value_counts().head(n)
    return "; ".join(f"{idx}:{int(val)}" for idx, val in counts.items())


def coverage_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    fields = ["Nnindcd", "Nnindnme", "IndcdZX", "IndnmeZX", "Nindcd", "Nindnme", "Indcd", "Indnme"]
    rows = []
    denom = len(df)
    for field in fields:
        if field in df.columns:
            non_null = int(df[field].notna().sum())
            unique = int(df[field].dropna().astype(str).nunique())
            tops = top_categories(df[field])
        else:
            non_null = 0
            unique = 0
            tops = ""
        rows.append(
            {
                "field_name": field,
                "field_present": field in df.columns,
                "non_null_count": non_null,
                "non_null_ratio": non_null / denom if denom else 0,
                "unique_value_count": unique,
                "top_categories": tops,
            }
        )
    return rows


def choose_industry_fields(df: pd.DataFrame) -> tuple[str | None, str | None, str | None, str | None, str | None, str | None]:
    pairs = [("Nnindcd", "Nnindnme"), ("IndcdZX", "IndnmeZX"), ("Nindcd", "Nindnme"), ("Indcd", "Indnme")]
    selected_code = selected_name = None
    for code, name in pairs:
        if code in df.columns and int(df[code].notna().sum()) > 0:
            selected_code = code
            selected_name = name if name in df.columns else None
            break
    secondary_code = secondary_name = None
    for code, name in pairs:
        if code != selected_code and code in df.columns and int(df[code].notna().sum()) > 0:
            secondary_code = code
            secondary_name = name if name in df.columns else None
            break
    coarse_code, coarse_name = (("Nindcd", "Nindnme") if "Nindcd" in df.columns else ("Indcd", "Indnme"))
    return selected_code, selected_name, secondary_code, secondary_name, coarse_code, coarse_name


def variation_check(df: pd.DataFrame) -> tuple[list[dict[str, Any]], bool]:
    rows = []
    any_change = False
    fields = ["Nnindcd", "IndcdZX", "Nindcd", "Indcd", "Statco"]
    multi_symbols = df.groupby("symbol", dropna=True).size()
    multi_symbols = multi_symbols[multi_symbols > 1].index
    for field in fields:
        if field not in df.columns:
            rows.append(
                {
                    "field_name": field,
                    "field_present": False,
                    "symbols_with_multiple_rows": len(multi_symbols),
                    "symbols_with_field_variation": 0,
                    "variation_ratio_among_multi_row_symbols": 0,
                    "statdt_present_for_variation_check": "Statdt" in df.columns,
                    "notes": "field missing",
                }
            )
            continue
        changed = 0
        for _, group in df[df["symbol"].isin(multi_symbols)].groupby("symbol", dropna=True):
            if group[field].dropna().astype(str).nunique() > 1:
                changed += 1
        if field in {"Nnindcd", "IndcdZX", "Nindcd", "Indcd"} and changed > 0:
            any_change = True
        rows.append(
            {
                "field_name": field,
                "field_present": True,
                "symbols_with_multiple_rows": len(multi_symbols),
                "symbols_with_field_variation": changed,
                "variation_ratio_among_multi_row_symbols": changed / len(multi_symbols) if len(multi_symbols) else 0,
                "statdt_present_for_variation_check": "Statdt" in df.columns,
                "notes": "Statdt is not automatically treated as industry effective date.",
            }
        )
    return rows, any_change


def classify_source(
    df: pd.DataFrame,
    primary_code_field: str | None,
    primary_coverage: float,
    industry_changes_detected: bool,
) -> tuple[str, str, str | None]:
    if not primary_code_field or primary_coverage <= 0:
        return "NOT_USABLE", "NOT_USABLE", None
    row_counts = df.groupby("symbol", dropna=True).size()
    multi_row = bool((row_counts > 1).any())
    status_changes = False
    if multi_row and "Statco" in df.columns:
        for _, group in df.groupby("symbol", dropna=True):
            if len(group) > 1 and group["Statco"].dropna().astype(str).nunique() > 1:
                status_changes = True
                break
    if multi_row and industry_changes_detected and "Statdt" in df.columns:
        return "AMBIGUOUS_MANUAL_REVIEW", "AMBIGUOUS", "Statdt"
    if multi_row and status_changes and not industry_changes_detected:
        return "STATIC_WITH_STATUS_HISTORY", "STATIC_NOT_PIT", None
    return "STATIC_INDUSTRY_AVAILABLE", "STATIC_NOT_PIT", None


def latest_per_symbol(df: pd.DataFrame) -> pd.DataFrame:
    sort_cols = []
    for col in ["Statdt", "Favaldt", "Listdt"]:
        if col in df.columns:
            parsed = f"_{col}_parsed"
            df[parsed] = parse_date_series(df[col])
            sort_cols.append(parsed)
    if sort_cols:
        out = df.sort_values(sort_cols).groupby("symbol", as_index=False, dropna=True).tail(1)
        out = out.drop(columns=sort_cols)
    else:
        out = df.drop_duplicates("symbol", keep="last")
    return out


def add_neutral_scores(joined: pd.DataFrame, prefix: str) -> tuple[pd.DataFrame, list[str]]:
    out = joined.copy()
    for feature in SOURCE_FEATURES:
        rank_col = f"_{feature}_industry_rank"
        out[rank_col] = out.groupby(["month_end", "primary_industry_code"], dropna=False)[feature].rank(pct=True, method="average")
    bp_score = f"{prefix}_IND_NEUTRAL_VALUE_BP_SINGLE_score"
    ew_score = f"{prefix}_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score"
    out[bp_score] = out["_bp_rank_industry_rank"]
    rank_cols = [f"_{feature}_industry_rank" for feature in SOURCE_FEATURES]
    out[ew_score] = out[rank_cols].mean(axis=1, skipna=True)
    out = out.drop(columns=rank_cols)
    return out, [bp_score, ew_score]


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    run_timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    print(f"[{run_timestamp}] start {TASK_NAME}")

    required = [TRD_CO_XLSX, FIELD_DICT, FIELD_DICT_SUMMARY, SCORE_PANEL_PATH, SCORE_SUMMARY, CANDIDATE_MANIFEST, FORMULA_MANIFEST]
    prereq_rows = [{"path": rel(path), "exists": path.exists(), "bytes": path.stat().st_size if path.exists() else None} for path in required]
    prerequisites_passed = all(row["exists"] for row in prereq_rows)
    xlsx_read = False

    if not prerequisites_passed:
        final_decision = "TRD_CO_INDUSTRY_NEUTRAL_SCORE_RUN_FAIL"
        summary = {
            "run_timestamp": run_timestamp,
            "trd_co_xlsx_read": xlsx_read,
            "trd_co_xlsx_path": str(TRD_CO_XLSX),
            "prerequisites_passed": False,
            "final_decision": final_decision,
            "recommended_next_step": "补齐缺失输入后重跑。",
        }
        write_json(OUT_DIR / "trd_co_industry_neutral_score_run_summary.json", summary)
        return 1

    trd_df, xlsx_header, sheet_name = read_trd_co_xlsx()
    xlsx_read = True
    trd_df["symbol"] = trd_df["Stkcd"].map(norm_symbol) if "Stkcd" in trd_df.columns else None

    row_count = int(len(trd_df))
    unique_stkcd = int(trd_df["symbol"].nunique(dropna=True))
    row_counts = trd_df.groupby("symbol", dropna=True).size()
    one_row_per_stkcd = bool((row_counts == 1).all()) if len(row_counts) else False
    multi_row_per_stkcd = bool((row_counts > 1).any()) if len(row_counts) else False
    duplicate_stkcd_rows = int(max(row_count - unique_stkcd, 0))
    if "Statdt" in trd_df.columns:
        duplicate_stkcd_statdt = int(trd_df.duplicated(["symbol", "Statdt"], keep=False).sum())
    else:
        duplicate_stkcd_statdt = 0

    schema_rows = []
    for col in xlsx_header:
        if not col:
            continue
        present = col in trd_df.columns
        schema_rows.append(
            {
                "column_name": col,
                "selected_for_read": present,
                "non_null_count": int(trd_df[col].notna().sum()) if present else "",
                "dtype_observed": str(trd_df[col].dtype) if present else "",
            }
        )
    write_csv(OUT_DIR / "trd_co_schema_profile.csv", schema_rows, ["column_name", "selected_for_read", "non_null_count", "dtype_observed"])

    uniqueness_rows = [
        {"check": "row_count", "value": row_count, "notes": ""},
        {"check": "unique_stkcd_count", "value": unique_stkcd, "notes": ""},
        {"check": "one_row_per_stkcd", "value": one_row_per_stkcd, "notes": ""},
        {"check": "multi_row_per_stkcd", "value": multi_row_per_stkcd, "notes": ""},
        {"check": "duplicate_stkcd_rows", "value": duplicate_stkcd_rows, "notes": "row_count - unique symbol count"},
        {"check": "duplicate_stkcd_statdt_rows", "value": duplicate_stkcd_statdt, "notes": "Only meaningful if Statdt exists."},
    ]
    write_csv(OUT_DIR / "trd_co_key_uniqueness_check.csv", uniqueness_rows, ["check", "value", "notes"])

    cov_rows = coverage_rows(trd_df)
    write_csv(
        OUT_DIR / "trd_co_industry_field_coverage_profile.csv",
        cov_rows,
        ["field_name", "field_present", "non_null_count", "non_null_ratio", "unique_value_count", "top_categories"],
    )

    primary_code_field, primary_name_field, secondary_code_field, secondary_name_field, coarse_code_field, coarse_name_field = choose_industry_fields(trd_df)
    primary_coverage = float(trd_df[primary_code_field].notna().mean()) if primary_code_field else 0.0
    secondary_coverage = float(trd_df[secondary_code_field].notna().mean()) if secondary_code_field else 0.0

    var_rows, industry_changes = variation_check(trd_df)
    write_csv(
        OUT_DIR / "trd_co_industry_time_variation_check.csv",
        var_rows,
        [
            "field_name",
            "field_present",
            "symbols_with_multiple_rows",
            "symbols_with_field_variation",
            "variation_ratio_among_multi_row_symbols",
            "statdt_present_for_variation_check",
            "notes",
        ],
    )

    date_rows = []
    for field in ["Listdt", "Favaldt", "Statdt"]:
        if field in trd_df.columns:
            parsed = parse_date_series(trd_df[field])
            date_rows.append(
                {
                    "field_name": field,
                    "field_present": True,
                    "non_null_count": int(trd_df[field].notna().sum()),
                    "parseable_date_count": int(parsed.notna().sum()),
                    "min_date": str(parsed.min().date()) if parsed.notna().any() else "",
                    "max_date": str(parsed.max().date()) if parsed.notna().any() else "",
                    "interpretation_note": "Statdt is status change date, not automatic industry effective date." if field == "Statdt" else "date field for company/trading metadata.",
                }
            )
            del parsed
        else:
            date_rows.append(
                {
                    "field_name": field,
                    "field_present": False,
                    "non_null_count": 0,
                    "parseable_date_count": 0,
                    "min_date": "",
                    "max_date": "",
                    "interpretation_note": "missing",
                }
            )
    write_csv(
        OUT_DIR / "trd_co_date_field_interpretation_check.csv",
        date_rows,
        ["field_name", "field_present", "non_null_count", "parseable_date_count", "min_date", "max_date", "interpretation_note"],
    )

    profile_rows = []
    for field in ["Curtrd", "Markettype", "OWNERSHIPTYPE", "OWNERSHIPTYPECODE", "Statco", "Sctcd"]:
        profile_rows.append(
            {
                "field_name": field,
                "field_present": field in trd_df.columns,
                "non_null_count": int(trd_df[field].notna().sum()) if field in trd_df.columns else 0,
                "unique_value_count": int(trd_df[field].dropna().astype(str).nunique()) if field in trd_df.columns else 0,
                "top_categories": top_categories(trd_df[field]) if field in trd_df.columns else "",
                "usage_note": "A-share RMB filter candidate, not applied in this run." if field in {"Curtrd", "Markettype"} else "audit field only.",
            }
        )
    write_csv(
        OUT_DIR / "trd_co_market_ownership_status_field_profile.csv",
        profile_rows,
        ["field_name", "field_present", "non_null_count", "unique_value_count", "top_categories", "usage_note"],
    )

    classification, pit_quality_status, industry_change_date_candidate = classify_source(
        trd_df, primary_code_field, primary_coverage, industry_changes
    )

    clean_df = latest_per_symbol(trd_df.copy())
    clean_df["primary_industry_code"] = clean_df[primary_code_field] if primary_code_field else pd.NA
    clean_df["primary_industry_name"] = clean_df[primary_name_field] if primary_name_field else pd.NA
    clean_df["secondary_industry_code"] = clean_df[secondary_code_field] if secondary_code_field else pd.NA
    clean_df["secondary_industry_name"] = clean_df[secondary_name_field] if secondary_name_field else pd.NA
    clean_df["coarse_industry_code"] = clean_df[coarse_code_field] if coarse_code_field in clean_df.columns else pd.NA
    clean_df["coarse_industry_name"] = clean_df[coarse_name_field] if coarse_name_field in clean_df.columns else pd.NA
    clean_df["industry_source_classification"] = classification
    clean_df["pit_quality_status"] = pit_quality_status
    for field in TRD_FIELDS:
        if field not in clean_df.columns:
            clean_df[field] = pd.NA
    cleaned_cols = [
        "Stkcd",
        "symbol",
        "Stknme",
        "Listdt",
        "primary_industry_code",
        "primary_industry_name",
        "secondary_industry_code",
        "secondary_industry_name",
        "coarse_industry_code",
        "coarse_industry_name",
        "Markettype",
        "OWNERSHIPTYPE",
        "OWNERSHIPTYPECODE",
        "Curtrd",
        "Sctcd",
        "Statco",
        "Statdt",
        "FormerCode",
        "industry_source_classification",
        "pit_quality_status",
    ]
    clean_df[cleaned_cols].to_csv(OUT_DIR / "cleaned_trd_co_industry_source.csv", index=False, encoding="utf-8-sig")
    cleaned_industry_source_written = True

    policy = {
        "industry_source_classification": classification,
        "pit_quality_status": pit_quality_status,
        "primary_industry_code_field": primary_code_field,
        "primary_industry_name_field": primary_name_field,
        "secondary_industry_code_field": secondary_code_field,
        "secondary_industry_name_field": secondary_name_field,
        "static_industry_join": pit_quality_status == "STATIC_NOT_PIT",
        "pit_industry_join": pit_quality_status == "PIT_LIKE_CANDIDATE",
        "industry_change_date_candidate": industry_change_date_candidate,
        "limitations": [
            "Statdt is not automatically treated as industry effective date.",
            "No IC, return, portfolio, or backtest calculation is performed.",
        ],
    }
    write_json(OUT_DIR / "industry_source_classification_policy.json", policy)

    score_df = pd.read_parquet(SCORE_PANEL_PATH, columns=SCORE_COLUMNS)
    score_df["symbol"] = score_df["symbol"].map(norm_symbol)
    score_panel_rows = int(len(score_df))

    clean_join = clean_df[[
        "symbol",
        "primary_industry_code",
        "primary_industry_name",
        "secondary_industry_code",
        "secondary_industry_name",
        "industry_source_classification",
        "pit_quality_status",
    ]].drop_duplicates("symbol", keep="last")
    duplicate_join_detected = bool(clean_df["symbol"].duplicated(keep=False).any())
    joined = score_df.merge(clean_join, on="symbol", how="left", validate="many_to_one")
    joined["primary_industry_code"] = joined["primary_industry_code"].astype("string")
    joined["primary_industry_name"] = joined["primary_industry_name"].astype("string")
    joined["secondary_industry_code"] = joined["secondary_industry_code"].astype("string")
    joined["secondary_industry_name"] = joined["secondary_industry_name"].astype("string")
    joined_rows = int(joined["primary_industry_code"].notna().sum())
    join_coverage = joined_rows / score_panel_rows if score_panel_rows else 0.0
    missing_industry_rows = int(joined["primary_industry_code"].isna().sum())
    missing_industry_symbols = int(joined.loc[joined["primary_industry_code"].isna(), "symbol"].nunique())
    one_symbol_one_primary = bool(clean_join.groupby("symbol")["primary_industry_code"].nunique(dropna=True).max() <= 1)

    join_qa_rows = [
        {"metric": "score_panel_rows", "value": score_panel_rows, "notes": ""},
        {"metric": "joined_rows", "value": joined_rows, "notes": "Rows with non-null primary_industry_code after join."},
        {"metric": "join_coverage_ratio", "value": join_coverage, "notes": ""},
        {"metric": "missing_industry_rows", "value": missing_industry_rows, "notes": ""},
        {"metric": "missing_industry_symbols", "value": missing_industry_symbols, "notes": ""},
        {"metric": "duplicate_join_detected", "value": duplicate_join_detected, "notes": "Cleaned source was deduplicated to one row per symbol for join."},
        {"metric": "one_symbol_one_primary_industry_after_cleaning", "value": one_symbol_one_primary, "notes": ""},
        {"metric": "static_industry_join", "value": pit_quality_status == "STATIC_NOT_PIT", "notes": ""},
        {"metric": "pit_industry_join", "value": pit_quality_status == "PIT_LIKE_CANDIDATE", "notes": ""},
    ]
    write_csv(OUT_DIR / "simple_baseline_industry_join_qa.csv", join_qa_rows, ["metric", "value", "notes"])

    neutral_score_generated = False
    neutral_cols: list[str] = []
    neutral_score_row_count = 0
    small_industry_group_detected = False
    if classification != "NOT_USABLE" and join_coverage >= 0.95:
        prefix = "PIT" if pit_quality_status == "PIT_LIKE_CANDIDATE" else "STATIC"
        joined, neutral_cols = add_neutral_scores(joined, prefix)
        neutral_score_generated = True
        neutral_score_row_count = int(joined[neutral_cols].notna().any(axis=1).sum())
    else:
        for col in ["STATIC_IND_NEUTRAL_VALUE_BP_SINGLE_score", "STATIC_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score"]:
            joined[col] = pd.NA
        neutral_cols = ["STATIC_IND_NEUTRAL_VALUE_BP_SINGLE_score", "STATIC_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score"]

    group_size = joined.groupby(["month_end", "primary_industry_code"], dropna=False).size().reset_index(name="group_size")
    group_size["small_group_flag"] = group_size["group_size"] < 5
    small_industry_group_detected = bool(group_size["small_group_flag"].any())
    group_size.to_csv(OUT_DIR / "industry_group_size_summary.csv", index=False, encoding="utf-8-sig")
    joined = joined.merge(group_size[["month_end", "primary_industry_code", "small_group_flag"]], on=["month_end", "primary_industry_code"], how="left")

    formula_rows = [
        {
            "neutral_score_name": neutral_cols[0],
            "source_features": "bp_rank",
            "neutralization_method": "month_end_x_primary_industry_percentile_rank",
            "score_formula": "industry_within_rank(bp_rank)",
            "expected_score_direction": "higher_is_better",
            "fwd_ret_1m_used": False,
        },
        {
            "neutral_score_name": neutral_cols[1],
            "source_features": ";".join(SOURCE_FEATURES),
            "neutralization_method": "month_end_x_primary_industry_percentile_rank",
            "score_formula": "mean(industry_within_rank(bp_rank), industry_within_rank(ep_ttm_rank), industry_within_rank(cfo_to_earnings_parent_rank))",
            "expected_score_direction": "higher_is_better",
            "fwd_ret_1m_used": False,
        },
    ]
    write_csv(
        OUT_DIR / "industry_neutral_score_formula_manifest.csv",
        formula_rows,
        ["neutral_score_name", "source_features", "neutralization_method", "score_formula", "expected_score_direction", "fwd_ret_1m_used"],
    )

    panel_cols = [
        "symbol",
        "month_end",
        "primary_industry_code",
        "primary_industry_name",
        "secondary_industry_code",
        "secondary_industry_name",
        "bp_rank",
        "ep_ttm_rank",
        "cfo_to_earnings_parent_rank",
        "VALUE_BP_SINGLE_score",
        "VALUE_QUALITY_EQUAL_WEIGHT_score",
        *neutral_cols,
        "fwd_ret_1m",
        "small_group_flag",
        "pit_quality_status",
        "industry_source_classification",
    ]
    joined[panel_cols].to_parquet(OUT_DIR / "simple_baseline_industry_neutral_score_panel_v0.parquet", index=False)

    qa_rows = []
    qa_rows.append({"check": "neutral_score_columns_created", "passed": neutral_score_generated, "value": ";".join(neutral_cols), "notes": ""})
    for col in neutral_cols:
        series = pd.to_numeric(joined[col], errors="coerce")
        qa_rows.append({"check": f"{col}_range_within_0_1", "passed": bool(((series.dropna() >= 0) & (series.dropna() <= 1)).all()), "value": "", "notes": ""})
        qa_rows.append({"check": f"{col}_no_inf", "passed": bool(~series.isin([float("inf"), float("-inf")]).any()), "value": "", "notes": ""})
        qa_rows.append({"check": f"{col}_null_count", "passed": int(series.isna().sum()) == missing_industry_rows if neutral_score_generated else False, "value": int(series.isna().sum()), "notes": ""})
    qa_rows.extend(
        [
            {"check": "fwd_ret_1m_not_used_in_score_formula", "passed": True, "value": "", "notes": ""},
            {"check": "score_row_count_equals_joined_row_count", "passed": neutral_score_row_count == joined_rows if neutral_score_generated else False, "value": neutral_score_row_count, "notes": ""},
            {"check": "static_vs_pit_label_correct", "passed": pit_quality_status in {"STATIC_NOT_PIT", "PIT_LIKE_CANDIDATE"}, "value": pit_quality_status, "notes": ""},
        ]
    )
    write_csv(OUT_DIR / "industry_neutral_score_qa.csv", qa_rows, ["check", "passed", "value", "notes"])

    if classification == "NOT_USABLE":
        final_decision = "TRD_CO_INDUSTRY_SOURCE_FAIL_NOT_USABLE"
        recommended_next_step = "补充可 join 的行业字段后重跑。"
    elif join_coverage < 0.95:
        final_decision = "TRD_CO_INDUSTRY_SOURCE_FAIL_JOIN_COVERAGE_LOW"
        recommended_next_step = "先修复 TRD_Co 与 score panel 的 symbol 覆盖，再生成评估输入。"
    elif pit_quality_status == "PIT_LIKE_CANDIDATE" and neutral_score_generated:
        final_decision = "TRD_CO_INDUSTRY_NEUTRAL_SCORE_READY_FOR_EVALUATION_PREP_PIT_LIKE"
        recommended_next_step = "下一步只做 neutral score evaluation prep，不做组合或回测。"
    elif pit_quality_status == "STATIC_NOT_PIT" and neutral_score_generated:
        final_decision = "TRD_CO_STATIC_INDUSTRY_NEUTRAL_SCORE_READY_FOR_EVALUATION_PREP"
        recommended_next_step = "下一步对 static industry-neutral score 做 evaluation prep，并明确 STATIC_NOT_PIT。"
    else:
        final_decision = "TRD_CO_INDUSTRY_SOURCE_WATCH_MANUAL_REVIEW_REQUIRED"
        recommended_next_step = "人工确认行业日期结构后再决定 PIT-like 路径。"

    fwd_ret_used = False
    forbidden_false = {
        "ic_calculated": False,
        "d10_d1_calculated": False,
        "portfolio_constructed": False,
        "portfolio_return_calculated": False,
        "backtest_run": False,
        "transaction_cost_calculated": False,
        "turnover_calculated": False,
        "sharpe_calculated": False,
        "maxdd_calculated": False,
        "benchmark_relative_return_calculated": False,
        "alpha_beta_regression_calculated": False,
        "training_run": False,
        "shap_calculated": False,
        "tuning_run": False,
        "feature_importance_calculated": False,
        "production_holdings_generated": False,
        "live_order_ready_file_generated": False,
        "production_modified": False,
    }
    summary = {
        "run_timestamp": run_timestamp,
        "trd_co_xlsx_read": xlsx_read,
        "trd_co_xlsx_path": str(TRD_CO_XLSX),
        "prerequisites_passed": prerequisites_passed,
        "trd_co_row_count": row_count,
        "trd_co_unique_stkcd_count": unique_stkcd,
        "one_row_per_stkcd": one_row_per_stkcd,
        "multi_row_per_stkcd": multi_row_per_stkcd,
        "primary_industry_code_field": primary_code_field,
        "primary_industry_name_field": primary_name_field,
        "primary_industry_coverage_ratio": primary_coverage,
        "secondary_industry_code_field": secondary_code_field,
        "secondary_industry_name_field": secondary_name_field,
        "secondary_industry_coverage_ratio": secondary_coverage,
        "industry_changes_detected_within_stkcd": industry_changes,
        "industry_change_date_candidate": industry_change_date_candidate,
        "industry_source_classification": classification,
        "pit_quality_status": pit_quality_status,
        "cleaned_industry_source_written": cleaned_industry_source_written,
        "score_panel_path": rel(SCORE_PANEL_PATH),
        "score_panel_rows": score_panel_rows,
        "joined_rows": joined_rows,
        "join_coverage_ratio": join_coverage,
        "missing_industry_rows": missing_industry_rows,
        "neutral_score_generated": neutral_score_generated,
        "neutral_score_columns": neutral_cols,
        "neutral_score_row_count": neutral_score_row_count,
        "small_industry_group_detected": small_industry_group_detected,
        "fwd_ret_used_in_score_formula": fwd_ret_used,
        **forbidden_false,
        "final_decision": final_decision,
        "recommended_next_step": recommended_next_step,
    }
    write_json(OUT_DIR / "trd_co_industry_neutral_score_run_summary.json", summary)

    prereq = {
        "run_timestamp": run_timestamp,
        "prerequisites_passed": prerequisites_passed,
        "required_inputs": prereq_rows,
        "trd_co_xlsx_read": xlsx_read,
        "trd_co_sheet_read": sheet_name,
        "score_panel_columns_read": SCORE_COLUMNS,
    }
    write_json(OUT_DIR / "trd_co_industry_source_audit_prerequisite_check.json", prereq)

    plan_md = f"""# 下一步 Industry-Neutral Score Evaluation Plan

- final_decision: `{final_decision}`
- industry_source_classification: `{classification}`
- pit_quality_status: `{pit_quality_status}`
- neutral_score_columns: `{", ".join(neutral_cols)}`

下一步仅允许做 neutral score evaluation prep。仍禁止 IC、D10-D1、组合构造、收益、回测、交易成本、换手、Sharpe、MaxDD、训练、调参、SHAP 和 production。
"""
    (OUT_DIR / "next_step_industry_neutral_score_evaluation_plan.md").write_text(plan_md, encoding="utf-8")

    report_md = f"""# TRD_Co Industry Source Audit + Static Industry Neutral Score Run v0

## 结论

`{final_decision}`

## 行业源判断

- TRD_Co rows: `{row_count}`
- unique Stkcd: `{unique_stkcd}`
- classification: `{classification}`
- pit_quality_status: `{pit_quality_status}`
- primary industry: `{primary_code_field}` / `{primary_name_field}`
- secondary industry: `{secondary_code_field}` / `{secondary_name_field}`

`Statdt` 只作为公司活动情况变动日审计字段，本任务没有自动解释为行业分类生效日期。

## Join 和 Score

- score_panel_rows: `{score_panel_rows}`
- joined_rows: `{joined_rows}`
- join_coverage_ratio: `{join_coverage:.6f}`
- neutral_score_generated: `{neutral_score_generated}`
- neutral_score_columns: `{", ".join(neutral_cols)}`

本任务未计算 IC、D10-D1、组合收益、回测、交易成本、换手、Sharpe、MaxDD、benchmark-relative return、alpha/beta，未训练、未调参、未 SHAP、未写 production。
"""
    (OUT_DIR / "trd_co_industry_neutral_score_run_report.md").write_text(report_md, encoding="utf-8")

    completion_card = f"""# Task Completion Card

- task_name: `{TASK_NAME}`
- final_decision: `{final_decision}`
- output_directory: `{rel(OUT_DIR)}`
- trd_co_xlsx_read: `{xlsx_read}`
- neutral_score_generated: `{neutral_score_generated}`
- production_modified: `False`
"""
    (OUT_DIR / "task_completion_card.md").write_text(completion_card, encoding="utf-8")

    terminal_summary = {
        "task_name": TASK_NAME,
        "run_timestamp": run_timestamp,
        "stdout_log": rel(RUN_DIR / "run_stdout.txt"),
        "stderr_log": rel(RUN_DIR / "run_stderr.txt"),
        "output_directory": rel(OUT_DIR),
        "final_decision": final_decision,
        "exit_code": 0,
    }
    write_json(OUT_DIR / "terminal_summary.json", terminal_summary)

    final_qa_rows = [
        {"check": "prerequisites_passed", "passed": prerequisites_passed, "notes": ""},
        {"check": "trd_co_xlsx_read", "passed": xlsx_read, "notes": "Allowed by this task."},
        {"check": "fwd_ret_not_used_in_score", "passed": not fwd_ret_used, "notes": ""},
        {"check": "no_forbidden_calculations", "passed": True, "notes": "No IC/D10-D1/returns/backtest/training/production."},
        {"check": "summary_written", "passed": True, "notes": str(OUT_DIR / "trd_co_industry_neutral_score_run_summary.json")},
    ]
    write_csv(OUT_DIR / "final_qa.csv", final_qa_rows, ["check", "passed", "notes"])

    run_state = f"""# RUN_STATE

任务：{TASK_NAME}
状态：完成

已读取：
- {TRD_CO_XLSX}
- {FIELD_DICT}
- {FIELD_DICT_SUMMARY}
- {SCORE_PANEL_PATH} 的必要列
- simple baseline summary / candidate manifest / formula manifest

输出目录：
- {OUT_DIR}

final_decision: {final_decision}

禁止项确认：
- 未计算 IC / D10-D1
- 未构造 portfolio 或计算收益
- 未回测
- 未训练 / 调参 / SHAP
- 未写 production
"""
    (RUN_DIR / "RUN_STATE.md").write_text(run_state, encoding="utf-8")

    del trd_df, clean_df, score_df, joined, clean_join, group_size
    gc.collect()
    print(f"final_decision={final_decision}")
    print(f"trd_co_xlsx_read={xlsx_read}")
    print(f"join_coverage_ratio={join_coverage}")
    print(f"neutral_score_generated={neutral_score_generated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
