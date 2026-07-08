from __future__ import annotations

import csv
import gc
import json
import os
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

warnings.filterwarnings("ignore", message="Workbook contains no default style.*")


ROOT = Path(r"C:\dev\quant")
TASK_NAME = "historical_industry_source_gap_resolution_v0"
OUT_DIR = ROOT / "output" / TASK_NAME
RUN_DIR = ROOT / "output" / "_agent_runs" / TASK_NAME

FORENSICS_DIR = ROOT / "output" / "trd_co_static_industry_join_forensics_v0"
STATIC_DIR = ROOT / "output" / "trd_co_static_industry_neutral_score_run_v1"
SCORE_PANEL = ROOT / "output" / "simple_baseline_score_run_v0" / "simple_baseline_score_panel_v0.parquet"

REQUIRED_INPUTS = [
    FORENSICS_DIR / "trd_co_static_industry_join_forensics_summary.json",
    FORENSICS_DIR / "join_method_comparison.csv",
    FORENSICS_DIR / "missing_symbol_profile.csv",
    FORENSICS_DIR / "missing_by_month.csv",
    FORENSICS_DIR / "trd_co_join_coverage_diagnosis.json",
    STATIC_DIR / "cleaned_trd_co_static_industry_source.csv",
    SCORE_PANEL,
]

SCAN_DIRS = [
    ROOT / "data" / "csmar_exports",
    ROOT / "data" / "raw" / "csmar",
    ROOT / "data" / "processed",
    ROOT / "data" / "raw" / "industry",
    ROOT / "data" / "processed" / "industry",
    ROOT / "output",
]

KEYWORDS = [
    "industry",
    "indcd",
    "indnme",
    "nnindcd",
    "nnindnme",
    "indcdzx",
    "indnmezx",
    "trd_co",
    "company",
    "stockinfo",
    "basic",
    "listed",
    "delist",
    "退市",
    "行业",
    "公司",
    "证券基本",
]

SKIP_DIR_NAMES = {".git", ".venv", "__pycache__", "logs", "cache", "xhs", "_agent_runs"}
DATA_SUFFIXES = {".csv", ".txt", ".tsv", ".parquet", ".xlsx", ".xlsm", ".xls"}
SYMBOL_CANDIDATES = ["symbol", "stkcd", "stockcode", "stock_code", "secu_code", "证券代码", "股票代码", "Stkcd"]
IND_CODE_CANDIDATES = ["indcd", "nnindcd", "nindcd", "indcdzx", "industry_code", "行业代码"]
IND_NAME_CANDIDATES = ["indnme", "nnindnme", "nindnme", "indnmezx", "industry_name", "行业名称", "行业"]
DATE_CANDIDATES = ["month_end", "statdt", "listdt", "date", "trdmnt", "enddate", "annodt", "截止日期", "日期"]


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def as_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def norm_symbol(value: Any) -> str | None:
    text = as_text(value).upper()
    if not text:
        return None
    if "." in text:
        text = text.split(".", 1)[0]
    digits = re.findall(r"\d+", text)
    if not digits:
        return None
    code = digits[0]
    if len(code) > 6:
        code = code[-6:]
    return code.zfill(6)


def code_prefix(symbol: Any) -> str:
    code = norm_symbol(symbol)
    return code[:2] if code else "NA"


def filename_matches(path: Path) -> bool:
    lower = path.name.lower()
    if any(k in path.name for k in ["退市", "行业", "公司", "证券基本"]):
        return True
    return any(k in lower for k in KEYWORDS if k.isascii())


def csv_header(path: Path) -> list[str]:
    with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        return [str(x).strip() for x in next(reader, [])]


def parquet_columns(path: Path) -> tuple[list[str], int | None]:
    import pyarrow.parquet as pq

    pf = pq.ParquetFile(path)
    cols = list(pf.schema.names)
    rows = pf.metadata.num_rows if pf.metadata else None
    del pf
    gc.collect()
    return cols, rows


def excel_header(path: Path) -> tuple[list[str], int | None]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.max_row
    header = [as_text(v) for v in next(ws.iter_rows(max_row=1, values_only=True), ())]
    wb.close()
    del wb, ws
    gc.collect()
    return header, rows


def find_col(columns: list[str], candidates: list[str]) -> str | None:
    lower_map = {c.lower(): c for c in columns}
    for candidate in candidates:
        cand = candidate.lower()
        if cand in lower_map:
            return lower_map[cand]
    for candidate in candidates:
        cand = candidate.lower()
        for lower, original in lower_map.items():
            if cand not in {"symbol", "stkcd"} and cand in lower:
                return original
    return None


def lightweight_symbols_with_industry(
    path: Path,
    source_type: str,
    symbol_col: str | None,
    ind_code_col: str | None,
    ind_name_col: str | None,
) -> set[str] | None:
    if not symbol_col or not (ind_code_col or ind_name_col):
        return None
    size = path.stat().st_size
    usecols = list(dict.fromkeys([c for c in [symbol_col, ind_code_col, ind_name_col] if c]))
    try:
        if source_type == "CSV" and size <= 50 * 1024 * 1024:
            df = pd.read_csv(path, dtype=str, usecols=usecols)
        elif source_type == "PARQUET" and size <= 200 * 1024 * 1024:
            df = pd.read_parquet(path, columns=usecols)
        elif source_type in {"XLSX", "XLSM"} and size <= 25 * 1024 * 1024:
            df = pd.read_excel(path, dtype=str, header=0, usecols=usecols)
        else:
            return None
        industry_mask = pd.Series(False, index=df.index)
        for col in [ind_code_col, ind_name_col]:
            if col and col in df.columns:
                industry_mask = industry_mask | df[col].notna()
        df = df.loc[industry_mask].copy()
        symbols = {s for s in df[symbol_col].map(norm_symbol).dropna().astype(str).tolist()}
        del df
        gc.collect()
        return symbols
    except Exception:
        return None


def inspect_candidate(path: Path, missing_symbols: set[str], total_symbols: set[str], trd_symbols: set[str]) -> tuple[dict[str, Any], dict[str, Any]]:
    suffix = path.suffix.lower()
    source_type = suffix.lstrip(".").upper() or "UNKNOWN"
    columns: list[str] = []
    row_count: int | None = None
    notes: list[str] = []
    try:
        if suffix in {".csv", ".txt", ".tsv"}:
            columns = csv_header(path)
        elif suffix == ".parquet":
            columns, row_count = parquet_columns(path)
        elif suffix in {".xlsx", ".xlsm"}:
            columns, row_count = excel_header(path)
        elif suffix == ".xls":
            notes.append("Legacy xls schema not read; manual review required.")
        else:
            notes.append("Unsupported suffix.")
    except Exception as exc:  # noqa: BLE001
        notes.append(f"schema/head check failed: {type(exc).__name__}: {exc}")

    symbol_col = find_col(columns, SYMBOL_CANDIDATES)
    ind_code_col = find_col(columns, IND_CODE_CANDIDATES)
    ind_name_col = find_col(columns, IND_NAME_CANDIDATES)
    date_col = find_col(columns, DATE_CANDIDATES)
    has_symbol = bool(symbol_col)
    has_code = bool(ind_code_col)
    has_name = bool(ind_name_col)
    has_date = bool(date_col)
    likely = "HISTORICAL_CANDIDATE" if has_symbol and (has_code or has_name) and has_date else "STATIC_CANDIDATE" if has_symbol and (has_code or has_name) else "MANUAL_REVIEW_REQUIRED"

    symbols = lightweight_symbols_with_industry(path, source_type, symbol_col, ind_code_col, ind_name_col)
    checked_overlap = symbols is not None
    overlap_missing = len(missing_symbols & symbols) if symbols is not None else 0
    overlap_total = len(total_symbols & symbols) if symbols is not None else 0
    estimated = (len(trd_symbols & total_symbols) + overlap_missing) / len(total_symbols) if total_symbols else 0.0
    can_cover = overlap_missing > 0 if checked_overlap else False
    scan_row = {
        "source_path": rel(path),
        "source_type": source_type,
        "row_count_if_lightweight_available": row_count if row_count is not None else "",
        "columns_found": ";".join(columns[:80]),
        "has_symbol_or_stkcd": has_symbol,
        "has_industry_code": has_code,
        "has_industry_name": has_name,
        "has_date_or_status_date": has_date,
        "likely_static_or_historical": likely,
        "can_cover_missing_symbols": can_cover,
        "checked_missing_symbol_overlap_count": overlap_missing if checked_overlap else "",
        "notes": " | ".join(notes),
    }
    overlap_row = {
        "source_path": rel(path),
        "source_type": source_type,
        "symbol_column": symbol_col or "",
        "industry_code_column": ind_code_col or "",
        "industry_name_column": ind_name_col or "",
        "date_column": date_col or "",
        "overlap_with_missing_symbols": overlap_missing if checked_overlap else "",
        "overlap_ratio_over_missing_symbols": overlap_missing / len(missing_symbols) if checked_overlap and missing_symbols else "",
        "overlap_with_total_score_symbols": overlap_total if checked_overlap else "",
        "estimated_join_coverage_if_used": estimated if checked_overlap else "",
        "manual_review_required": not checked_overlap,
        "notes": "" if checked_overlap else "Unable or unsafe to read symbol column lightly.",
    }
    return scan_row, overlap_row


def possible_reason(symbol: str, first_month: pd.Timestamp, trd_symbols: set[str], missing_profile_reason: str | None) -> str:
    if symbol in trd_symbols:
        return "FORMER_CODE_OR_CODE_CHANGE_NEEDED"
    if missing_profile_reason == "FORMER_CODE_MAPPING_NEEDED":
        return "FORMER_CODE_OR_CODE_CHANGE_NEEDED"
    if symbol[:2] not in {"00", "30", "60", "68", "83", "87", "43", "92"}:
        return "NON_A_SHARE_OR_SYNTHETIC"
    if pd.notna(first_month) and first_month.year < 2015:
        return "DELISTED_OR_HISTORICAL_CONSTITUENT"
    return "MISSING_FROM_CURRENT_TRD_CO"


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    run_timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    print(f"[{run_timestamp}] start {TASK_NAME}")

    prereq_rows = [
        {"path": rel(path), "exists": path.exists(), "bytes": path.stat().st_size if path.exists() else None}
        for path in REQUIRED_INPUTS
    ]
    prerequisites_passed = all(row["exists"] for row in prereq_rows)
    if not prerequisites_passed:
        summary = {
            "run_timestamp": run_timestamp,
            "prerequisites_passed": False,
            "final_decision": "HISTORICAL_INDUSTRY_SOURCE_GAP_RESOLUTION_FAIL",
            "recommended_next_step": "补齐缺失输入后重跑。",
        }
        write_json(OUT_DIR / "historical_industry_source_gap_resolution_summary.json", summary)
        return 1

    forensics = read_json(FORENSICS_DIR / "trd_co_static_industry_join_forensics_summary.json")
    diagnosis = read_json(FORENSICS_DIR / "trd_co_join_coverage_diagnosis.json")
    missing_profile_prev = pd.read_csv(FORENSICS_DIR / "missing_symbol_profile.csv", dtype=str)
    prev_reason_map = dict(zip(missing_profile_prev["normalized_symbol"], missing_profile_prev["missing_reason_guess"]))
    cleaned = pd.read_csv(STATIC_DIR / "cleaned_trd_co_static_industry_source.csv", dtype=str, usecols=["symbol"])
    trd_symbols = {s for s in cleaned["symbol"].map(norm_symbol).dropna().astype(str).tolist()}
    score = pd.read_parquet(SCORE_PANEL, columns=["symbol", "month_end"])
    score["symbol"] = score["symbol"].map(norm_symbol)
    score["month_end"] = pd.to_datetime(score["month_end"], errors="coerce")
    score_panel_unique_symbols = int(score["symbol"].nunique(dropna=True))
    total_symbols = set(score["symbol"].dropna().astype(str).tolist())
    missing_score = score.loc[~score["symbol"].isin(trd_symbols)].copy()
    missing_symbols = set(missing_score["symbol"].dropna().astype(str).tolist())
    missing_unique_symbol_count = len(missing_symbols)
    missing_row_count = int(len(missing_score))
    missing_row_ratio = missing_row_count / len(score) if len(score) else 0.0

    master_rows = []
    for symbol, group in missing_score.groupby("symbol", dropna=True):
        first_month = group["month_end"].min()
        last_month = group["month_end"].max()
        reason = possible_reason(str(symbol), first_month, trd_symbols, prev_reason_map.get(str(symbol)))
        master_rows.append(
            {
                "symbol": symbol,
                "row_count_in_score_panel": int(len(group)),
                "first_month": str(first_month.date()) if pd.notna(first_month) else "",
                "last_month": str(last_month.date()) if pd.notna(last_month) else "",
                "active_month_count": int(group["month_end"].nunique(dropna=True)),
                "appears_in_trd_co": symbol in trd_symbols,
                "possible_reason": reason,
                "notes": "",
            }
        )
    master_rows = sorted(master_rows, key=lambda r: int(r["row_count_in_score_panel"]), reverse=True)
    write_csv(
        OUT_DIR / "missing_industry_symbol_master.csv",
        master_rows,
        ["symbol", "row_count_in_score_panel", "first_month", "last_month", "active_month_count", "appears_in_trd_co", "possible_reason", "notes"],
    )

    by_month = missing_score.groupby(missing_score["month_end"].dt.to_period("M")).size().rename("missing_rows").to_frame()
    total_month = score.groupby(score["month_end"].dt.to_period("M")).size().rename("total_rows")
    by_month = by_month.join(total_month, how="right").fillna(0)
    by_month["missing_rows"] = by_month["missing_rows"].astype(int)
    by_month["total_rows"] = by_month["total_rows"].astype(int)
    by_month["missing_ratio"] = by_month["missing_rows"] / by_month["total_rows"]
    by_month.reset_index().assign(month_end=lambda d: d["month_end"].astype(str))[["month_end", "total_rows", "missing_rows", "missing_ratio"]].to_csv(
        OUT_DIR / "missing_industry_by_month.csv", index=False, encoding="utf-8-sig"
    )
    by_year = missing_score.groupby(missing_score["month_end"].dt.year).size().rename("missing_rows").to_frame()
    total_year = score.groupby(score["month_end"].dt.year).size().rename("total_rows")
    by_year = by_year.join(total_year, how="right").fillna(0)
    by_year["missing_rows"] = by_year["missing_rows"].astype(int)
    by_year["total_rows"] = by_year["total_rows"].astype(int)
    by_year["missing_ratio"] = by_year["missing_rows"] / by_year["total_rows"]
    by_year.reset_index().rename(columns={"month_end": "year"}).to_csv(OUT_DIR / "missing_industry_by_year.csv", index=False, encoding="utf-8-sig")

    missing_score["code_prefix"] = missing_score["symbol"].map(code_prefix)
    prefix_rows = []
    for prefix, group in missing_score.groupby("code_prefix", dropna=False):
        prefix_rows.append(
            {
                "code_prefix": prefix,
                "row_count": int(len(group)),
                "unique_symbol_count": int(group["symbol"].nunique(dropna=True)),
                "notes": "A-share-like prefix" if prefix in {"00", "30", "60", "68", "83", "87", "43", "92"} else "review symbol type",
            }
        )
    write_csv(OUT_DIR / "missing_industry_by_code_prefix.csv", sorted(prefix_rows, key=lambda r: r["row_count"], reverse=True), ["code_prefix", "row_count", "unique_symbol_count", "notes"])

    candidates: list[Path] = []
    seen: set[Path] = set()
    for base in SCAN_DIRS:
        if not base.exists():
            continue
        for dirpath, dirnames, filenames in os.walk(base):
            current = Path(dirpath)
            try:
                current.relative_to(OUT_DIR)
                dirnames[:] = []
                continue
            except ValueError:
                pass
            dirnames[:] = [d for d in dirnames if d.lower() not in SKIP_DIR_NAMES]
            for filename in filenames:
                path = current / filename
                if path.suffix.lower() not in DATA_SUFFIXES:
                    continue
                if filename_matches(path) and path not in seen:
                    seen.add(path)
                    candidates.append(path)

    scan_rows: list[dict[str, Any]] = []
    overlap_rows: list[dict[str, Any]] = []
    for path in sorted(candidates):
        scan_row, overlap_row = inspect_candidate(path, missing_symbols, total_symbols, trd_symbols)
        scan_rows.append(scan_row)
        overlap_rows.append(overlap_row)

    scan_fields = [
        "source_path",
        "source_type",
        "row_count_if_lightweight_available",
        "columns_found",
        "has_symbol_or_stkcd",
        "has_industry_code",
        "has_industry_name",
        "has_date_or_status_date",
        "likely_static_or_historical",
        "can_cover_missing_symbols",
        "checked_missing_symbol_overlap_count",
        "notes",
    ]
    write_csv(OUT_DIR / "local_industry_source_candidate_scan.csv", scan_rows, scan_fields)
    overlap_fields = [
        "source_path",
        "source_type",
        "symbol_column",
        "industry_code_column",
        "industry_name_column",
        "date_column",
        "overlap_with_missing_symbols",
        "overlap_ratio_over_missing_symbols",
        "overlap_with_total_score_symbols",
        "estimated_join_coverage_if_used",
        "manual_review_required",
        "notes",
    ]
    write_csv(OUT_DIR / "local_industry_source_overlap_check.csv", overlap_rows, overlap_fields)

    checked_overlap_rows = [
        r
        for r in overlap_rows
        if r["estimated_join_coverage_if_used"] != "" and int(r["overlap_with_missing_symbols"] or 0) > 0
    ]
    if checked_overlap_rows:
        best_overlap = max(checked_overlap_rows, key=lambda r: float(r["estimated_join_coverage_if_used"]))
        best_candidate_path = best_overlap["source_path"]
        best_candidate_overlap_missing = int(best_overlap["overlap_with_missing_symbols"])
        best_candidate_estimated = float(best_overlap["estimated_join_coverage_if_used"])
    else:
        best_candidate_path = None
        best_candidate_overlap_missing = 0
        best_candidate_estimated = float(forensics.get("best_join_coverage_ratio", 0.0))

    historical_source_found = best_candidate_estimated >= 0.95
    csmar_reexport_needed = not historical_source_found
    download_rows = [
        {"required_field": "Stkcd", "reason": "join key for all score panel symbols including historical/delisted names", "priority": "HIGH"},
        {"required_field": "Stknme", "reason": "manual validation and issuer identity check", "priority": "MEDIUM"},
        {"required_field": "Indcd", "reason": "coarse industry fallback", "priority": "HIGH"},
        {"required_field": "Indnme", "reason": "coarse industry fallback name", "priority": "HIGH"},
        {"required_field": "IndcdZX", "reason": "primary static industry code candidate", "priority": "HIGH"},
        {"required_field": "IndnmeZX", "reason": "primary static industry name candidate", "priority": "HIGH"},
        {"required_field": "Nnindcd", "reason": "CSRC 2012 industry code candidate", "priority": "HIGH"},
        {"required_field": "Nnindnme", "reason": "CSRC 2012 industry name candidate", "priority": "HIGH"},
        {"required_field": "Listdt", "reason": "listing date and historical universe validation", "priority": "MEDIUM"},
        {"required_field": "Statco", "reason": "listing/status audit", "priority": "HIGH"},
        {"required_field": "Statdt", "reason": "status change date audit; not automatic industry effective date", "priority": "HIGH"},
        {"required_field": "FormerCode", "reason": "historical code mapping and join miss diagnosis", "priority": "HIGH"},
        {"required_field": "Markettype", "reason": "A-share market segment filtering", "priority": "MEDIUM"},
        {"required_field": "Curtrd", "reason": "RMB/A-share sample filter validation", "priority": "MEDIUM"},
    ]
    write_csv(OUT_DIR / "historical_industry_download_request.csv", download_rows, ["required_field", "reason", "priority"])
    download_symbol_rows = [
        {
            "symbol": r["symbol"],
            "first_month": r["first_month"],
            "last_month": r["last_month"],
            "row_count_in_score_panel": r["row_count_in_score_panel"],
            "reason": r["possible_reason"],
        }
        for r in master_rows
    ]
    write_csv(OUT_DIR / "missing_symbols_for_industry_download.csv", download_symbol_rows, ["symbol", "first_month", "last_month", "row_count_in_score_panel", "reason"])

    if best_candidate_estimated >= 0.95:
        final_decision = "HISTORICAL_INDUSTRY_SOURCE_READY_TO_RERUN_JOIN"
        recommended_next_step = "使用本地候选历史行业源重跑 static/historical industry join。"
    elif best_candidate_estimated >= 0.80:
        final_decision = "HISTORICAL_INDUSTRY_SOURCE_WATCH_PARTIAL_SOURCE_FOUND"
        recommended_next_step = "人工确认是否接受 universe loss 后再决定是否重跑。"
    else:
        final_decision = "HISTORICAL_INDUSTRY_SOURCE_NEEDS_CSMAR_REEXPORT"
        recommended_next_step = "按下载清单从 CSMAR 重新导出包含历史/退市/暂停上市/曾用代码样本的行业源。"

    policy = {
        "best_candidate_source_path": best_candidate_path,
        "best_candidate_estimated_join_coverage": best_candidate_estimated,
        "historical_source_found": historical_source_found,
        "csmar_reexport_needed": csmar_reexport_needed,
        "safe_to_generate_neutral_score_now": False,
        "download_requirement": "下载时必须包含历史/退市/暂停上市/曾用代码样本，不要只下载当前正常上市公司。目标覆盖 score panel 全部 1323 symbols，最低覆盖率 >= 0.95。",
        "recommended_next_step": recommended_next_step,
    }
    write_json(OUT_DIR / "historical_industry_source_resolution_policy.json", policy)

    forbidden_false = {
        "safe_to_generate_neutral_score_now": False,
        "neutral_score_generated": False,
        "ic_calculated": False,
        "d10_d1_calculated": False,
        "portfolio_constructed": False,
        "portfolio_return_calculated": False,
        "backtest_run": False,
        "training_run": False,
        "shap_calculated": False,
        "tuning_run": False,
        "production_modified": False,
    }
    summary = {
        "run_timestamp": run_timestamp,
        "prerequisites_passed": prerequisites_passed,
        "previous_best_join_coverage_ratio": float(forensics.get("best_join_coverage_ratio", 0.0)),
        "score_panel_unique_symbols": score_panel_unique_symbols,
        "previous_joined_unique_symbols": int(forensics.get("best_joined_unique_symbols", 0)),
        "missing_unique_symbol_count": missing_unique_symbol_count,
        "missing_row_count": missing_row_count,
        "missing_row_ratio": missing_row_ratio,
        "local_candidate_source_count": len(scan_rows),
        "best_candidate_source_path": best_candidate_path,
        "best_candidate_overlap_missing_symbols": best_candidate_overlap_missing,
        "best_candidate_estimated_join_coverage": best_candidate_estimated,
        "historical_source_found": historical_source_found,
        "csmar_reexport_needed": csmar_reexport_needed,
        "missing_symbols_download_list_generated": True,
        "download_request_generated": True,
        **forbidden_false,
        "final_decision": final_decision,
        "recommended_next_step": recommended_next_step,
    }
    write_json(OUT_DIR / "historical_industry_source_gap_resolution_summary.json", summary)

    prereq = {
        "run_timestamp": run_timestamp,
        "prerequisites_passed": prerequisites_passed,
        "required_inputs": prereq_rows,
        "score_panel_columns_read": ["symbol", "month_end"],
        "fwd_ret_1m_read": False,
    }
    write_json(OUT_DIR / "historical_industry_gap_prerequisite_check.json", prereq)

    report = f"""# Historical Industry Source Gap Resolution v0

## 结论

`{final_decision}`

## 缺口

- score_panel_unique_symbols: `{score_panel_unique_symbols}`
- previous_joined_unique_symbols: `{int(forensics.get("best_joined_unique_symbols", 0))}`
- missing_unique_symbol_count: `{missing_unique_symbol_count}`
- missing_row_count: `{missing_row_count}`
- missing_row_ratio: `{missing_row_ratio:.6f}`

## 本地候选源

- local_candidate_source_count: `{len(scan_rows)}`
- best_candidate_source_path: `{best_candidate_path}`
- best_candidate_overlap_missing_symbols: `{best_candidate_overlap_missing}`
- best_candidate_estimated_join_coverage: `{best_candidate_estimated:.6f}`

未找到足够覆盖率的本地历史行业源时，应按下载清单重新导出 CSMAR 行业源，必须包含历史/退市/暂停上市/曾用代码样本，不要只下载当前正常上市公司。

本任务未生成 neutral score，未计算 IC、D10-D1、收益、回测，未训练、调参、SHAP 或写 production。
"""
    (OUT_DIR / "historical_industry_source_gap_resolution_report.md").write_text(report, encoding="utf-8")

    next_plan = f"""# Next Step Historical Industry Source Plan

- final_decision: `{final_decision}`
- csmar_reexport_needed: `{csmar_reexport_needed}`
- safe_to_generate_neutral_score_now: `False`

下一步应优先补齐历史/退市/暂停上市/曾用代码样本的行业源。补齐后先重跑 join coverage audit，覆盖率达到 `>= 0.95` 后才允许进入 static/historical industry-neutral score run。
"""
    (OUT_DIR / "next_step_historical_industry_source_plan.md").write_text(next_plan, encoding="utf-8")

    terminal_summary = {
        "task_name": TASK_NAME,
        "run_timestamp": run_timestamp,
        "stdout_log": rel(RUN_DIR / "run_stdout.txt"),
        "stderr_log": rel(RUN_DIR / "run_stderr.txt"),
        "output_directory": rel(OUT_DIR),
        "final_decision": final_decision,
        "exit_code": 0,
    }
    write_json(OUT_DIR / "terminal_summary.json", terminal_summary)
    write_csv(
        OUT_DIR / "final_qa.csv",
        [
            {"check": "prerequisites_passed", "passed": prerequisites_passed, "notes": ""},
            {"check": "score_panel_only_symbol_month_read", "passed": True, "notes": ""},
            {"check": "download_lists_generated", "passed": True, "notes": ""},
            {"check": "no_forbidden_calculations", "passed": True, "notes": "No neutral score/IC/return/backtest/training/production."},
        ],
        ["check", "passed", "notes"],
    )
    (OUT_DIR / "task_completion_card.md").write_text(
        f"""# Task Completion Card

- task_name: `{TASK_NAME}`
- final_decision: `{final_decision}`
- output_directory: `{rel(OUT_DIR)}`
- neutral_score_generated: `False`
- production_modified: `False`
""",
        encoding="utf-8",
    )
    (RUN_DIR / "RUN_STATE.md").write_text(
        f"""# RUN_STATE

任务：{TASK_NAME}
状态：完成

输出目录：
- {OUT_DIR}

final_decision: {final_decision}
best_candidate_estimated_join_coverage: {best_candidate_estimated}

禁止项确认：
- 未生成 neutral score
- 未计算 IC / D10-D1 / 收益 / 回测
- 未训练 / 调参 / SHAP
- 未写 production
""",
        encoding="utf-8",
    )

    del score, missing_score, cleaned, missing_profile_prev
    gc.collect()
    print(f"final_decision={final_decision}")
    print(f"missing_unique_symbol_count={missing_unique_symbol_count}")
    print(f"best_candidate_estimated_join_coverage={best_candidate_estimated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
