from __future__ import annotations

import csv
import gc
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


TASK_NAME = "simple_baseline_industry_neutralization_prep_v0"
ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / TASK_NAME
RUN_DIR = ROOT / "output" / "_agent_runs" / TASK_NAME

SCORE_PANEL = ROOT / "output" / "simple_baseline_score_run_v0" / "simple_baseline_score_panel_v0.parquet"

REQUIRED_INPUTS = [
    ROOT / "output" / "simple_baseline_score_run_v0" / "simple_baseline_score_run_summary.json",
    ROOT / "output" / "simple_baseline_score_evaluation_run_v0" / "simple_baseline_score_evaluation_run_summary.json",
    ROOT / "output" / "simple_baseline_portfolio_construction_run_v0" / "simple_baseline_portfolio_construction_run_summary.json",
    ROOT / "output" / "simple_baseline_portfolio_evaluation_run_v0" / "simple_baseline_portfolio_evaluation_run_summary.json",
    ROOT / "output" / "simple_baseline_portfolio_review_v0" / "simple_baseline_portfolio_review_summary.json",
    ROOT / "output" / "simple_baseline_candidate_selection_v0" / "simple_baseline_candidate_selection_summary.json",
    ROOT / "output" / "simple_baseline_candidate_selection_v0" / "simple_baseline_candidate_manifest.csv",
    ROOT / "output" / "simple_baseline_score_prep_v0" / "simple_baseline_score_formula_manifest.csv",
    SCORE_PANEL,
    ROOT / "output" / "compact_f_v3_full_training_panel_price_label_v0" / "compact_f_v3_full_training_panel_price_label_summary.json",
    ROOT / "output" / "compact_f_v3_full_training_panel_price_label_v0" / "compact_f_v3_full_price_label_unique13_feature_list.txt",
    ROOT / "output" / "compact_f_v3_full_training_config_freeze_v0" / "frozen_feature_target_config.json",
]

SCAN_DIRS = [
    ROOT / "output",
    ROOT / "data" / "processed",
    ROOT / "data" / "benchmark",
    ROOT / "data" / "raw" / "benchmark",
    ROOT / "data" / "raw" / "csmar",
    ROOT / "data" / "raw" / "industry",
    ROOT / "data" / "processed" / "industry",
]

KEYWORDS = [
    "industry",
    "sw",
    "shenwan",
    "citic",
    "csrc",
    "sector",
    "ind",
    "classification",
    "行业",
    "申万",
    "中信",
]

DATA_SUFFIXES = {".parquet", ".csv", ".txt", ".tsv", ".xlsx", ".xlsm", ".xls"}

SKIP_DIR_NAMES = {
    ".git",
    ".venv",
    "__pycache__",
    "logs",
    "cache",
    "xhs",
}

SOURCE_FEATURES = ["bp_rank", "ep_ttm_rank", "cfo_to_earnings_parent_rank"]
NEUTRAL_SCORE_COLUMNS = [
    "IND_NEUTRAL_VALUE_BP_SINGLE_score",
    "IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score",
]
RAW_SCORE_COLUMNS = ["VALUE_BP_SINGLE_score", "VALUE_QUALITY_EQUAL_WEIGHT_score"]


def rel(path: Path | None) -> str | None:
    if path is None:
        return None
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def csv_header(path: Path) -> list[str]:
    with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        return next(reader, [])


def parquet_columns(path: Path) -> list[str]:
    import pyarrow.parquet as pq

    parquet_file = pq.ParquetFile(path)
    cols = list(parquet_file.schema.names)
    del parquet_file
    gc.collect()
    return cols


def excel_header_light(path: Path) -> list[str]:
    if path.stat().st_size > 25 * 1024 * 1024:
        return []

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row = next(ws.iter_rows(max_row=1, values_only=True), ())
    headers = [str(v) for v in row if v is not None]
    wb.close()
    del wb, ws, row
    gc.collect()
    return headers


def find_col(columns: list[str], groups: list[list[str]]) -> str:
    lower_map = {c.lower(): c for c in columns}
    for group in groups:
        for col_lower, original in lower_map.items():
            if any(token in col_lower for token in group):
                return original
    return ""


def guess_taxonomy(path: Path, columns: list[str]) -> str:
    text = (path.name + " " + " ".join(columns)).lower()
    if "shenwan" in text or "申万" in text or "sw" in text:
        return "SHENWAN_OR_SW_GUESS"
    if "citic" in text or "中信" in text:
        return "CITIC_GUESS"
    if "csrc" in text or "证监" in text:
        return "CSRC_GUESS"
    if "sector" in text:
        return "SECTOR_GUESS"
    if "industry" in text or "行业" in text or "ind" in text:
        return "INDUSTRY_UNKNOWN_TAXONOMY"
    return "UNKNOWN"


def inspect_candidate(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    columns: list[str] = []
    notes: list[str] = []
    inspected = False

    try:
        if suffix == ".parquet":
            columns = parquet_columns(path)
            inspected = True
        elif suffix in {".csv", ".txt", ".tsv"}:
            columns = csv_header(path)
            inspected = True
        elif suffix in {".xlsx", ".xlsm"}:
            columns = excel_header_light(path)
            inspected = bool(columns)
            if not inspected:
                notes.append("Excel file not opened because file is large; filename-only candidate.")
        elif suffix == ".xls":
            notes.append("Legacy xls not opened in prep; filename-only candidate.")
        else:
            notes.append("Unsupported format for schema preview; filename-only candidate.")
    except Exception as exc:  # noqa: BLE001
        notes.append(f"Light schema inspection failed: {type(exc).__name__}: {exc}")

    symbol_col = find_col(
        columns,
        [["symbol"], ["stkcd"], ["stockcode"], ["stock_code"], ["证券代码"], ["股票代码"], ["code"]],
    )
    date_col = find_col(
        columns,
        [
            ["month_end"],
            ["monthend"],
            ["month"],
            ["enddate"],
            ["trdmnt"],
            ["date"],
            ["截止日期"],
            ["交易月份"],
        ],
    )
    industry_col = find_col(
        columns,
        [
            ["industry"],
            ["indname"],
            ["indcd"],
            ["sector"],
            ["sw"],
            ["shenwan"],
            ["citic"],
            ["csrc"],
            ["行业"],
            ["申万"],
            ["中信"],
        ],
    )
    taxonomy = guess_taxonomy(path, columns)

    has_symbol = bool(symbol_col)
    has_industry = bool(industry_col)
    has_date = bool(date_col)
    if inspected and has_symbol and has_industry and has_date:
        source_status = "PIT_MONTHLY_AVAILABLE"
        pit_status = "PIT_COMPATIBLE"
        notes.append("Schema contains symbol/date/industry candidates; supports as-of month join subject to next-run QA.")
    elif inspected and has_symbol and has_industry:
        source_status = "STATIC_AVAILABLE_WATCH"
        pit_status = "STATIC_NOT_PIT"
        notes.append("Schema contains symbol/industry but no date candidate; usable only as static WATCH source.")
    elif inspected and columns:
        source_status = "AMBIGUOUS_MANUAL_REVIEW"
        pit_status = "UNKNOWN"
        notes.append("Schema preview does not clearly identify required join fields.")
    else:
        source_status = "AMBIGUOUS_MANUAL_REVIEW"
        pit_status = "UNKNOWN"
        notes.append("Candidate identified by filename only; manual schema review required.")

    return {
        "candidate_source_name": path.name,
        "source_path": rel(path),
        "source_format": suffix.lstrip(".").upper() or "UNKNOWN",
        "file_found": True,
        "symbol_column_candidate": symbol_col,
        "date_column_candidate": date_col,
        "industry_column_candidate": industry_col,
        "industry_taxonomy_guess": taxonomy,
        "industry_source_status": source_status,
        "pit_quality_status": pit_status,
        "notes": " | ".join(notes),
        "_score": (
            (100 if source_status == "PIT_MONTHLY_AVAILABLE" else 0)
            + (50 if source_status == "STATIC_AVAILABLE_WATCH" else 0)
            + (5 if suffix == ".parquet" else 0)
            + (3 if "industry" in path.name.lower() or "行业" in path.name else 0)
        ),
    }


def filename_matches_keyword(filename: str) -> bool:
    lower = filename.lower()
    if any(token in filename for token in ["行业", "申万", "中信"]):
        return True
    long_keywords = ["industry", "shenwan", "citic", "csrc", "sector"]
    if any(keyword in lower for keyword in long_keywords):
        return True
    tokens = [token for token in re.split(r"[^a-z0-9]+", lower) if token]
    return "sw" in tokens or "ind" in tokens or "classification" in tokens


def scan_industry_sources() -> list[dict[str, Any]]:
    seen: set[Path] = set()
    candidates: list[Path] = []
    for base in SCAN_DIRS:
        if not base.exists():
            continue
        for dirpath, dirnames, filenames in os.walk(base):
            current = Path(dirpath)
            parts_lower = {part.lower() for part in current.parts}
            try:
                current.relative_to(OUT_DIR)
                dirnames[:] = []
                continue
            except ValueError:
                pass
            if "_agent_runs" in parts_lower:
                dirnames[:] = []
                continue
            dirnames[:] = [
                d
                for d in dirnames
                if d.lower() not in SKIP_DIR_NAMES
                and d.lower() != "_agent_runs"
                and not d.lower().endswith("cache")
            ]
            for filename in filenames:
                path = current / filename
                if path.suffix.lower() not in DATA_SUFFIXES:
                    continue
                if filename_matches_keyword(filename):
                    if path not in seen:
                        seen.add(path)
                        candidates.append(path)

    rows = [inspect_candidate(path) for path in sorted(candidates)]
    if not rows:
        rows.append(
            {
                "candidate_source_name": "NO_CANDIDATE_FOUND",
                "source_path": "",
                "source_format": "",
                "file_found": False,
                "symbol_column_candidate": "",
                "date_column_candidate": "",
                "industry_column_candidate": "",
                "industry_taxonomy_guess": "",
                "industry_source_status": "MISSING",
                "pit_quality_status": "NOT_USABLE",
                "notes": "No candidate industry source found in allowed scan directories.",
                "_score": -1,
            }
        )
    return rows


def select_source(scan_rows: list[dict[str, Any]]) -> dict[str, Any]:
    usable = [row for row in scan_rows if row["industry_source_status"] != "MISSING"]
    if not usable:
        return scan_rows[0]
    return sorted(usable, key=lambda row: row["_score"], reverse=True)[0]


def score_panel_schema() -> tuple[list[str], bool]:
    cols = parquet_columns(SCORE_PANEL)
    industry_like = [
        col
        for col in cols
        if any(token in col.lower() for token in ["industry", "sector", "shenwan", "citic", "csrc", "sw"])
    ]
    del cols
    gc.collect()
    return industry_like, False


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    run_timestamp = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    print(f"[{run_timestamp}] start {TASK_NAME}")

    prerequisite_rows = []
    for path in REQUIRED_INPUTS:
        prerequisite_rows.append(
            {
                "path": rel(path),
                "exists": path.exists(),
                "bytes": path.stat().st_size if path.exists() else None,
            }
        )
    prerequisites_passed = all(row["exists"] for row in prerequisite_rows)

    summary_inputs: dict[str, Any] = {}
    for path in REQUIRED_INPUTS:
        if path.suffix.lower() == ".json" and path.exists():
            summary_inputs[rel(path) or path.name] = read_json(path)
    candidate_manifest_header = csv_header(
        ROOT / "output" / "simple_baseline_candidate_selection_v0" / "simple_baseline_candidate_manifest.csv"
    )
    score_formula_header = csv_header(
        ROOT / "output" / "simple_baseline_score_prep_v0" / "simple_baseline_score_formula_manifest.csv"
    )

    score_panel_industry_like_columns: list[str] = []
    score_panel_full_read = False
    if SCORE_PANEL.exists():
        score_panel_industry_like_columns, score_panel_full_read = score_panel_schema()

    scan_rows = scan_industry_sources()
    selected = select_source(scan_rows)
    industry_source_status = selected["industry_source_status"]
    pit_quality_status = selected["pit_quality_status"]
    selected_source_path = selected["source_path"] if selected.get("file_found") else None
    industry_taxonomy = selected["industry_taxonomy_guess"] or "UNKNOWN"
    industry_source_found = industry_source_status != "MISSING"

    manual_review_required = industry_source_status == "AMBIGUOUS_MANUAL_REVIEW"
    within_rank_allowed = industry_source_status == "PIT_MONTHLY_AVAILABLE"
    constrained_allowed = industry_source_status in {"PIT_MONTHLY_AVAILABLE", "STATIC_AVAILABLE_WATCH"}
    raw_audit_allowed = industry_source_status in {
        "PIT_MONTHLY_AVAILABLE",
        "STATIC_AVAILABLE_WATCH",
        "AMBIGUOUS_MANUAL_REVIEW",
    }

    if not prerequisites_passed:
        final_decision = "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_FAIL"
        recommended_next_step = "先补齐缺失输入文件，再重新运行行业中性化 prep。"
    elif industry_source_status == "PIT_MONTHLY_AVAILABLE":
        final_decision = "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_READY_FOR_NEUTRAL_SCORE_RUN"
        recommended_next_step = "下一步运行 industry-neutral score run，仅计算中性化 score 和 join QA，不做回测。"
    elif industry_source_status == "STATIC_AVAILABLE_WATCH":
        final_decision = "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_READY_FOR_STATIC_EXPOSURE_AUDIT"
        recommended_next_step = "下一步只做 static 行业暴露审计或行业约束组合 prep，不声称 PIT-clean。"
    elif industry_source_status == "AMBIGUOUS_MANUAL_REVIEW":
        final_decision = "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_WATCH_MANUAL_REVIEW_REQUIRED"
        recommended_next_step = "先人工确认候选行业源字段、分类体系和日期口径，再决定是否进入 neutral score run。"
    else:
        final_decision = "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_FAIL_NEEDS_INDUSTRY_DATA"
        recommended_next_step = "先准备 symbol + month_end + industry 的 PIT 月度行业源。"

    prerequisite_check = {
        "run_timestamp": run_timestamp,
        "task_name": TASK_NAME,
        "prerequisites_passed": prerequisites_passed,
        "required_inputs": prerequisite_rows,
        "candidate_manifest_header": candidate_manifest_header,
        "score_formula_manifest_header": score_formula_header,
        "score_panel_path": rel(SCORE_PANEL),
        "score_panel_parquet_full_read": score_panel_full_read,
        "score_panel_industry_like_columns": score_panel_industry_like_columns,
        "summary_inputs_loaded_count": len(summary_inputs),
    }
    write_json(OUT_DIR / "industry_neutralization_prep_prerequisite_check.json", prerequisite_check)

    scan_fields = [
        "candidate_source_name",
        "source_path",
        "source_format",
        "file_found",
        "symbol_column_candidate",
        "date_column_candidate",
        "industry_column_candidate",
        "industry_taxonomy_guess",
        "industry_source_status",
        "pit_quality_status",
        "notes",
    ]
    write_csv(OUT_DIR / "industry_source_availability_scan.csv", scan_rows, scan_fields)

    selected_policy = {
        "selected_source_path": selected_source_path,
        "industry_source_status": industry_source_status,
        "pit_quality_status": pit_quality_status,
        "symbol_join_key": selected.get("symbol_column_candidate") or None,
        "date_join_key": selected.get("date_column_candidate") or None,
        "industry_column": selected.get("industry_column_candidate") or None,
        "industry_taxonomy": industry_taxonomy,
        "join_policy": (
            "PIT/as-of month join on symbol + month_end; next run must verify one symbol-month one industry."
            if industry_source_status == "PIT_MONTHLY_AVAILABLE"
            else "Static symbol-level join only; not PIT-clean."
            if industry_source_status == "STATIC_AVAILABLE_WATCH"
            else "No executable join policy until manual review/data prep completes."
        ),
        "limitations": [
            "本任务只做 prep，不计算行业暴露或中性化 score。",
            "候选行业源只做轻量 schema/header 检查，未做全量读取。",
            "下一步必须验证 join coverage、重复 symbol-month、行业缺失和分类 stale 风险。",
        ],
        "manual_review_required": manual_review_required,
    }
    write_json(OUT_DIR / "selected_industry_source_policy.json", selected_policy)

    method_rows = [
        {
            "method_name": "INDUSTRY_WITHIN_RANK_SCORE",
            "method_type": "score_neutralization",
            "requires_pit_industry": True,
            "allowed_next_run": within_rank_allowed,
            "blocked_reason": "" if within_rank_allowed else "Requires PIT_MONTHLY_AVAILABLE industry source.",
            "output_score_or_portfolio": ";".join(NEUTRAL_SCORE_COLUMNS),
            "notes": "每月、每行业内对 raw rank feature 再排序后构造 score。",
        },
        {
            "method_name": "INDUSTRY_CONSTRAINED_PORTFOLIO",
            "method_type": "portfolio_constraint",
            "requires_pit_industry": False,
            "allowed_next_run": constrained_allowed,
            "blocked_reason": "" if constrained_allowed else "Requires usable industry source.",
            "output_score_or_portfolio": "industry_constrained_value_quality_portfolio",
            "notes": "保留 raw score，在组合构造阶段约束行业权重接近 universe distribution。",
        },
        {
            "method_name": "RAW_BASELINE_EXPOSURE_AUDIT",
            "method_type": "exposure_audit",
            "requires_pit_industry": False,
            "allowed_next_run": raw_audit_allowed,
            "blocked_reason": "" if raw_audit_allowed else "Requires usable or reviewable industry source.",
            "output_score_or_portfolio": "raw_value_quality_industry_exposure_audit",
            "notes": "只审计 raw baseline 行业暴露，不改变 score 或 portfolio。",
        },
    ]
    write_csv(
        OUT_DIR / "industry_neutralization_method_plan.csv",
        method_rows,
        [
            "method_name",
            "method_type",
            "requires_pit_industry",
            "allowed_next_run",
            "blocked_reason",
            "output_score_or_portfolio",
            "notes",
        ],
    )

    formula_rows = [
        {
            "neutral_score_name": "IND_NEUTRAL_VALUE_BP_SINGLE_score",
            "source_features": "bp_rank",
            "neutralization_method": "industry_within_rank",
            "score_formula": "industry_within_rank(bp_rank)",
            "expected_score_direction": "higher_is_better",
            "allowed_next_run": within_rank_allowed,
            "notes": "仅当 PIT 月度行业源通过下一步 join QA 后执行。",
        },
        {
            "neutral_score_name": "IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score",
            "source_features": ";".join(SOURCE_FEATURES),
            "neutralization_method": "industry_within_rank",
            "score_formula": "mean(industry_within_rank(bp_rank), industry_within_rank(ep_ttm_rank), industry_within_rank(cfo_to_earnings_parent_rank))",
            "expected_score_direction": "higher_is_better",
            "allowed_next_run": within_rank_allowed,
            "notes": "仅当 PIT 月度行业源通过下一步 join QA 后执行。",
        },
    ]
    write_csv(
        OUT_DIR / "industry_neutral_score_formula_plan.csv",
        formula_rows,
        [
            "neutral_score_name",
            "source_features",
            "neutralization_method",
            "score_formula",
            "expected_score_direction",
            "allowed_next_run",
            "notes",
        ],
    )

    qa_rows = [
        {
            "qa_check": "symbol join coverage",
            "required_in_next_run": True,
            "blocking_if_fail": True,
            "notes": "按 score panel symbol 覆盖率审计，低覆盖率阻断后续。",
        },
        {
            "qa_check": "month_end join coverage",
            "required_in_next_run": True,
            "blocking_if_fail": True,
            "notes": "PIT 方案必须确认 month_end/as-of join 覆盖率。",
        },
        {
            "qa_check": "missing industry count",
            "required_in_next_run": True,
            "blocking_if_fail": True,
            "notes": "输出缺失行业 symbol-month 数和比例。",
        },
        {
            "qa_check": "stale industry classification check",
            "required_in_next_run": True,
            "blocking_if_fail": True,
            "notes": "检查行业分类长时间未变化或日期早于样本窗口的 stale 风险。",
        },
        {
            "qa_check": "one symbol-month one industry check",
            "required_in_next_run": True,
            "blocking_if_fail": True,
            "notes": "每个 symbol-month 必须唯一映射一个行业。",
        },
        {
            "qa_check": "industry category count",
            "required_in_next_run": True,
            "blocking_if_fail": False,
            "notes": "记录月度行业类别数量，发现异常波动需 review。",
        },
        {
            "qa_check": "raw vs neutral score row count consistency",
            "required_in_next_run": True,
            "blocking_if_fail": True,
            "notes": "neutral score 输出行数必须与可计算 raw score panel 对齐。",
        },
    ]
    write_csv(
        OUT_DIR / "industry_join_qa_plan.csv",
        qa_rows,
        ["qa_check", "required_in_next_run", "blocking_if_fail", "notes"],
    )

    run_config = {
        "score_panel_path": rel(SCORE_PANEL),
        "selected_industry_source_path": selected_source_path,
        "industry_source_status": industry_source_status,
        "pit_quality_status": pit_quality_status,
        "selected_methods": [row["method_name"] for row in method_rows if row["allowed_next_run"]],
        "source_features": SOURCE_FEATURES,
        "neutral_score_columns": NEUTRAL_SCORE_COLUMNS,
        "raw_score_columns_for_comparison": RAW_SCORE_COLUMNS,
        "output_directory_for_next_run": "output\\simple_baseline_industry_neutral_score_run_v0\\",
        "calculate_scores_next_run_allowed": within_rank_allowed,
        "calculate_portfolio_returns_next_run_allowed": False,
        "backtest_allowed_next_run": False,
        "production_allowed_next_run": False,
        "calculated_now": False,
    }
    write_json(OUT_DIR / "industry_neutralization_run_config_draft.json", run_config)

    guardrail_rows = [
        {
            "guardrail": "prep_only_no_exposure_calculation",
            "passed": True,
            "notes": "未计算行业暴露。",
        },
        {
            "guardrail": "prep_only_no_neutral_score_calculation",
            "passed": True,
            "notes": "未生成行业中性 score。",
        },
        {
            "guardrail": "prep_only_no_portfolio_return_or_backtest",
            "passed": True,
            "notes": "未计算组合收益、回测、交易成本、换手、Sharpe、MaxDD。",
        },
        {
            "guardrail": "score_panel_not_full_read",
            "passed": not score_panel_full_read,
            "notes": "仅读取 parquet schema。",
        },
        {
            "guardrail": "industry_source_not_full_read",
            "passed": True,
            "notes": "候选源只做文件名和轻量 schema/header 检查。",
        },
        {
            "guardrail": "production_not_modified",
            "passed": True,
            "notes": "未写 production。",
        },
    ]
    write_csv(OUT_DIR / "industry_neutralization_guardrail_checklist.csv", guardrail_rows, ["guardrail", "passed", "notes"])

    summary = {
        "run_timestamp": run_timestamp,
        "prerequisites_passed": prerequisites_passed,
        "industry_source_found": industry_source_found,
        "industry_source_status": industry_source_status,
        "pit_quality_status": pit_quality_status,
        "selected_industry_source_path": selected_source_path,
        "industry_taxonomy": industry_taxonomy,
        "manual_review_required": manual_review_required,
        "industry_within_rank_score_allowed": within_rank_allowed,
        "industry_constrained_portfolio_allowed": constrained_allowed,
        "raw_exposure_audit_allowed": raw_audit_allowed,
        "neutral_score_count_planned": len(NEUTRAL_SCORE_COLUMNS),
        "neutral_score_columns_planned": NEUTRAL_SCORE_COLUMNS,
        "source_features": SOURCE_FEATURES,
        "score_panel_path_recorded": rel(SCORE_PANEL),
        "score_panel_parquet_full_read": score_panel_full_read,
        "industry_source_full_read": False,
        "industry_exposure_calculated": False,
        "neutral_score_calculated": False,
        "portfolio_return_calculated": False,
        "backtest_run": False,
        "transaction_cost_calculated": False,
        "turnover_calculated": False,
        "sharpe_calculated": False,
        "maxdd_calculated": False,
        "benchmark_relative_return_calculated": False,
        "alpha_beta_regression_calculated": False,
        "training_run": False,
        "shap_calculated": False,
        "tuning_run": False,
        "feature_importance_calculated": False,
        "production_holdings_generated": False,
        "live_order_ready_file_generated": False,
        "production_modified": False,
        "compact_f_rescue_blocked": True,
        "sign_flip_production_blocked": True,
        "lightgbm_first_blocked": True,
        "final_decision": final_decision,
        "recommended_next_step": recommended_next_step,
    }
    write_json(OUT_DIR / "simple_baseline_industry_neutralization_prep_summary.json", summary)

    next_step_md = f"""# 下一步 Industry-Neutral Run Plan

## 当前决策

- final_decision: `{final_decision}`
- selected_source_path: `{selected_source_path}`
- industry_source_status: `{industry_source_status}`
- pit_quality_status: `{pit_quality_status}`

## 允许的下一步

- calculate_scores_next_run_allowed: `{within_rank_allowed}`
- calculate_portfolio_returns_next_run_allowed: `False`
- backtest_allowed_next_run: `False`
- production_allowed_next_run: `False`

## 下一步必须先做的 QA

1. symbol join coverage
2. month_end join coverage
3. missing industry count
4. stale industry classification check
5. one symbol-month one industry check
6. industry category count
7. raw vs neutral score row count consistency

## 备注

本任务未计算行业暴露、未生成中性化 score、未回测、未写 production。
"""
    (OUT_DIR / "next_step_industry_neutralization_plan.md").write_text(next_step_md, encoding="utf-8")

    report_md = f"""# Simple Baseline Industry Neutralization Prep Report

## 结论

`{final_decision}`

## 行业源

- selected_source_path: `{selected_source_path}`
- industry_source_status: `{industry_source_status}`
- pit_quality_status: `{pit_quality_status}`
- industry_taxonomy: `{industry_taxonomy}`
- manual_review_required: `{manual_review_required}`

## 方法冻结

- INDUSTRY_WITHIN_RANK_SCORE allowed: `{within_rank_allowed}`
- INDUSTRY_CONSTRAINED_PORTFOLIO allowed: `{constrained_allowed}`
- RAW_BASELINE_EXPOSURE_AUDIT allowed: `{raw_audit_allowed}`

## 明确未执行

未计算行业暴露、行业中性 score、组合收益、回测、交易成本、换手、Sharpe、MaxDD、benchmark-relative return、alpha/beta、训练、调参、SHAP 或 production holdings。
"""
    (OUT_DIR / "simple_baseline_industry_neutralization_prep_report.md").write_text(report_md, encoding="utf-8")

    completion_card = f"""# Task Completion Card

- task_name: `{TASK_NAME}`
- status: completed
- final_decision: `{final_decision}`
- output_directory: `{rel(OUT_DIR)}`
- run_log_directory: `{rel(RUN_DIR)}`
- calculated_now: `False`
- production_modified: `False`
"""
    (OUT_DIR / "task_completion_card.md").write_text(completion_card, encoding="utf-8")

    final_qa_rows = [
        {"check": "required_inputs_exist", "passed": prerequisites_passed, "notes": "全部必需输入文件存在。" if prerequisites_passed else "存在缺失输入。"},
        {"check": "score_panel_full_read", "passed": not score_panel_full_read, "notes": "仅读取 parquet schema。"},
        {"check": "industry_source_full_read", "passed": True, "notes": "未全量读取行业源。"},
        {"check": "no_forbidden_calculation", "passed": True, "notes": "未计算暴露、score、组合收益或回测。"},
        {"check": "production_not_modified", "passed": True, "notes": "未写 production。"},
        {"check": "final_decision_valid", "passed": final_decision in {
            "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_READY_FOR_NEUTRAL_SCORE_RUN",
            "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_READY_FOR_STATIC_EXPOSURE_AUDIT",
            "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_WATCH_MANUAL_REVIEW_REQUIRED",
            "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_FAIL_NEEDS_INDUSTRY_DATA",
            "SIMPLE_BASELINE_INDUSTRY_NEUTRAL_PREP_FAIL",
        }, "notes": "final_decision 使用允许枚举值。"},
    ]
    write_csv(OUT_DIR / "final_qa.csv", final_qa_rows, ["check", "passed", "notes"])

    terminal_summary = {
        "task_name": TASK_NAME,
        "run_timestamp": run_timestamp,
        "script_path": rel(ROOT / "scripts" / "prep_simple_baseline_industry_neutralization_v0.py"),
        "stdout_log": rel(RUN_DIR / "run_stdout.txt"),
        "stderr_log": rel(RUN_DIR / "run_stderr.txt"),
        "exit_code": 0,
        "output_directory": rel(OUT_DIR),
        "final_decision": final_decision,
    }
    write_json(OUT_DIR / "terminal_summary.json", terminal_summary)
    write_json(RUN_DIR / "terminal_summary.json", terminal_summary)

    run_state = f"""# RUN_STATE

任务：{TASK_NAME}
状态：完成

已完成：
- 必需输入存在性检查
- score panel parquet schema 检查，未全量读取
- 允许目录内候选行业源文件名扫描和轻量 schema/header 检查
- 行业源 policy、method plan、formula plan、QA plan、run config draft、summary/report 生成

final_decision: {final_decision}
selected_industry_source_path: {selected_source_path}

禁止项确认：
- 未计算行业暴露
- 未生成行业中性 score
- 未计算 portfolio return / backtest / transaction cost / turnover / Sharpe / MaxDD
- 未训练 / 调参 / SHAP
- 未写 production
"""
    (RUN_DIR / "RUN_STATE.md").write_text(run_state, encoding="utf-8")

    del summary_inputs, scan_rows, prerequisite_rows, method_rows, formula_rows, qa_rows, guardrail_rows
    gc.collect()
    print(f"[{datetime.now(timezone.utc).astimezone().isoformat(timespec='seconds')}] completed {TASK_NAME}")
    print(f"final_decision={final_decision}")
    print(f"output_directory={rel(OUT_DIR)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
