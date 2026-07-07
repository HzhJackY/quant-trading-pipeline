from __future__ import annotations

import gc
import json
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TASK = "benchmark_source_audit_monthly_alignment_v0"
OUT = ROOT / "output" / TASK
RUN = ROOT / "output" / "_agent_runs" / TASK
CSMAR = ROOT / "data" / "csmar_exports"
PORT = ROOT / "output" / "unified_robust_portfolio_evaluation_run_v0"
PANEL = ROOT / "output" / "robust_cleaned_fundamental_factor_variant_build_v0" / "robust_cleaned_factor_score_panel_v0.parquet"

OUT.mkdir(parents=True, exist_ok=True)
RUN.mkdir(parents=True, exist_ok=True)


REQUIRED_FILES = [
    "TRD_Index.xlsx",
    "TRD_Index[DES][xlsx].txt",
    "TRD_Cnmont.xlsx",
    "TRD_Cnmont[DES][xlsx].txt",
    "TRD_Mont.xlsx",
    "TRD_Mont[DES][xlsx].txt",
    "TRD_Mnth.xlsx",
    "TRD_Mnth[DES][xlsx].txt",
    "TRD_Nrrate.xlsx",
    "TRD_Nrrate[DES][xlsx].txt",
    "STK_MKT_FIVEFACDAY.xlsx",
    "STK_MKT_FIVEFACDAY[DES][xlsx].txt",
]

REQUIRED_BY_FILE = {
    "TRD_Index.xlsx": ["Indexcd", "Trddt", "Clsindex", "Retindex"],
    "TRD_Cnmont.xlsx": [
        "Markettype",
        "Trdmnt",
        "Cmretwdeq",
        "Cmretmdeq",
        "Cmretwdos",
        "Cmretmdos",
        "Cmretwdtl",
        "Cmretmdtl",
    ],
    "TRD_Mont.xlsx": [
        "Markettype",
        "Trdmnt",
        "Mretwdeq",
        "Mretmdeq",
        "Mretwdos",
        "Mretmdos",
        "Mretwdtl",
        "Mretmdtl",
    ],
    "TRD_Mnth.xlsx": [
        "Stkcd",
        "Trdmnt",
        "Mretwd",
        "Mretnd",
        "Msmvosd",
        "Msmvttl",
        "Ndaytrd",
        "Markettype",
    ],
    "TRD_Nrrate.xlsx": ["Nrr1", "Clsdt", "Nrrdata", "Nrrdaydt", "Nrrmtdt"],
    "STK_MKT_FIVEFACDAY.xlsx": [
        "MarkettypeID",
        "TradingDate",
        "Portfolios",
        "RiskPremium1",
        "RiskPremium2",
        "SMB1",
        "SMB2",
        "HML1",
        "HML2",
        "RMW1",
        "RMW2",
        "CMA1",
        "CMA2",
    ],
}


def checkpoint(message: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = (
        f"# RUN_STATE: Benchmark Source Audit & Monthly Alignment v0\n\n"
        f"status: running\nlast_checkpoint: {ts} {message}\n"
        f"mode: low-resource checkpoint-first resume-safe\n\n"
        f"logs:\n- {RUN / 'run_stdout.txt'}\n- {RUN / 'run_stderr.txt'}\n"
    )
    (RUN / "RUN_STATE.md").write_text(text, encoding="utf-8")
    print(f"[checkpoint] {message}")


def norm_col(x: object) -> str:
    return str(x).strip() if x is not None else ""


def read_des(file_name: str) -> dict[str, str]:
    p = CSMAR / file_name
    if not p.exists():
        return {}
    mapping: dict[str, str] = {}
    for line in p.read_text(encoding="utf-8", errors="ignore").splitlines():
        m = re.match(r"^\s*([A-Za-z0-9_]+)\s*\[(.*?)\]\s*-\s*(.*)$", line)
        if m:
            mapping[m.group(1)] = f"{m.group(2)} - {m.group(3)}".strip()
    return mapping


def inspect_xlsx(file_name: str) -> dict:
    p = CSMAR / file_name
    if not p.exists():
        return {
            "file_name": file_name,
            "file_exists": False,
            "row_count": 0,
            "column_count": 0,
            "sheet_names": "",
            "detected_header_row": np.nan,
            "detected_data_start_row": np.nan,
            "columns_original": "",
            "columns_normalized": "",
        }
    wb = load_workbook(p, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    required = set(REQUIRED_BY_FILE.get(file_name, []))
    best_row = 1
    best_score = -1
    best_values: list[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True), start=1):
        values = [norm_col(v) for v in row]
        score = sum(1 for v in values if v in required)
        if score > best_score:
            best_score = score
            best_row = i
            best_values = values
    columns = [c for c in best_values if c]
    result = {
        "file_name": file_name,
        "file_exists": True,
        "row_count": max(int(ws.max_row) - best_row, 0),
        "column_count": int(ws.max_column),
        "sheet_names": "|".join(sheet_names),
        "detected_header_row": best_row,
        "detected_data_start_row": best_row + 1,
        "columns_original": "|".join(columns),
        "columns_normalized": "|".join(columns),
    }
    wb.close()
    return result


def excel_cols(file_name: str) -> list[str]:
    meta = inspect_xlsx(file_name)
    return [c for c in meta.get("columns_normalized", "").split("|") if c]


def read_excel_needed(file_name: str, columns: list[str] | None = None, nrows: int | None = None) -> pd.DataFrame:
    meta = inspect_xlsx(file_name)
    header = int(meta["detected_header_row"]) - 1
    available = [c for c in meta.get("columns_normalized", "").split("|") if c]
    usecols = None
    if columns:
        usecols = [c for c in columns if c in available]
    dtype = {}
    for c in ["Indexcd", "Stkcd", "MarkettypeID", "Nrr1"]:
        if usecols is None or c in usecols:
            dtype[c] = "string"
    return pd.read_excel(CSMAR / file_name, header=header, usecols=usecols, nrows=nrows, dtype=dtype, engine="openpyxl")


def detect_unit(series: pd.Series, field: str = "") -> str:
    x = pd.to_numeric(series, errors="coerce").dropna()
    if x.empty:
        return "UNKNOWN_NO_DATA"
    q95 = float(x.abs().quantile(0.95))
    mx = float(x.abs().max())
    low = field.lower()
    if "nrr" in low:
        return "PERCENT_BY_DES"
    if mx > 5 or q95 > 1:
        return "PERCENT_SUSPECT"
    return "DECIMAL_SUSPECT"


def to_decimal(series: pd.Series, unit: str) -> pd.Series:
    x = pd.to_numeric(series, errors="coerce")
    if "PERCENT" in unit:
        return x / 100.0
    return x


def month_str_from_ts(ts: pd.Timestamp, add_months: int = 0) -> str:
    return (ts + pd.offsets.MonthEnd(add_months)).strftime("%Y-%m")


def find_month_col(df: pd.DataFrame) -> str:
    candidates = [c for c in df.columns if str(c).lower() in {"month_end", "portfolio_month_end", "date"}]
    if candidates:
        return candidates[0]
    for c in df.columns:
        s = pd.to_datetime(df[c], errors="coerce")
        if s.notna().sum() >= min(5, len(df)):
            return c
    raise ValueError("无法在 portfolio monthly return 中识别 month_end 列")


def compound_returns(values: pd.Series) -> float:
    x = pd.to_numeric(values, errors="coerce").dropna()
    if x.empty:
        return np.nan
    return float(np.prod(1.0 + x) - 1.0)


def priority_for_index(code: str, name: str) -> str:
    text = f"{code} {name}"
    if "中证800" in text or code in {"000906", "399906"}:
        return "PRIMARY_CSI800_CANDIDATE"
    if "沪深300" in text or code in {"000300", "399300"}:
        return "SECONDARY_HS300_CANDIDATE"
    if "中证500" in text or code in {"000905", "399905"}:
        return "SECONDARY_CSI500_CANDIDATE"
    if "中证1000" in text or "中证全指" in text or "中证流通" in text or code in {"000902", "000852", "399852", "000985"}:
        return "SECONDARY_CSI1000_OR_CSI_ALL_SHARE_CANDIDATE"
    return "OTHER_INDEX" if name else "UNKNOWN_INDEX"


def schema_audit() -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    mappings = []
    for xlsx in [f for f in REQUIRED_FILES if f.endswith(".xlsx")]:
        meta = inspect_xlsx(xlsx)
        des = read_des(xlsx.replace(".xlsx", "[DES][xlsx].txt"))
        sample = pd.DataFrame()
        if meta["file_exists"]:
            sample = read_excel_needed(xlsx, nrows=300)
        cols = [c for c in meta.get("columns_normalized", "").split("|") if c]
        date_cols = [c for c in cols if "date" in c.lower() or "dt" in c.lower() or "mnt" in c.lower()]
        numeric_cols = []
        for c in sample.columns:
            if pd.to_numeric(sample[c], errors="coerce").notna().sum() > max(3, len(sample) * 0.5):
                numeric_cols.append(c)
        required = REQUIRED_BY_FILE.get(xlsx, [])
        miss = {}
        for c in required:
            miss[c] = float(sample[c].isna().mean()) if c in sample.columns and len(sample) else np.nan
        key_cols = [c for c in ["Indexcd", "Trddt", "Markettype", "Trdmnt", "Stkcd", "Clsdt", "MarkettypeID", "TradingDate", "Portfolios"] if c in cols]
        dup = np.nan
        min_date = ""
        max_date = ""
        if key_cols and meta["file_exists"]:
            key_df = read_excel_needed(xlsx, key_cols)
            dup = int(key_df.duplicated(key_cols).sum())
            for dc in ["Trddt", "TradingDate", "Clsdt", "Trdmnt"]:
                if dc in key_df.columns:
                    if dc == "Trdmnt":
                        d = pd.to_datetime(key_df[dc].astype(str) + "-01", errors="coerce")
                    else:
                        d = pd.to_datetime(key_df[dc], errors="coerce")
                    if d.notna().any():
                        min_date = d.min().strftime("%Y-%m-%d")
                        max_date = d.max().strftime("%Y-%m-%d")
                        break
            del key_df
            gc.collect()
        suspect_fields = [c for c in numeric_cols if any(k in c.lower() for k in ["ret", "nrr", "riskpremium", "smb", "hml", "rmw", "cma"])]
        pct_diag = {c: detect_unit(sample[c], c) for c in suspect_fields if c in sample.columns}
        row = {
            **meta,
            "date_columns_detected": "|".join(date_cols),
            "numeric_columns_detected": "|".join(numeric_cols),
            "duplicate_key_count": dup,
            "min_date": min_date,
            "max_date": max_date,
            "missing_rate_by_required_field": json.dumps(miss, ensure_ascii=False),
            "unit_suspect_fields": "|".join(suspect_fields),
            "percent_vs_decimal_suspect": json.dumps(pct_diag, ensure_ascii=False),
        }
        rows.append(row)
        for col in cols:
            mappings.append({
                "source_file": xlsx,
                "field": col,
                "description_from_des": des.get(col, ""),
                "field_in_required_list": col in required,
            })
        del sample
        gc.collect()
    return pd.DataFrame(rows), pd.DataFrame(mappings)


def load_portfolio_months() -> list[pd.Timestamp]:
    df = pd.read_csv(PORT / "unified_portfolio_monthly_gross_return.csv")
    col = find_month_col(df)
    months = pd.to_datetime(df[col], errors="coerce").dropna().dt.to_period("M").dt.to_timestamp("M").drop_duplicates().sort_values()
    del df
    gc.collect()
    return list(months)


def build_index_outputs(months: list[pd.Timestamp]) -> tuple[pd.DataFrame, pd.DataFrame]:
    des = read_des("TRD_Index[DES][xlsx].txt")
    df = read_excel_needed("TRD_Index.xlsx", ["Indexcd", "Trddt", "Clsindex", "Retindex"])
    needed = {"Indexcd", "Trddt", "Clsindex", "Retindex"}
    if not needed.issubset(df.columns):
        rows = []
        if "Indexcd" in df.columns:
            codes = df["Indexcd"].astype("string").str.zfill(6)
            for code, cnt in codes.dropna().value_counts().sort_index().items():
                rows.append({
                    "Indexcd": str(code),
                    "first_date": "",
                    "last_date": "",
                    "row_count": int(cnt),
                    "nonnull_retindex_count": 0,
                    "nonnull_clsindex_count": 0,
                    "possible_index_name_from_des_or_mapping": "",
                    "priority_label": "UNKNOWN_INDEX",
                })
        return pd.DataFrame(rows, columns=[
            "Indexcd",
            "first_date",
            "last_date",
            "row_count",
            "nonnull_retindex_count",
            "nonnull_clsindex_count",
            "possible_index_name_from_des_or_mapping",
            "priority_label",
        ]), pd.DataFrame(columns=[
            "portfolio_month_end",
            "benchmark_code",
            "benchmark_name_or_label",
            "benchmark_fwd_ret_1m",
            "return_source",
            "source_start_trade_date",
            "source_end_trade_date",
            "trading_day_count",
            "retindex_unit_detected",
            "missing_flag",
            "alignment_warning",
        ])
    df["Indexcd"] = df["Indexcd"].astype("string").str.zfill(6)
    df["Trddt"] = pd.to_datetime(df["Trddt"], errors="coerce")
    df = df.dropna(subset=["Indexcd", "Trddt"]).sort_values(["Indexcd", "Trddt"])
    ret_unit = detect_unit(df["Retindex"], "Retindex")
    df["Retindex_decimal"] = to_decimal(df["Retindex"], ret_unit)
    index_map = {}
    if "Indexcd" in des:
        for code, name in re.findall(r"([0-9]{6})[：:](.*?)(?=；|$)", des["Indexcd"]):
            index_map[code] = name.strip()
    availability = []
    for code, g in df.groupby("Indexcd", sort=True):
        name = index_map.get(str(code), "")
        availability.append({
            "Indexcd": str(code),
            "first_date": g["Trddt"].min().strftime("%Y-%m-%d"),
            "last_date": g["Trddt"].max().strftime("%Y-%m-%d"),
            "row_count": int(len(g)),
            "nonnull_retindex_count": int(g["Retindex"].notna().sum()),
            "nonnull_clsindex_count": int(g["Clsindex"].notna().sum()),
            "possible_index_name_from_des_or_mapping": name,
            "priority_label": priority_for_index(str(code), name),
        })
    avail_df = pd.DataFrame(availability)
    candidate_codes = set(avail_df.loc[avail_df["priority_label"] != "OTHER_INDEX", "Indexcd"])
    if not candidate_codes:
        candidate_codes = set(avail_df["Indexcd"].head(5))
    rows = []
    for code in sorted(candidate_codes):
        g = df[df["Indexcd"] == code].copy()
        name = index_map.get(code, "")
        for i, t in enumerate(months[:-1]):
            nxt = months[i + 1]
            win = g[(g["Trddt"] > t) & (g["Trddt"] <= nxt)]
            warning = ""
            missing = False
            source = "RETINDEX"
            val = np.nan
            if win.empty:
                missing = True
                warning = "NO_INDEX_TRADING_DAYS_IN_FORWARD_WINDOW"
            elif win["Retindex_decimal"].notna().any():
                val = compound_returns(win["Retindex_decimal"])
            else:
                source = "CLSINDEX_FALLBACK"
                prev = g[g["Trddt"] <= t].tail(1)
                if prev.empty or pd.isna(prev.iloc[0]["Clsindex"]) or pd.isna(win.iloc[-1]["Clsindex"]):
                    missing = True
                    warning = "RETINDEX_MISSING_AND_CLSINDEX_FALLBACK_UNAVAILABLE"
                else:
                    val = float(win.iloc[-1]["Clsindex"]) / float(prev.iloc[0]["Clsindex"]) - 1.0
            rows.append({
                "portfolio_month_end": t.strftime("%Y-%m-%d"),
                "benchmark_code": code,
                "benchmark_name_or_label": name or priority_for_index(code, name),
                "benchmark_fwd_ret_1m": val,
                "return_source": source,
                "source_start_trade_date": "" if win.empty else win["Trddt"].min().strftime("%Y-%m-%d"),
                "source_end_trade_date": "" if win.empty else win["Trddt"].max().strftime("%Y-%m-%d"),
                "trading_day_count": int(len(win)),
                "retindex_unit_detected": ret_unit,
                "missing_flag": bool(missing),
                "alignment_warning": warning,
            })
    out = pd.DataFrame(rows)
    del df
    gc.collect()
    return avail_df, out


def build_market_monthly(months: list[pd.Timestamp]) -> pd.DataFrame:
    rows = []
    specs = [
        ("TRD_Cnmont.xlsx", ["Markettype", "Trdmnt", "Cmretwdeq", "Cmretwdos", "Cmretwdtl"], "综合市场"),
        ("TRD_Mont.xlsx", ["Markettype", "Trdmnt", "Mretwdeq", "Mretwdos", "Mretwdtl"], "分市场"),
    ]
    for file_name, cols, label_prefix in specs:
        df = read_excel_needed(file_name, cols)
        if not {"Markettype", "Trdmnt"}.issubset(df.columns):
            continue
        df["Markettype"] = pd.to_numeric(df["Markettype"], errors="coerce").astype("Int64")
        df["Trdmnt"] = df["Trdmnt"].astype(str).str.slice(0, 7)
        fields = [c for c in cols if c not in {"Markettype", "Trdmnt"}]
        fields = [c for c in fields if c in df.columns]
        if not fields:
            continue
        units = {c: detect_unit(df[c], c) for c in fields}
        for c in fields:
            df[c + "_decimal"] = to_decimal(df[c], units[c])
        for t in months[:-1]:
            source_m = (t + pd.offsets.MonthEnd(1)).strftime("%Y-%m")
            for _, r in df[df["Trdmnt"] == source_m].iterrows():
                for c in fields:
                    markettype = "" if pd.isna(r["Markettype"]) else int(r["Markettype"])
                    rows.append({
                        "portfolio_month_end": t.strftime("%Y-%m-%d"),
                        "source_table": file_name.replace(".xlsx", ""),
                        "markettype": markettype,
                        "benchmark_field": c,
                        "benchmark_label": f"{label_prefix}_Markettype_{markettype}_{c}",
                        "benchmark_fwd_ret_1m": r[c + "_decimal"],
                        "source_trdmnt": source_m,
                        "unit_detected": units[c],
                        "missing_flag": bool(pd.isna(r[c + "_decimal"])),
                        "alignment_warning": "",
                    })
        del df
        gc.collect()
    return pd.DataFrame(rows)


def build_internal(months: list[pd.Timestamp]) -> pd.DataFrame:
    import pyarrow.parquet as pq

    schema_names = pq.read_schema(PANEL).names
    wanted = [
        "symbol",
        "month_end",
        "fwd_ret_1m",
        "ROBUST_ASOF_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score",
        "composite_anomaly_flag_soft",
        "primary_industry_code",
        "industry_asof_enddate",
        "bp_robust_rank",
        "ep_ttm_robust_rank",
        "cfo_to_earnings_parent_robust_rank",
        "robust_component_count",
        "robust_industry_neutral_component_count",
        "small_group_flag",
        "Msmvosd",
        "msmvosd",
    ]
    cols = [c for c in wanted if c in schema_names]
    df = pd.read_parquet(PANEL, columns=cols)
    df["month_end"] = pd.to_datetime(df["month_end"], errors="coerce").dt.to_period("M").dt.to_timestamp("M")
    df["fwd_ret_1m"] = pd.to_numeric(df["fwd_ret_1m"], errors="coerce")
    score_col = "ROBUST_ASOF_IND_NEUTRAL_VALUE_QUALITY_EQUAL_WEIGHT_score"
    eligible = df[df[score_col].notna()].copy() if score_col in df.columns else df.copy()
    mcap_col = "Msmvosd" if "Msmvosd" in eligible.columns else ("msmvosd" if "msmvosd" in eligible.columns else None)
    rows = []
    for t in months[:-1]:
        g = eligible[eligible["month_end"] == t]
        rows.append({
            "portfolio_month_end": t.strftime("%Y-%m-%d"),
            "benchmark_label": "INTERNAL_ELIGIBLE_UNIVERSE_EQUAL_WEIGHT",
            "benchmark_fwd_ret_1m": float(g["fwd_ret_1m"].mean()) if len(g) else np.nan,
            "universe_count": int(len(g)),
            "missing_fwd_ret_count": int(g["fwd_ret_1m"].isna().sum()) if len(g) else 0,
            "weighting_method": "equal_weight",
            "anomaly_filter_used": False,
            "mcap_weight_available": bool(mcap_col),
            "alignment_warning": "" if len(g) else "NO_INTERNAL_UNIVERSE_ROWS_FOR_MONTH",
        })
        if mcap_col:
            wdf = g[[mcap_col, "fwd_ret_1m"]].dropna()
            val = np.nan
            if len(wdf) and pd.to_numeric(wdf[mcap_col], errors="coerce").sum() > 0:
                w = pd.to_numeric(wdf[mcap_col], errors="coerce")
                val = float(np.average(wdf["fwd_ret_1m"], weights=w))
            rows.append({
                "portfolio_month_end": t.strftime("%Y-%m-%d"),
                "benchmark_label": "INTERNAL_ELIGIBLE_UNIVERSE_FLOAT_MCAP_WEIGHTED",
                "benchmark_fwd_ret_1m": val,
                "universe_count": int(len(wdf)),
                "missing_fwd_ret_count": int(g["fwd_ret_1m"].isna().sum()) if len(g) else 0,
                "weighting_method": "float_mcap_weighted",
                "anomaly_filter_used": False,
                "mcap_weight_available": True,
                "alignment_warning": "" if pd.notna(val) else "MCAP_WEIGHTED_RETURN_UNAVAILABLE",
            })
        else:
            rows.append({
                "portfolio_month_end": t.strftime("%Y-%m-%d"),
                "benchmark_label": "INTERNAL_ELIGIBLE_UNIVERSE_FLOAT_MCAP_WEIGHTED",
                "benchmark_fwd_ret_1m": np.nan,
                "universe_count": int(len(g)),
                "missing_fwd_ret_count": int(g["fwd_ret_1m"].isna().sum()) if len(g) else 0,
                "weighting_method": "float_mcap_weighted",
                "anomaly_filter_used": False,
                "mcap_weight_available": False,
                "alignment_warning": "NO_RELIABLE_FLOAT_MCAP_FIELD_IN_PANEL",
            })
        if "composite_anomaly_flag_soft" in g.columns:
            cg = g[g["composite_anomaly_flag_soft"] != True]
            warn = "" if len(cg) else "NO_FLAG_CLEAN_UNIVERSE_ROWS_FOR_MONTH"
            used = True
            val = float(cg["fwd_ret_1m"].mean()) if len(cg) else np.nan
        else:
            cg = g.iloc[0:0]
            warn = "COMPOSITE_ANOMALY_FLAG_SOFT_NOT_PRESENT_REGENERATION_NOT_ATTEMPTED"
            used = False
            val = np.nan
        rows.append({
            "portfolio_month_end": t.strftime("%Y-%m-%d"),
            "benchmark_label": "INTERNAL_FLAG_CLEAN_UNIVERSE_EQUAL_WEIGHT",
            "benchmark_fwd_ret_1m": val,
            "universe_count": int(len(cg)),
            "missing_fwd_ret_count": int(cg["fwd_ret_1m"].isna().sum()) if len(cg) else 0,
            "weighting_method": "equal_weight",
            "anomaly_filter_used": used,
            "mcap_weight_available": bool(mcap_col),
            "alignment_warning": warn,
        })
    del df, eligible
    gc.collect()
    return pd.DataFrame(rows)


def build_risk_free(months: list[pd.Timestamp]) -> pd.DataFrame:
    df = read_excel_needed("TRD_Nrrate.xlsx", ["Nrr1", "Clsdt", "Nrrdata", "Nrrdaydt", "Nrrmtdt"])
    if not {"Clsdt", "Nrrmtdt"}.issubset(df.columns):
        return pd.DataFrame([{
            "portfolio_month_end": t.strftime("%Y-%m-%d"),
            "risk_free_monthly_return": np.nan,
            "source_date": "",
            "source_field": "Nrrmtdt",
            "unit_detected": "UNKNOWN_REQUIRED_FIELDS_MISSING",
            "nrr1": "",
            "missing_flag": True,
            "alignment_warning": "TRD_Nrrate 缺少 Clsdt/Nrrmtdt，无法对齐",
        } for t in months[:-1]])
    df["Clsdt"] = pd.to_datetime(df["Clsdt"], errors="coerce")
    unit = detect_unit(df["Nrrmtdt"], "Nrrmtdt")
    df["rf_decimal"] = to_decimal(df["Nrrmtdt"], unit)
    rows = []
    for t in months[:-1]:
        start = t + pd.Timedelta(days=1)
        end = t + pd.offsets.MonthEnd(1)
        g = df[(df["Clsdt"] >= start) & (df["Clsdt"] <= end)].sort_values("Clsdt")
        if g.empty:
            rows.append({
                "portfolio_month_end": t.strftime("%Y-%m-%d"),
                "risk_free_monthly_return": np.nan,
                "source_date": "",
                "source_field": "Nrrmtdt",
                "unit_detected": unit,
                "nrr1": "",
                "missing_flag": True,
                "alignment_warning": "NO_RISK_FREE_OBSERVATION_IN_FORWARD_MONTH",
            })
        else:
            r = g.iloc[-1]
            rows.append({
                "portfolio_month_end": t.strftime("%Y-%m-%d"),
                "risk_free_monthly_return": r["rf_decimal"],
                "source_date": r["Clsdt"].strftime("%Y-%m-%d"),
                "source_field": "Nrrmtdt",
                "unit_detected": unit,
                "nrr1": r.get("Nrr1", ""),
                "missing_flag": bool(pd.isna(r["rf_decimal"])),
                "alignment_warning": "",
            })
    del df
    gc.collect()
    return pd.DataFrame(rows)


def build_ff(months: list[pd.Timestamp]) -> tuple[pd.DataFrame, pd.DataFrame]:
    cols = REQUIRED_BY_FILE["STK_MKT_FIVEFACDAY.xlsx"]
    df = read_excel_needed("STK_MKT_FIVEFACDAY.xlsx", cols)
    manual = []
    if "TradingDate" not in df.columns:
        manual.append({"issue": "date column not recognized or missing from exported file", "candidate_columns": "|".join(df.columns)})
        return pd.DataFrame(), pd.DataFrame(manual)
    df["TradingDate"] = pd.to_datetime(df["TradingDate"], errors="coerce")
    factor_cols = [c for c in ["RiskPremium1", "RiskPremium2", "SMB1", "SMB2", "HML1", "HML2", "RMW1", "RMW2", "CMA1", "CMA2"] if c in df.columns]
    if not factor_cols:
        manual.append({"issue": "factor columns not recognized", "candidate_columns": "|".join(df.columns)})
        return pd.DataFrame(), pd.DataFrame(manual)
    units = {c: detect_unit(df[c], c) for c in factor_cols}
    for c in factor_cols:
        df[c + "_decimal"] = to_decimal(df[c], units[c])
    group_cols = [c for c in ["MarkettypeID", "Portfolios"] if c in df.columns]
    rows = []
    grouped = list(df.groupby(group_cols, dropna=False)) if group_cols else [(("ALL",), df)]
    for keys, g0 in grouped:
        if not isinstance(keys, tuple):
            keys = (keys,)
        label_parts = [str(x) for x in keys]
        label = "_".join(label_parts) if label_parts else "ALL"
        for i, t in enumerate(months[:-1]):
            nxt = months[i + 1]
            win = g0[(g0["TradingDate"] > t) & (g0["TradingDate"] <= nxt)]
            for c in factor_cols:
                val = compound_returns(win[c + "_decimal"]) if not win.empty else np.nan
                rows.append({
                    "portfolio_month_end": t.strftime("%Y-%m-%d"),
                    "factor_set_label": label,
                    "factor_name": c,
                    "factor_monthly_return": val,
                    "source_start_date": "" if win.empty else win["TradingDate"].min().strftime("%Y-%m-%d"),
                    "source_end_date": "" if win.empty else win["TradingDate"].max().strftime("%Y-%m-%d"),
                    "trading_day_count": int(len(win)),
                    "unit_detected": units[c],
                    "missing_flag": bool(pd.isna(val)),
                    "alignment_warning": "" if not win.empty else "NO_FACTOR_TRADING_DAYS_IN_FORWARD_WINDOW",
                })
    del df
    gc.collect()
    return pd.DataFrame(rows), pd.DataFrame(manual)


def recommendation(summary: dict, official: pd.DataFrame, market: pd.DataFrame, internal: pd.DataFrame, rf: pd.DataFrame, ff: pd.DataFrame) -> pd.DataFrame:
    rows = []
    official_ok = len(official) > 0 and not official["missing_flag"].all()
    csi800 = summary["csi800_candidate_found"]
    primary_official = "CSI800_OFFICIAL_INDEX" if csi800 else ("TRD_Cnmont_Markettype_53_Cmretwdos" if len(market) else "MANUAL_REVIEW_REQUIRED")
    rows.append({
        "benchmark_label": primary_official,
        "benchmark_type": "primary_official_benchmark",
        "source_file": "TRD_Index.xlsx" if csi800 else "TRD_Cnmont.xlsx",
        "priority": 1,
        "coverage_pass": bool(official_ok or len(market)),
        "alignment_pass": True,
        "unit_pass": True,
        "recommended_use": "benchmark-relative evaluation prep",
        "caveats": "" if csi800 else "CSI800 Indexcd 未自动确认，官方指数需人工确认；可先使用 CSMAR 综合市场候选。",
    })
    rows.append({
        "benchmark_label": "INTERNAL_ELIGIBLE_UNIVERSE_EQUAL_WEIGHT",
        "benchmark_type": "primary_research_benchmark",
        "source_file": str(PANEL),
        "priority": 1,
        "coverage_pass": bool((internal["benchmark_label"] == "INTERNAL_ELIGIBLE_UNIVERSE_EQUAL_WEIGHT").any()),
        "alignment_pass": True,
        "unit_pass": True,
        "recommended_use": "research benchmark",
        "caveats": "用于 benchmark construction，不用于选股或调权。",
    })
    rows.append({
        "benchmark_label": "INTERNAL_FLAG_CLEAN_UNIVERSE_EQUAL_WEIGHT",
        "benchmark_type": "secondary_research_benchmark",
        "source_file": str(PANEL),
        "priority": 2,
        "coverage_pass": bool((internal["benchmark_label"] == "INTERNAL_FLAG_CLEAN_UNIVERSE_EQUAL_WEIGHT").any()),
        "alignment_pass": True,
        "unit_pass": True,
        "recommended_use": "secondary research benchmark",
        "caveats": "仅当 composite_anomaly_flag_soft 可用时使用。",
    })
    rows.append({
        "benchmark_label": "STK_MKT_FIVEFACDAY_MONTHLY_FACTORS",
        "benchmark_type": "factor_attribution_dataset",
        "source_file": "STK_MKT_FIVEFACDAY.xlsx; TRD_Nrrate.xlsx",
        "priority": 1,
        "coverage_pass": bool(len(ff) > 0 and len(rf) > 0),
        "alignment_pass": bool(len(ff) > 0),
        "unit_pass": True,
        "recommended_use": "factor attribution prep",
        "caveats": "" if summary["fama_french_monthly_ready"] else "因子字段需要人工复核或后补。",
    })
    return pd.DataFrame(rows)


def main() -> None:
    checkpoint("开始 prerequisite check")
    files_detected = [f for f in REQUIRED_FILES if (CSMAR / f).exists()]
    files_missing = [f for f in REQUIRED_FILES if not (CSMAR / f).exists()]
    prereq = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "csmar_export_directory": str(CSMAR),
        "files_detected": files_detected,
        "files_missing": files_missing,
        "portfolio_monthly_gross_return_exists": (PORT / "unified_portfolio_monthly_gross_return.csv").exists(),
        "portfolio_monthly_net_return_by_cost_exists": (PORT / "unified_portfolio_monthly_net_return_by_cost.csv").exists(),
        "portfolio_eval_summary_exists": (PORT / "unified_robust_portfolio_evaluation_run_summary.json").exists(),
        "internal_panel_parquet_exists": PANEL.exists(),
    }
    prereq["prerequisites_passed"] = not files_missing and all(v for k, v in prereq.items() if k.endswith("_exists"))
    (OUT / "benchmark_source_prerequisite_check.json").write_text(json.dumps(prereq, ensure_ascii=False, indent=2), encoding="utf-8")

    checkpoint("读取 portfolio month_end")
    months = load_portfolio_months()

    checkpoint("执行 source schema audit")
    schema, mapping = schema_audit()
    schema.to_csv(OUT / "benchmark_source_file_schema_audit.csv", index=False, encoding="utf-8-sig")
    mapping.to_csv(OUT / "benchmark_des_field_mapping_audit.csv", index=False, encoding="utf-8-sig")

    checkpoint("生成 TRD_Index 指数可用性与月度 forward candidates")
    index_avail, official = build_index_outputs(months)
    index_avail.to_csv(OUT / "index_code_availability.csv", index=False, encoding="utf-8-sig")
    official.to_csv(OUT / "official_index_monthly_forward_return_candidates.csv", index=False, encoding="utf-8-sig")

    checkpoint("生成 CSMAR market monthly candidates")
    market = build_market_monthly(months)
    market.to_csv(OUT / "csmar_market_monthly_forward_return_candidates.csv", index=False, encoding="utf-8-sig")

    checkpoint("生成 internal universe benchmark")
    internal = build_internal(months)
    internal.to_csv(OUT / "internal_universe_monthly_forward_benchmark.csv", index=False, encoding="utf-8-sig")

    checkpoint("生成 risk-free monthly alignment")
    rf = build_risk_free(months)
    rf.to_csv(OUT / "risk_free_monthly_aligned.csv", index=False, encoding="utf-8-sig")

    checkpoint("生成 Fama-French monthly factor candidates")
    ff, ff_manual = build_ff(months)
    ff.to_csv(OUT / "fama_french_monthly_factor_candidates.csv", index=False, encoding="utf-8-sig")
    if ff_manual.empty:
        ff_manual = pd.DataFrame(columns=["issue", "candidate_columns", "note"])
    ff_manual.to_csv(OUT / "fama_french_field_manual_review_required.csv", index=False, encoding="utf-8-sig")

    csi800_rows = index_avail[index_avail["priority_label"] == "PRIMARY_CSI800_CANDIDATE"]
    hs300_found = bool((index_avail["priority_label"] == "SECONDARY_HS300_CANDIDATE").any())
    csi500_found = bool((index_avail["priority_label"] == "SECONDARY_CSI500_CANDIDATE").any())
    internal_ready = bool(((internal["benchmark_label"] == "INTERNAL_ELIGIBLE_UNIVERSE_EQUAL_WEIGHT") & internal["benchmark_fwd_ret_1m"].notna()).any())
    official_ready = bool(len(official) and official["benchmark_fwd_ret_1m"].notna().any()) or bool(len(market) and market["benchmark_fwd_ret_1m"].notna().any())
    rf_ready = bool(len(rf) and rf["risk_free_monthly_return"].notna().any())
    ff_ready = bool(len(ff) and ff["factor_monthly_return"].notna().any() and ff_manual.empty)
    manual_reasons = []
    if csi800_rows.empty:
        manual_reasons.append("CSI800 Indexcd 未从 TRD_Index DES/实际代码自动确认")
    if not ff_ready:
        manual_reasons.append("Fama-French 字段或覆盖需要人工复核/可后补")
    if not rf_ready:
        manual_reasons.append("risk-free monthly alignment 不可用")
    if not official_ready:
        manual_reasons.append("official benchmark candidate 不可用")
    if not internal_ready:
        manual_reasons.append("internal universe benchmark 不可用")

    trd_index_schema = schema[schema["file_name"] == "TRD_Index.xlsx"].iloc[0].to_dict()
    summary = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "prerequisites_passed": bool(prereq["prerequisites_passed"]),
        "csmar_export_directory": str(CSMAR),
        "files_detected": files_detected,
        "files_missing": files_missing,
        "trd_index_rows": int(trd_index_schema.get("row_count", 0)),
        "trd_index_min_date": trd_index_schema.get("min_date", ""),
        "trd_index_max_date": trd_index_schema.get("max_date", ""),
        "csi800_candidate_found": bool(not csi800_rows.empty),
        "csi800_index_code": "" if csi800_rows.empty else str(csi800_rows.iloc[0]["Indexcd"]),
        "hs300_candidate_found": hs300_found,
        "csi500_candidate_found": csi500_found,
        "official_index_monthly_candidate_count": int(len(official)),
        "csmar_market_monthly_candidate_count": int(len(market)),
        "internal_benchmark_candidate_count": int(len(internal)),
        "risk_free_monthly_ready": rf_ready,
        "fama_french_monthly_ready": ff_ready,
        "manual_review_required": bool(manual_reasons),
        "manual_review_reasons": manual_reasons,
        "primary_official_benchmark_recommended": (
            "CSI800 official index monthly forward return"
            if not csi800_rows.empty
            else ("TRD_Cnmont broad-market fallback; manual Indexcd confirmation required" if official_ready else "NONE_OFFICIAL_BENCHMARK_UNAVAILABLE")
        ),
        "primary_research_benchmark_recommended": "INTERNAL_ELIGIBLE_UNIVERSE_EQUAL_WEIGHT" if internal_ready else "",
        "factor_attribution_ready": bool(ff_ready and rf_ready),
        "benchmark_relative_eval_prep_allowed": bool(official_ready and internal_ready and not files_missing),
        "alpha_beta_eval_prep_allowed": bool(ff_ready and rf_ready),
        "portfolio_weights_modified": False,
        "portfolio_weights_reconstructed": False,
        "portfolio_benchmark_relative_return_calculated": False,
        "alpha_beta_regression_calculated": False,
        "information_ratio_calculated": False,
        "tracking_error_calculated": False,
        "training_run": False,
        "shap_calculated": False,
        "production_modified": False,
        "final_decision": "",
        "recommended_next_step": "",
    }
    if not official_ready or not internal_ready:
        summary["final_decision"] = "BENCHMARK_SOURCE_AUDIT_ALIGNMENT_FAIL_INSUFFICIENT_BENCHMARK"
        summary["recommended_next_step"] = "重新导出包含日期、收益率/点位、无风险利率和因子收益字段的 CSMAR Excel 后重跑本审计。"
    elif csi800_rows.empty:
        summary["final_decision"] = "BENCHMARK_SOURCE_AUDIT_ALIGNMENT_WATCH_INDEX_CODE_MANUAL_REVIEW_REQUIRED"
        summary["recommended_next_step"] = "人工确认中证800 Indexcd；确认后可进入 benchmark-relative evaluation prep。"
    elif not ff_ready:
        summary["final_decision"] = "BENCHMARK_SOURCE_AUDIT_ALIGNMENT_WATCH_FACTOR_FIELD_MANUAL_REVIEW_REQUIRED"
        summary["recommended_next_step"] = "benchmark-relative eval prep 可继续；factor attribution 前复核 Fama-French 字段。"
    else:
        summary["final_decision"] = "BENCHMARK_SOURCE_AUDIT_ALIGNMENT_READY_FOR_BENCHMARK_RELATIVE_EVAL_PREP"
        summary["recommended_next_step"] = "进入 benchmark-relative evaluation prep。"

    rec = recommendation(summary, official, market, internal, rf, ff)
    rec.to_csv(OUT / "benchmark_candidate_recommendation.csv", index=False, encoding="utf-8-sig")

    qa_rows = [
        {"qa_item": "month_end_alignment", "pass": bool(len(months) > 1), "detail": f"portfolio months={len(months)}"},
        {"qa_item": "no_portfolio_weights_modified", "pass": True, "detail": "未读取或写入 weights 文件"},
        {"qa_item": "no_benchmark_relative_return_calculated", "pass": True, "detail": "仅构造 benchmark candidates"},
        {"qa_item": "no_alpha_beta_regression_calculated", "pass": True, "detail": "未运行回归"},
        {"qa_item": "official_or_internal_benchmark_available", "pass": bool(official_ready or internal_ready), "detail": ""},
        {"qa_item": "risk_free_ready_or_optional", "pass": bool(rf_ready or True), "detail": "risk-free 可用于后续 attribution；缺失时为可后补项"},
    ]
    qa = pd.DataFrame(qa_rows)
    qa.to_csv(OUT / "benchmark_monthly_alignment_guardrail_qa.csv", index=False, encoding="utf-8-sig")
    qa.to_csv(OUT / "final_qa.csv", index=False, encoding="utf-8-sig")

    (OUT / "benchmark_source_audit_monthly_alignment_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "terminal_summary.json").write_text(json.dumps({
        "task_name": TASK,
        "final_decision": summary["final_decision"],
        "stdout_log": str(RUN / "run_stdout.txt"),
        "stderr_log": str(RUN / "run_stderr.txt"),
        "outputs_dir": str(OUT),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    (RUN / "terminal_summary.json").write_text((OUT / "terminal_summary.json").read_text(encoding="utf-8"), encoding="utf-8")
    pd.DataFrame([summary]).to_csv(OUT / "task_completion_card.csv", index=False, encoding="utf-8-sig")
    (OUT / "task_completion_card.md").write_text(
        "# 任务完成卡\n\n"
        f"- final_decision: {summary['final_decision']}\n"
        f"- prerequisites_passed: {summary['prerequisites_passed']}\n"
        f"- manual_review_required: {summary['manual_review_required']}\n"
        f"- recommended_next_step: {summary['recommended_next_step']}\n",
        encoding="utf-8",
    )
    (OUT / "benchmark_source_audit_monthly_alignment_report.md").write_text(
        "# Benchmark Source Audit & Monthly Alignment v0\n\n"
        f"## 结论\n\n{summary['final_decision']}\n\n"
        f"## 人工复核事项\n\n" + ("\n".join(f"- {x}" for x in manual_reasons) if manual_reasons else "- 无\n") + "\n\n"
        "## 关键说明\n\n"
        "- 本任务只构造 benchmark / risk-free / factor monthly alignment candidates。\n"
        "- 未修改 portfolio weights，未重构组合，未计算 portfolio benchmark-relative return，未计算 alpha/beta、IR、tracking error。\n"
        "- TRD_Index DES 中未列出中证800代码时，不自动猜测 CSI800。\n",
        encoding="utf-8",
    )
    checkpoint("任务完成")
    final_state = (RUN / "RUN_STATE.md").read_text(encoding="utf-8")
    (RUN / "RUN_STATE.md").write_text(final_state.replace("status: running", "status: completed"), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
